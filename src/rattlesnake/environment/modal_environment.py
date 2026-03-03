# -*- coding: utf-8 -*-
"""
This file defines a Modal Testing Environment where users can perform
hammer or shaker modal tests and export FRFs and other relevant data.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
from rattlesnake.utilities import GlobalCommands, VerboseMessageQueue, flush_queue, load_python_module
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.environment.abstract_environment import EnvironmentProcess, EnvironmentMetadata, EnvironmentInstructions
from rattlesnake.process.signal_generation_utilities import (
    BurstRandomSignalGenerator,
    ChirpSignalGenerator,
    PseudorandomSignalGenerator,
    RandomSignalGenerator,
    SineSignalGenerator,
    SquareSignalGenerator,
)
from rattlesnake.process.data_collector import (  # noqa # pylint: disable=wrong-import-position
    Acceptance,
    AcquisitionType,
    CollectorMetadata,
    DataCollectorCommands,
    TriggerSlope,
    Window,
    data_collector_process,
)
from rattlesnake.process.signal_generation import (  # noqa # pylint: disable=wrong-import-position
    SignalGenerationCommands,
    SignalGenerationMetadata,
    signal_generation_process,
)
from rattlesnake.process.spectral_processing import (  # noqa # pylint: disable=wrong-import-position
    AveragingTypes,
    Estimator,
    SpectralProcessingCommands,
    SpectralProcessingMetadata,
    spectral_processing_process,
)
from rattlesnake.user_interface.ui_utilities import UICommands
import multiprocessing as mp
import threading
import time
import openpyxl
from enum import Enum
from glob import glob
import netCDF4 as nc4
import numpy as np
from typing import List


CONTROL_TYPE = ControlTypes.MODAL
WAIT_TIME = 0.02


# region: Commands
class ModalCommands(Enum):
    """Valid commands for the modal environment"""

    ACCEPT_FRAME = 2
    RUN_CONTROL = 3
    CHECK_FOR_COMPLETE_SHUTDOWN = 4

    @property
    def label(self):
        """Used by UI as names for commands in profile table"""
        return self.name.replace("_", " ").title()

    @property
    def valid_data(self):
        return {
            ModalCommands.ACCEPT_FRAME: int,
            ModalCommands.RUN_CONTROL: type(None),
            ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN: type(None),
        }.get(self)


class ModalUICommands(Enum):
    SPECTRAL_UPDATE = 1
    FINISHED = 2


# region: Metadata
class ModalMetadata(EnvironmentMetadata):
    """Class for storing metadata for an environment.

    This class is used as a storage container for parameters used by an
    environment.  It is returned by the environment UI's
    ``collect_environment_definition_parameters`` function as well as its
    ``initialize_environment`` function.  Various parts of the controller and
    environment will query the class's data members for parameter values.
    """

    def __init__(
        self,
        environment_name: str,
        channel_list_bools: list,
        sample_rate: int,
        samples_per_frame: int,
        averaging_type: str,
        num_averages: int,
        averaging_coefficient: float,
        frf_technique: str,
        frf_window: str,
        overlap_percent: float,
        trigger_type: str,
        accept_type: str,
        wait_for_steady_state: float,
        trigger_channel: int,
        pretrigger_percent: float,
        trigger_slope_positive: bool,
        trigger_level_percent: float,
        hysteresis_level_percent: float,
        hysteresis_frame_percent: float,
        signal_generator_type: str,
        signal_generator_level: float,
        signal_generator_min_frequency: float,
        signal_generator_max_frequency: float,
        signal_generator_on_percent: float,
        acceptance_function,
        reference_channel_indices,
        response_channel_indices,
        output_channel_indices,
        output_oversample: int,
        exponential_window_value_at_frame_end: float,
    ):
        super().__init__(CONTROL_TYPE, environment_name, channel_list_bools, sample_rate)
        self.samples_per_frame = samples_per_frame
        self.averaging_type = averaging_type
        self.num_averages = num_averages
        self.averaging_coefficient = averaging_coefficient
        self.frf_technique = frf_technique
        self.frf_window = frf_window
        self.overlap = overlap_percent / 100
        self.trigger_type = trigger_type
        self.accept_type = accept_type
        self.wait_for_steady_state = wait_for_steady_state
        self.trigger_channel = trigger_channel
        self.pretrigger = pretrigger_percent / 100
        self.trigger_slope_positive = trigger_slope_positive
        self.trigger_level = trigger_level_percent / 100
        self.hysteresis_level = hysteresis_level_percent / 100
        self.hysteresis_length = hysteresis_frame_percent / 100
        self.signal_generator_type = signal_generator_type
        self.signal_generator_level = signal_generator_level
        self.signal_generator_min_frequency = signal_generator_min_frequency
        self.signal_generator_max_frequency = signal_generator_max_frequency
        self.signal_generator_on_fraction = signal_generator_on_percent / 100
        self.acceptance_function = acceptance_function
        self.reference_channel_indices = reference_channel_indices
        self.response_channel_indices = response_channel_indices
        self.output_channel_indices = output_channel_indices
        self.exponential_window_value_at_frame_end = exponential_window_value_at_frame_end
        # Set up signal generator
        self.output_oversample = output_oversample
        self.signal_generator = self.get_signal_generator()

    def validate(self, hardware_metadata):
        return super().validate(hardware_metadata)

    def get_signal_generator(self):
        """Gets a signal generator object that the modal environment will use to generate signals"""
        if self.signal_generator_type == "none":
            signal_generator = PseudorandomSignalGenerator(
                rms=0.0,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                low_frequency_cutoff=self.signal_generator_min_frequency,
                high_frequency_cutoff=self.signal_generator_max_frequency,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "random":
            signal_generator = RandomSignalGenerator(
                rms=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                low_frequency_cutoff=self.signal_generator_min_frequency,
                high_frequency_cutoff=self.signal_generator_max_frequency,
                cola_overlap=0.5,
                cola_window="hann",
                cola_exponent=0.5,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "pseudorandom":
            signal_generator = PseudorandomSignalGenerator(
                rms=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                low_frequency_cutoff=self.signal_generator_min_frequency,
                high_frequency_cutoff=self.signal_generator_max_frequency,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "burst":
            signal_generator = BurstRandomSignalGenerator(
                rms=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                low_frequency_cutoff=self.signal_generator_min_frequency,
                high_frequency_cutoff=self.signal_generator_max_frequency,
                on_fraction=self.signal_generator_on_fraction,
                ramp_fraction=0.05,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "chirp":
            signal_generator = ChirpSignalGenerator(
                level=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                low_frequency_cutoff=self.signal_generator_min_frequency,
                high_frequency_cutoff=self.signal_generator_max_frequency,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "square":
            signal_generator = SquareSignalGenerator(
                level=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                frequency=self.signal_generator_min_frequency,
                phase=0,
                on_fraction=self.signal_generator_on_fraction,
                output_oversample=self.output_oversample,
            )
        elif self.signal_generator_type == "sine":
            signal_generator = SineSignalGenerator(
                level=self.signal_generator_level,
                sample_rate=self.sample_rate,
                num_samples_per_frame=self.samples_per_frame,
                num_signals=len(self.output_channel_indices),
                frequency=self.signal_generator_min_frequency,
                phase=0,
                output_oversample=self.output_oversample,
            )
        else:
            raise ValueError(f"Invalid Signal Type {self.signal_generator_type}")
        return signal_generator

    @property
    def samples_per_acquire(self):
        """Property returning the samples per acquisition step given the overlap"""
        return int(self.samples_per_frame * (1 - self.overlap))

    @property
    def frame_time(self):
        """Property returning the time per measurement frame"""
        return self.samples_per_frame / self.sample_rate

    @property
    def nyquist_frequency(self):
        """Property returning half the sample rate"""
        return self.sample_rate / 2

    @property
    def fft_lines(self):
        """Property returning the frequency lines given the sampling parameters"""
        return self.samples_per_frame // 2 + 1

    @property
    def skip_frames(self):
        """Property returning the number of frames to skip while waiting for steady state"""
        return int(np.ceil(self.wait_for_steady_state * self.sample_rate / (self.samples_per_frame * (1 - self.overlap))))

    @property
    def frequency_spacing(self):
        """Property returning frequency line spacing given the sampling parameters"""
        return self.sample_rate / self.samples_per_frame

    def get_trigger_levels(self, channels):
        """Gets the trigger levels for a channel based on the channel table information

        Parameters
        ----------
        channels : list of Channel
            A list of channels in the environment

        Returns
        -------
        trigger_level_v
            The trigger level in volts
        trigger_level_eu
            The trigger level in engineering units defined in the channel table
        hysterisis_level_v
            The level that the signal must return below before another trigger can be accepted,
            in volts
        hysterisis_level_eu
            The level that the signal must return below before another trigger can be accepted,
            in engineering units defined in the channel table
        """
        channel = channels[self.trigger_channel]
        try:
            volt_range = float(channel.maximum_value)
            if volt_range == 0.0:
                volt_range = 10.0
        except (ValueError, TypeError):
            volt_range = 10.0
        try:
            mv_per_eu = float(channel.sensitivity)
            if mv_per_eu == 0.0:
                mv_per_eu = 1000.0
        except (ValueError, TypeError):
            mv_per_eu = 1000.0
        v_per_eu = mv_per_eu / 1000.0
        trigger_level_v = self.trigger_level * volt_range
        trigger_level_eu = trigger_level_v / v_per_eu
        hysterisis_level_v = self.hysteresis_level * volt_range
        hysterisis_level_eu = hysterisis_level_v / v_per_eu
        return (
            trigger_level_v,
            trigger_level_eu,
            hysterisis_level_v,
            hysterisis_level_eu,
        )

    @property
    def disabled_signals(self):
        """Returns a list of indices corresponding to output signals that have been disabled"""
        return [
            i
            for i, index in enumerate(self.output_channel_indices)
            if not (index in self.response_channel_indices or index in self.reference_channel_indices)
        ]

    @property
    def hysteresis_samples(self):
        """Property returning the number of samples that a signal must be below the hysterisis
        level"""
        return int(self.hysteresis_length * self.samples_per_frame)

    # region: Loading
    def store_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,  # pylint: disable=c-extension-no-member
    ):
        """Store parameters to a group in a netCDF streaming file.

        This function stores parameters from the environment into the netCDF
        file in a group with the environment's name as its name.  The function
        will receive a reference to the group within the dataset and should
        store the environment's parameters into that group in the form of
        attributes, dimensions, or variables.

        This function is the "write" counterpart to the retrieve_metadata
        function in the AbstractUI class, which will read parameters from
        the netCDF file to populate the parameters in the user interface.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A reference to the Group within the netCDF dataset where the
            environment's metadata is stored.

        """
        netcdf_group_handle.samples_per_frame = self.samples_per_frame
        netcdf_group_handle.averaging_type = self.averaging_type
        netcdf_group_handle.num_averages = self.num_averages
        netcdf_group_handle.averaging_coefficient = self.averaging_coefficient
        netcdf_group_handle.frf_technique = self.frf_technique
        netcdf_group_handle.frf_window = self.frf_window
        netcdf_group_handle.overlap = self.overlap
        netcdf_group_handle.trigger_type = self.trigger_type
        netcdf_group_handle.accept_type = self.accept_type
        netcdf_group_handle.wait_for_steady_state = self.wait_for_steady_state
        netcdf_group_handle.trigger_channel = self.trigger_channel
        netcdf_group_handle.pretrigger = self.pretrigger
        netcdf_group_handle.trigger_slope_positive = 1 if self.trigger_slope_positive else 0
        netcdf_group_handle.trigger_level = self.trigger_level
        netcdf_group_handle.hysteresis_level = self.hysteresis_level
        netcdf_group_handle.hysteresis_length = self.hysteresis_length
        netcdf_group_handle.signal_generator_type = self.signal_generator_type
        netcdf_group_handle.signal_generator_level = self.signal_generator_level
        netcdf_group_handle.signal_generator_min_frequency = self.signal_generator_min_frequency
        netcdf_group_handle.signal_generator_max_frequency = self.signal_generator_max_frequency
        netcdf_group_handle.signal_generator_on_fraction = self.signal_generator_on_fraction
        netcdf_group_handle.exponential_window_value_at_frame_end = self.exponential_window_value_at_frame_end
        netcdf_group_handle.acceptance_function = (
            self.acceptance_function[0] + ":" + self.acceptance_function[1] if self.acceptance_function is not None else "None"
        )
        # Reference channels
        netcdf_group_handle.createDimension("reference_channels", len(self.reference_channel_indices))
        var = netcdf_group_handle.createVariable("reference_channel_indices", "i4", ("reference_channels"))
        var[...] = self.reference_channel_indices
        # Response channels
        netcdf_group_handle.createDimension("response_channels", len(self.response_channel_indices))
        var = netcdf_group_handle.createVariable("response_channel_indices", "i4", ("response_channels"))
        var[...] = self.response_channel_indices

    @classmethod
    def retrieve_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Dataset,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf
        function in the ModalMetadata class, which will write parameters to
        the netCDF file to document the metadata.

        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.

        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``

        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset :
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.

        """
        samples_per_frame = netcdf_group_handle.samples_per_frame
        averaging_type = netcdf_group_handle.averaging_type
        num_averages = netcdf_group_handle.num_averages
        averaging_coefficient = netcdf_group_handle.averaging_coefficient
        frf_technique = netcdf_group_handle.frf_technique
        frf_window = netcdf_group_handle.frf_window
        overlap_percent = netcdf_group_handle.overlap * 100
        trigger_type = netcdf_group_handle.trigger_type
        accept_type = netcdf_group_handle.accept_type
        if accept_type == "Autoreject...":
            acceptance_function = netcdf_group_handle.acceptance_function.split(":")
        else:
            acceptance_function = None
        wait_for_steady_state = netcdf_group_handle.wait_for_steady_state
        trigger_channel = netcdf_group_handle.trigger_channel
        pretrigger_percent = netcdf_group_handle.pretrigger * 100
        trigger_slope_positive = netcdf_group_handle.trigger_slope_positive
        trigger_level_percent = netcdf_group_handle.trigger_level * 100
        hysteresis_level_percent = netcdf_group_handle.hysteresis_level * 100
        hysteresis_frame_percent = netcdf_group_handle.hysteresis_length * 100
        signal_generator_type = netcdf_group_handle.signal_generator_type
        signal_generator_level = netcdf_group_handle.signal_generator_level
        signal_generator_min_frequency = netcdf_group_handle.signal_generator_min_frequency
        signal_generator_max_frequency = netcdf_group_handle.signal_generator_max_frequency
        signal_generator_on_percent = netcdf_group_handle.signal_generator_level * 100
        reference_channel_indices = netcdf_group_handle.variables["reference_channel_indices"][...]
        response_channel_indices = netcdf_group_handle.variables["response_channel_indices"][...]
        environment_channel_list = [channel for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools) if channel_bool]
        output_channel_indices = [index for index, channel in enumerate(environment_channel_list) if channel.feedback_device is not None]
        exponential_window_value_at_frame_end = netcdf_group_handle.exponential_window_value_at_frame_end

        return cls(
            environment_name,
            channel_list_bools,
            hardware_metadata.sample_rate,
            samples_per_frame,
            averaging_type,
            num_averages,
            averaging_coefficient,
            frf_technique,
            frf_window,
            overlap_percent,
            trigger_type,
            accept_type,
            wait_for_steady_state,
            trigger_channel,
            pretrigger_percent,
            trigger_slope_positive,
            trigger_level_percent,
            hysteresis_level_percent,
            hysteresis_frame_percent,
            signal_generator_type,
            signal_generator_level,
            signal_generator_min_frequency,
            signal_generator_max_frequency,
            signal_generator_on_percent,
            acceptance_function,
            reference_channel_indices,
            response_channel_indices,
            output_channel_indices,
            hardware_metadata.output_oversample,
            exponential_window_value_at_frame_end,
        )

    def store_to_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        """Creates a template worksheet in an Excel workbook defining the
        environment.

        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the
        environment.

        This function is the "write" counterpart to the
        ``set_parameters_from_template`` function in the ``ModalUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.worksheet.worksheet.Worksheet :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet.cell(1, 1, "Control Type")
        worksheet.cell(1, 2, "Modal")
        worksheet.cell(2, 1, "Samples Per Frame:")
        worksheet.cell(2, 3, "# Number of Samples per Measurement Frame")
        worksheet.cell(3, 1, "Averaging Type:")
        worksheet.cell(3, 3, "# Averaging Type")
        worksheet.cell(4, 1, "Number of Averages:")
        worksheet.cell(4, 3, "# Number of Averages used when computing the FRF")
        worksheet.cell(5, 1, "Averaging Coefficient:")
        worksheet.cell(5, 3, "# Averaging Coefficient for Exponential Averaging")
        worksheet.cell(6, 1, "FRF Technique:")
        worksheet.cell(6, 3, "# FRF Technique")
        worksheet.cell(7, 1, "FRF Window:")
        worksheet.cell(7, 3, "# Window used to compute FRF")
        worksheet.cell(8, 1, "Exponential Window End Value:")
        worksheet.cell(
            8,
            3,
            "# Exponential Window Value at the end of the measurement frame (0.5 or 50%, not 50)",
        )
        worksheet.cell(9, 1, "FRF Overlap:")
        worksheet.cell(9, 3, "# Overlap for FRF calculations (0.5 or 50%, not 50)")
        worksheet.cell(10, 1, "Triggering Type:")
        worksheet.cell(10, 3, '# One of "Free Run", "First Frame", or "Every Frame"')
        worksheet.cell(11, 1, "Average Acceptance:")
        worksheet.cell(11, 3, '# One of "Accept All", "Manual", or "Autoreject"')
        worksheet.cell(12, 1, "Trigger Channel")
        worksheet.cell(12, 3, "# Channel number (1-based) to use for triggering")
        worksheet.cell(13, 1, "Pretrigger")
        worksheet.cell(13, 3, "# Amount of frame to use as pretrigger (0.5 or 50%, not 50)")
        worksheet.cell(14, 1, "Trigger Slope")
        worksheet.cell(14, 3, '# One of "Positive" or "Negative"')
        worksheet.cell(15, 1, "Trigger Level")
        worksheet.cell(
            15,
            3,
            "# Level to use to trigger the test as a fraction of the total range of the channel " "(0.5 or 50%, not 50)",
        )
        worksheet.cell(16, 1, "Hysteresis Level")
        worksheet.cell(
            16,
            3,
            "# Level that a channel must fall below before another trigger can be considered " "(0.5 or 50%, not 50)",
        )
        worksheet.cell(17, 1, "Hysteresis Frame Fraction")
        worksheet.cell(
            17,
            3,
            "# Fraction of the frame that a channel maintain hysteresis condition before another " "trigger can be considered (0.5 or 50%, not 50)",
        )
        worksheet.cell(18, 1, "Signal Generator Type")
        worksheet.cell(
            18,
            3,
            '# One of "None", "Random", "Burst Random", "Pseudorandom", "Chirp", "Square", or ' '"Sine"',
        )
        worksheet.cell(19, 1, "Signal Generator Level")
        worksheet.cell(
            19,
            3,
            "# RMS voltage level for random signals, Peak voltage level for chirp, sine, and " "square pulse",
        )
        worksheet.cell(20, 1, "Signal Generator Frequency 1")
        worksheet.cell(
            20,
            3,
            "# Minimum frequency for broadband signals or frequency for sine and square pulse",
        )
        worksheet.cell(21, 1, "Signal Generator Frequency 2")
        worksheet.cell(
            21,
            3,
            "# Maximum frequency for broadband signals.  Ignored for sine and square pulse",
        )
        worksheet.cell(22, 1, "Signal Generator On Fraction")
        worksheet.cell(
            22,
            3,
            "# Fraction of time that the burst or square wave is on (0.5 or 50%, not 50)",
        )
        worksheet.cell(23, 1, "Wait Time for Steady State")
        worksheet.cell(
            23,
            3,
            "# Time to wait after output starts to allow the system to reach steady state",
        )
        worksheet.cell(24, 1, "Autoaccept Script")
        worksheet.cell(24, 3, "# File in which an autoacceptance function is defined")
        worksheet.cell(25, 1, "Autoaccept Function")
        worksheet.cell(25, 3, "# Function name in which the autoacceptance function is defined")
        worksheet.cell(26, 1, "Reference Channels")
        worksheet.cell(26, 3, "# List of channels, one per cell on this row")
        worksheet.cell(27, 1, "Disabled Channels")
        worksheet.cell(27, 3, "# List of channels, one per cell on this row")

        if self.samples_per_frame is not None:
            worksheet.cell(2, 2, self.samples_per_frame)
        if self.averaging_type is not None:
            worksheet.cell(3, 2, self.averaging_type)
        if self.num_averages is not None:
            worksheet.cell(4, 2, self.num_averages)
        if self.averaging_coefficient is not None:
            worksheet.cell(5, 2, self.averaging_coefficient)
        if self.frf_technique is not None:
            worksheet.cell(6, 2, self.frf_technique)
        if self.frf_window is not None:
            worksheet.cell(7, 2, self.frf_window)
        if self.exponential_window_value_at_frame_end:
            worksheet.cell(8, 2, self.exponential_window_value_at_frame_end)
        if self.overlap is not None:
            worksheet.cell(9, 2, self.overlap)
        if self.trigger_type is not None:
            worksheet.cell(10, 2, self.trigger_type)
        if self.accept_type is not None:
            worksheet.cell(11, 2, self.accept_type)
        if self.trigger_channel is not None:
            worksheet.cell(12, 2, self.trigger_channel)
        if self.pretrigger is not None:
            worksheet.cell(13, 2, self.pretrigger)
        if self.trigger_slope_positive is not None:
            worksheet.cell(14, 2, self.trigger_slope_positive)
        if self.trigger_level is not None:
            worksheet.cell(15, 2, self.trigger_level)
        if self.hysteresis_level is not None:
            worksheet.cell(16, 2, self.hysteresis_level)
        if self.hysteresis_length is not None:
            worksheet.cell(17, 2, self.hysteresis_length)
        if self.signal_generator_type is not None:
            worksheet.cell(18, 2, self.signal_generator_type)
        if self.signal_generator_level is not None:
            worksheet.cell(19, 2, self.signal_generator_level)
        if self.signal_generator_min_frequency is not None:
            worksheet.cell(20, 2, self.signal_generator_min_frequency)
        if self.signal_generator_max_frequency is not None:
            worksheet.cell(21, 2, self.signal_generator_max_frequency)
        if self.signal_generator_on_fraction is not None:
            worksheet.cell(22, 2, self.signal_generator_on_fraction)
        if self.wait_for_steady_state is not None:
            worksheet.cell(23, 2, self.wait_for_steady_state)
        if self.acceptance_function is not None:
            worksheet.cell(24, 2, self.acceptance_function[0])
            worksheet.cell(25, 2, self.acceptance_function[1])
        if self.reference_channel_indices is not None:
            for idx, channel_ind in enumerate(self.reference_channel_indices):
                col_idx = idx + 2
                worksheet.cell(26, col_idx, channel_ind + 1)
        num_channels = sum(self.channel_list_bools)
        if self.response_channel_indices is not None:
            col_idx = 2
            for channel_ind in range(num_channels):
                if channel_ind not in self.response_channel_indices and channel_ind not in self.reference_channel_indices:
                    worksheet.cell(27, col_idx, channel_ind + 1)
                    col_idx += 1

    @classmethod
    def retrieve_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """
        Collects parameters for the user interface from the Excel template file

        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.

        This function is the "read" counterpart to the
        ``create_environment_template`` function in the ``ModalUI`` class,
        which writes a template file that can be filled out by a user.


        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        samples_per_frame = worksheet.cell(2, 2).value
        averaging_type = worksheet.cell(3, 2).value
        num_averages = worksheet.cell(4, 2).value
        averaging_coefficient = worksheet.cell(5, 2).value
        frf_technique = worksheet.cell(6, 2).value
        frf_window = worksheet.cell(7, 2).value
        overlap = worksheet.cell(9, 2).value
        overlap_percent = overlap * 100 if overlap else 0
        trigger_type = worksheet.cell(10, 2).value
        exponential_window_value_at_frame_end = worksheet.cell(8, 2).value
        accept_type = worksheet.cell(11, 2).value
        if accept_type == "Autoreject":
            acceptance_function = [
                worksheet.cell(24, 2).value,
                worksheet.cell(25, 2).value,
            ]
        else:
            acceptance_function = None
        trigger_channel = worksheet.cell(12, 2).value
        pretrigger = worksheet.cell(13, 2).value
        pretrigger_percent = pretrigger * 100 if pretrigger else 0
        trigger_slope_positive = worksheet.cell(14, 2).value
        trigger_level = worksheet.cell(15, 2).value
        trigger_level_percent = trigger_level * 100 if trigger_level else 0
        hysteresis_level = worksheet.cell(16, 2).value
        hysteresis_level_percent = hysteresis_level * 100 if hysteresis_level else 0
        hysteresis_frame = worksheet.cell(17, 2).value
        hysteresis_frame_percent = hysteresis_frame * 100 if hysteresis_frame else 0
        signal_generator_type = worksheet.cell(18, 2).value
        signal_generator_level = worksheet.cell(19, 2).value
        signal_generator_min_frequency = worksheet.cell(20, 2).value
        signal_generator_max_frequency = worksheet.cell(21, 2).value
        signal_generator_on_fraction = worksheet.cell(22, 2).value
        signal_generator_on_percent = signal_generator_on_fraction * 100 if signal_generator_on_fraction else 0
        wait_for_steady_state = worksheet.cell(23, 2).value
        max_row = 0
        column_index = 2
        reference_channel_indices = []
        response_channel_indices = []
        output_channel_indices = []
        while True:
            value = worksheet.cell(26, column_index).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            reference_channel_indices.append(int(value) - 1)
            column_index += 1
        max_row = len(channel_list_bools)
        for i in range(max_row):
            response_channel_indices.append(int(i))
        column_index = 2
        while True:
            value = worksheet.cell(27, column_index).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            response_channel_indices.remove(value - 1)
            column_index += 1
        environment_channel_list = [channel for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools) if channel_bool]
        output_channel_indices = [index for index, channel in enumerate(environment_channel_list) if channel.feedback_device is not None]

        return cls(
            environment_name,
            channel_list_bools,
            hardware_metadata.sample_rate,
            samples_per_frame,
            averaging_type,
            num_averages,
            averaging_coefficient,
            frf_technique,
            frf_window,
            overlap_percent,
            trigger_type,
            accept_type,
            wait_for_steady_state,
            trigger_channel,
            pretrigger_percent,
            trigger_slope_positive,
            trigger_level_percent,
            hysteresis_level_percent,
            hysteresis_frame_percent,
            signal_generator_type,
            signal_generator_level,
            signal_generator_min_frequency,
            signal_generator_max_frequency,
            signal_generator_on_percent,
            acceptance_function,
            reference_channel_indices,
            response_channel_indices,
            output_channel_indices,
            hardware_metadata.output_oversample,
            exponential_window_value_at_frame_end,
        )

    def generate_signal(self):
        """Generates a single frame of data"""
        if self.signal_generator is None:
            return np.zeros(
                (
                    len(self.output_channel_indices),
                    self.samples_per_frame * self.output_oversample,
                )
            )
        else:
            return self.signal_generator.generate_frame()[0]


class ModalInstructions(EnvironmentInstructions):
    def __init__(self, environment_name):
        super().__init__(CONTROL_TYPE, environment_name)


class ModalQueues:
    """A set of queues used by the modal environment"""

    def __init__(
        self,
        environment_name: str,
        environment_command_queue: VerboseMessageQueue,
        gui_update_queue: mp.Queue,
        controller_communication_queue: VerboseMessageQueue,
        data_in_queue: mp.Queue,
        data_out_queue: mp.Queue,
        log_file_queue: mp.Queue,
    ):
        """
        Creates a namespace to store all the queues used by the Modal Environment

        Parameters
        ----------
        environment_command_queue : VerboseMessageQueue
            Queue from which the environment will receive instructions.
        gui_update_queue : mp.queues.Queue
            Queue to which the environment will put GUI updates.
        controller_communication_queue : VerboseMessageQueue
            Queue to which the environment will put global contorller instructions.
        data_in_queue : mp.queues.Queue
            Queue from which the environment will receive data from acquisition.
        data_out_queue : mp.queues.Queue
            Queue to which the environment will write data for output.
        log_file_queue : VerboseMessageQueue
            Queue to which the environment will write log file messages.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.log_file_queue = log_file_queue
        self.data_for_spectral_computation_queue = mp.Queue()
        self.updated_spectral_quantities_queue = mp.Queue()
        self.signal_generation_update_queue = mp.Queue()
        self.spectral_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Spectral Computation Command Queue")
        self.collector_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Data Collector Command Queue")
        self.signal_generation_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Signal Generation Command Queue")


# region: Environment
class ModalEnvironment(EnvironmentProcess):
    """Modal Environment class defining the interface with the controller"""

    def __init__(
        self,
        environment_name: str,
        queue_name: str,
        queue_container: ModalQueues,
        acquisition_active_event: mp.synchronize.Event,
        output_active_event: mp.synchronize.Event,
        active_event: mp.synchronize.Event,
        ready_event: mp.synchronize.Event,
    ):
        super().__init__(
            environment_name,
            queue_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
        )
        self.queue_container = queue_container
        self.hardware_metadata = None
        self.metadata = None
        self.frame_number = 0
        self.siggen_shutdown_achieved = False
        self.collector_shutdown_achieved = False
        self.spectral_shutdown_achieved = False

        # Map commands
        self.map_command(ModalCommands.ACCEPT_FRAME, self.accept_frame)
        self.map_command(GlobalCommands.START_ENVIRONMENT, self.start_environment)
        self.map_command(ModalCommands.RUN_CONTROL, self.run_control)
        self.map_command(ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN, self.check_for_shutdown)
        self.map_command(SignalGenerationCommands.SHUTDOWN_ACHIEVED, self.siggen_shutdown_achieved_fn)
        self.map_command(DataCollectorCommands.SHUTDOWN_ACHIEVED, self.collector_shutdown_achieved_fn)
        self.map_command(
            SpectralProcessingCommands.SHUTDOWN_ACHIEVED,
            self.spectral_shutdown_achieved_fn,
        )

    # region: Initialize
    def initialize_hardware(self, hardware_metadata: HardwareMetadata):
        """Initialize the data acquisition parameters in the environment.

        The environment will receive the global data acquisition parameters from
        the controller, and must set itself up accordingly.

        Parameters
        ----------
        hardware_metadata : DataAcquisitionParameters :
            A container containing data acquisition parameters, including
            channels active in the environment as well as sampling parameters.
        """
        self.hardware_metadata = hardware_metadata
        self.set_ready()

    def initialize_environment(self, metadata: ModalMetadata):
        """
        Initialize the environment parameters specific to this environment

        The environment will recieve parameters defining itself from the
        user interface and must set itself up accordingly.

        Parameters
        ----------
        environment_parameters : ModalMetadata
            A container containing the parameters defining the environment

        """
        self.metadata = metadata

        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.INITIALIZE_COLLECTOR,
                self.get_data_collector_metadata(),
            ),
        )
        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_PARAMETERS,
                self.get_signal_generation_metadata(),
            ),
        )
        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (
                SpectralProcessingCommands.INITIALIZE_PARAMETERS,
                self.get_spectral_processing_metadata(),
            ),
        )

        self.set_ready()

    def get_data_collector_metadata(self) -> CollectorMetadata:
        """Collects metadata used to define the data collector"""
        num_channels = len(self.hardware_metadata.channel_list)
        reference_channel_indices = self.metadata.reference_channel_indices
        response_channel_indices = self.metadata.response_channel_indices
        if self.metadata.trigger_type == "Free Run":
            acquisition_type = AcquisitionType.FREE_RUN
        elif self.metadata.trigger_type == "First Frame":
            acquisition_type = AcquisitionType.TRIGGER_FIRST_FRAME
        elif self.metadata.trigger_type == "Every Frame":
            acquisition_type = AcquisitionType.TRIGGER_EVERY_FRAME
        else:
            raise ValueError(f"Invalid Acquisition Type: {self.metadata.trigger_type}")
        if self.metadata.accept_type == "Accept All":
            acceptance = Acceptance.AUTOMATIC
            acceptance_function = None
        elif self.metadata.accept_type == "Manual":
            acceptance = Acceptance.MANUAL
            acceptance_function = None
        elif self.metadata.accept_type == "Autoreject...":
            acceptance = Acceptance.AUTOMATIC
            acceptance_function = self.metadata.acceptance_function
        else:
            raise ValueError(f"Invalid Acceptance Type: {self.metadata.accept_type}")
        overlap_fraction = self.metadata.overlap
        trigger_channel_index = self.metadata.trigger_channel
        trigger_slope = TriggerSlope.POSITIVE if self.metadata.trigger_slope_positive else TriggerSlope.NEGATIVE
        (_, trigger_level, _, trigger_hysteresis) = self.metadata.get_trigger_levels(self.hardware_metadata.channel_list)
        trigger_hysteresis_samples = self.metadata.hysteresis_samples
        pretrigger_fraction = self.metadata.pretrigger
        frame_size = self.metadata.samples_per_frame
        if self.metadata.frf_window == "hann":
            window = Window.HANN
        elif self.metadata.frf_window == "rectangle":
            window = Window.RECTANGLE
        elif self.metadata.frf_window == "exponential":
            window = Window.EXPONENTIAL
        else:
            raise ValueError(f"Invalid Window Type: {self.metadata.frf_window}")
        window_parameter = -(frame_size) / np.log(self.metadata.exponential_window_value_at_frame_end)
        return CollectorMetadata(
            num_channels,
            response_channel_indices,
            reference_channel_indices,
            acquisition_type,
            acceptance,
            acceptance_function,
            overlap_fraction,
            trigger_channel_index,
            trigger_slope,
            trigger_level,
            trigger_hysteresis,
            trigger_hysteresis_samples,
            pretrigger_fraction,
            frame_size,
            window,
            response_transformation_matrix=None,
            reference_transformation_matrix=None,
            window_parameter_2=window_parameter,
        )

    def get_spectral_processing_metadata(self) -> SpectralProcessingMetadata:
        """Collects metadata to define the spectral processing"""
        averaging_type = AveragingTypes.LINEAR if self.metadata.averaging_type == "Linear" else AveragingTypes.EXPONENTIAL
        averages = self.metadata.num_averages
        exponential_averaging_coefficient = self.metadata.averaging_coefficient
        if self.metadata.frf_technique == "H1":
            frf_estimator = Estimator.H1
        elif self.metadata.frf_technique == "H2":
            frf_estimator = Estimator.H2
        elif self.metadata.frf_technique == "H3":
            frf_estimator = Estimator.H3
        elif self.metadata.frf_technique == "Hv":
            frf_estimator = Estimator.HV
        else:
            raise ValueError(f"Invalid FRF Estimator {self.metadata.frf_technique}. " "How did you get here?")
        num_response_channels = len(self.metadata.response_channel_indices)
        num_reference_channels = len(self.metadata.reference_channel_indices)
        frequency_spacing = self.metadata.frequency_spacing
        sample_rate = self.metadata.sample_rate
        num_frequency_lines = self.metadata.fft_lines
        return SpectralProcessingMetadata(
            averaging_type,
            averages,
            exponential_averaging_coefficient,
            frf_estimator,
            num_response_channels,
            num_reference_channels,
            frequency_spacing,
            sample_rate,
            num_frequency_lines,
            compute_cpsd=False,
            compute_apsd=True,
        )

    def get_signal_generation_metadata(self) -> SignalGenerationMetadata:
        """Collects metadata to define the signal generator"""
        return SignalGenerationMetadata(
            samples_per_write=self.hardware_metadata.samples_per_write,
            level_ramp_samples=1,
            output_transformation_matrix=None,
            disabled_signals=self.metadata.disabled_signals,
        )

    def get_signal_generator(self):
        """Gets the signal generator object used to generate signals for the environment"""
        return self.metadata.get_signal_generator()

    # region: Control Loop
    def start_environment(self, data):  # pylint: disable=unused-argument
        """Starts the environment

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        self.log("Starting Modal")
        self.siggen_shutdown_achieved = False
        self.collector_shutdown_achieved = False
        self.spectral_shutdown_achieved = False

        # Set up the collector
        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.FORCE_INITIALIZE_COLLECTOR,
                self.get_data_collector_metadata(),
            ),
        )

        self.queue_container.collector_command_queue.put(
            self.environment_name,
            (
                DataCollectorCommands.SET_TEST_LEVEL,
                (self.metadata.skip_frames, 1),
            ),
        )
        time.sleep(0.01)

        # Set up the signal generation
        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_PARAMETERS,
                self.get_signal_generation_metadata(),
            ),
        )

        self.queue_container.signal_generation_command_queue.put(
            self.environment_name,
            (
                SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
                self.get_signal_generator(),
            ),
        )

        self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.MUTE, None))

        self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.ADJUST_TEST_LEVEL, 1.0))

        # Tell the collector to start acquiring data
        self.queue_container.collector_command_queue.put(self.environment_name, (DataCollectorCommands.ACQUIRE, None))

        # Tell the signal generation to start generating signals
        self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.GENERATE_SIGNALS, None))

        # Set up the spectral processing
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (
                SpectralProcessingCommands.INITIALIZE_PARAMETERS,
                self.get_spectral_processing_metadata(),
            ),
        )

        # Tell the spectral analysis to clear and start acquiring
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.CLEAR_SPECTRAL_PROCESSING, None),
        )

        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.RUN_SPECTRAL_PROCESSING, None),
        )

        self.queue_container.environment_command_queue.put(self.environment_name, (ModalCommands.RUN_CONTROL, None))

        self.set_active()

    def run_control(self, data):  # pylint: disable=unused-argument
        """Runs the environment

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        # Pull data off the spectral queue
        spectral_data = flush_queue(self.queue_container.updated_spectral_quantities_queue, timeout=WAIT_TIME)
        if len(spectral_data) > 0:
            self.log("Received Data")
            (
                frames,
                frequencies,
                frf,
                coherence,
                response_cpsd,
                reference_cpsd,
                condition,
            ) = spectral_data[-1]
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (
                        ModalUICommands.SPECTRAL_UPDATE,
                        (
                            frames,
                            self.metadata.num_averages,
                            frequencies,
                            frf,
                            coherence,
                            response_cpsd,
                            reference_cpsd,
                            condition,
                        ),
                    ),
                )
            )
        else:
            time.sleep(WAIT_TIME)
        self.queue_container.environment_command_queue.put(self.environment_name, (ModalCommands.RUN_CONTROL, None))

    # region: Shutdown
    def siggen_shutdown_achieved_fn(self, data):  # pylint: disable=unused-argument
        """Sets the signal generation shutdown flag to True

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        self.siggen_shutdown_achieved = True

    def collector_shutdown_achieved_fn(self, data):  # pylint: disable=unused-argument
        """Sets the collector shutdown flag to True

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        self.collector_shutdown_achieved = True

    def spectral_shutdown_achieved_fn(self, data):  # pylint: disable=unused-argument
        """Sets the spectral processing shutdown flag to True

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        self.spectral_shutdown_achieved = True

    def check_for_shutdown(self, data):  # pylint: disable=unused-argument
        """Checks if all environment subprocesses have shut down successfully.

        Parameters
        ----------
        data : NoneType
            Requred by the message/data data-passing strategy in Rattlesnake, but not needed by
            this method
        """
        if self.siggen_shutdown_achieved and self.collector_shutdown_achieved and self.spectral_shutdown_achieved:
            self.log("Shutdown Achieved")
            self.clear_active()
            # self.gui_update_queue.put((self.environment_name, (UICommands.ENVIRONMENT_ENDED, None)))
        else:
            # Recheck some time later
            time.sleep(1)
            self.environment_command_queue.put(self.environment_name, (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None))

    def accept_frame(self, data):
        """Accepts or rejects the previous measurement frame"""
        self.queue_container.collector_command_queue.put(self.environment_name, (DataCollectorCommands.ACCEPT, data))

    def stop_environment(self, data):
        """Stop the environment gracefully

        This function defines the operations to shut down the environment
        gracefully so there is no hard stop that might damage test equipment
        or parts.

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        """
        self.log("Stopping Control")
        flush_queue(self.queue_container.environment_command_queue)
        self.queue_container.collector_command_queue.put(self.environment_name, (DataCollectorCommands.SET_TEST_LEVEL, (1000, 1)))
        self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.START_SHUTDOWN, None))
        self.queue_container.spectral_command_queue.put(
            self.environment_name,
            (SpectralProcessingCommands.STOP_SPECTRAL_PROCESSING, None),
        )
        self.queue_container.environment_command_queue.put(self.environment_name, (ModalCommands.CHECK_FOR_COMPLETE_SHUTDOWN, None))

    def quit(self, data):
        """Returns True to stop the ``run`` while loop and exit the process

        Parameters
        ----------
        data : Ignored
            This parameter is not used by the function but must be present
            due to the calling signature of functions called through the
            ``command_map``

        Returns
        -------
        True :
            This function returns True to signal to the ``run`` while loop
            that it is time to close down the environment.

        """
        for queue in [
            self.queue_container.spectral_command_queue,
            self.queue_container.signal_generation_command_queue,
            self.queue_container.collector_command_queue,
        ]:
            queue.put(self.environment_name, (GlobalCommands.QUIT, None))
        return True


# region: Process
def modal_process(
    environment_name: str,
    queue_name: str,
    input_queue: VerboseMessageQueue,
    gui_update_queue: mp.Queue,
    controller_command_queue: VerboseMessageQueue,
    log_file_queue: mp.Queue,
    data_in_queue: mp.Queue,
    data_out_queue: mp.Queue,
    acquisition_active_event: mp.synchronize.Event,
    output_active_event: mp.synchronize.Event,
    active_event: mp.synchronize.Event,
    ready_event: mp.synchronize.Event,
    shutdown_event: mp.synchronize.Event,
    threaded: bool,
):
    """Modal environment process function called by multiprocessing

    This function defines the Modal Environment process that
    gets run by the multiprocessing module when it creates a new process.  It
    creates a ModalEnvironment object and runs it.

    Parameters
    ----------
    environment_name : str :
        Name of the environment
    input_queue : VerboseMessageQueue :
        Queue containing instructions for the environment
    gui_update_queue : Queue :
        Queue where GUI updates are put
    controller_communication_queue : Queue :
        Queue for global communications with the controller
    log_file_queue : Queue :
        Queue for writing log file messages
    data_in_queue : Queue :
        Queue from which data will be read by the environment
    data_out_queue : Queue :
        Queue to which data will be written that will be output by the hardware.

    """
    if threaded:
        new_process = threading.Thread  # worker threads
    else:
        new_process = mp.Process  # worker processes

    queue_container = ModalQueues(
        environment_name,
        input_queue,
        gui_update_queue,
        controller_command_queue,
        data_in_queue,
        data_out_queue,
        log_file_queue,
    )

    spectral_proc = new_process(
        target=spectral_processing_process,
        args=(
            environment_name,
            queue_container.spectral_command_queue,
            queue_container.data_for_spectral_computation_queue,
            queue_container.updated_spectral_quantities_queue,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.log_file_queue,
        ),
    )
    spectral_proc.start()

    siggen_proc = new_process(
        target=signal_generation_process,
        args=(
            environment_name,
            queue_container.signal_generation_command_queue,
            queue_container.signal_generation_update_queue,
            queue_container.data_out_queue,
            queue_container.environment_command_queue,
            queue_container.log_file_queue,
            queue_container.gui_update_queue,
        ),
    )
    siggen_proc.start()

    collection_proc = new_process(
        target=data_collector_process,
        args=(
            environment_name,
            queue_container.collector_command_queue,
            queue_container.data_in_queue,
            [queue_container.data_for_spectral_computation_queue],
            queue_container.environment_command_queue,
            queue_container.log_file_queue,
            queue_container.gui_update_queue,
        ),
    )
    collection_proc.start()

    process_class = ModalEnvironment(
        environment_name, queue_name, queue_container, acquisition_active_event, output_active_event, active_event, ready_event
    )
    process_class.run(shutdown_event)

    # Rejoin all the processes
    process_class.log("Joining Subprocesses")
    process_class.log("Joining Spectral Computation")
    spectral_proc.join()
    process_class.log("Joining Signal Generation")
    siggen_proc.join()
    process_class.log("Joining Data Collection")
    collection_proc.join()
