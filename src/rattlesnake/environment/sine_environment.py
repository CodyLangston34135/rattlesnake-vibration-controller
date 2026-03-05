# -*- coding: utf-8 -*-
"""
This file defines a sine environment that utilizes system
identification.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
from rattlesnake.utilities import (
    VerboseMessageQueue,
    GlobalCommands,
    flush_queue,
    scale2db,
    wrap,
    db2scale,
    read_transformation_matrix_from_worksheet,
)
from rattlesnake.user_interface.ui_utilities import UICommands
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.abstract_environment import EnvironmentInstructions, EnvironmentCommands
from rattlesnake.environment.abstract_sysid_environment import SysIdEnvironmentProcess, SysIdEnvironmentMetadata
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.environment.sine_utilities import (
    DefaultSineControlLaw,
    SineSpecification,
    digital_tracking_filter_generator,
    sine_sweep,
    vold_kalman_filter_generator,
    load_specification,
)
from rattlesnake.process.sysid_data_analysis import SysIdMetadata, sysid_data_analysis_process
from rattlesnake.process.data_collector import data_collector_process
from rattlesnake.process.signal_generation_utilities import ContinuousTransientSignalGenerator
from rattlesnake.process.signal_generation import SignalGenerationCommands, SignalGenerationMetadata, signal_generation_process
from rattlesnake.process.spectral_processing import spectral_processing_process
import openpyxl
import importlib
import multiprocessing as mp
import threading
import os
import time
import traceback
import netCDF4 as nc4
import numpy as np
import scipy.signal as sig
from enum import Enum
from typing import List


# region: Global Variables
MAXIMUM_SAMPLES_TO_PLOT = 1000000
CONTROL_TYPE = ControlTypes.SINE

DEBUG = False

if DEBUG:
    from glob import glob

    FILE_OUTPUT = "debug_data/sine_control_{:}.npz"


# region: Commands
class SineCommands(EnvironmentCommands):
    """Enumeration containing sine commands"""

    START_CONTROL = 0
    STOP_CONTROL = 1
    SAVE_CONTROL_DATA = 2
    PERFORM_CONTROL_PREDICTION = 3
    SEND_EXCITATION_PREDICTION = 4
    SEND_RESPONSE_PREDICTION = 5
    SET_TEST_LEVEL = 6

    VALID_PROFILE_COMMANDS = (SET_TEST_LEVEL, SAVE_CONTROL_DATA)
    VALID_DATA = {
        START_CONTROL: type(None),
        STOP_CONTROL: type(None),
        SAVE_CONTROL_DATA: str,
        PERFORM_CONTROL_PREDICTION: type(None),
        SEND_EXCITATION_PREDICTION: type(None),
        SEND_RESPONSE_PREDICTION: type(None),
        SET_TEST_LEVEL: int,
    }


class SineUICommands(Enum):
    REQUEST_PREDICTION_PLOT_CHOICES = 0
    EXCITATION_PRECDICTION = 1
    RESPONSE_PREDICTION = 2
    RESPONSE_ERROR_MATRIX = 3
    EXCITATION_VOLTAGE_LIST = 4
    SPECIFICATION_FOR_PLOTTING = 5
    TIME_DATA = 6
    CONTROL_DATA = 7
    ENABLE_CONTROL = 8


# region: Queues
class SineQueues:
    """A container class for the queues that this environment will manage."""

    def __init__(
        self,
        environment_name: str,
        environment_command_queue: VerboseMessageQueue,
        gui_update_queue: mp.Queue,
        controller_communication_queue: VerboseMessageQueue,
        data_in_queue: mp.Queue,
        data_out_queue: mp.Queue,
        log_file_queue: VerboseMessageQueue,
    ):
        """A container class for the queues that sine vibration will manage.

        The environment uses many queues to pass data between the various
        pieces.  This class organizes those queues into one common namespace.

        Parameters
        ----------
        environment_name : str
            Name of the environment
        environment_command_queue : VerboseMessageQueue
            Queue that is read by the environment for environment commands
        gui_update_queue : mp.queues.Queue
            Queue where various subtasks put instructions for updating the
            widgets in the user interface
        controller_communication_queue : VerboseMessageQueue
            Queue that is read by the controller for global controller commands
        data_in_queue : mp.queues.Queue
            Multiprocessing queue that connects the acquisition subtask to the
            environment subtask.  Each environment will retrieve acquired data
            from this queue.
        data_out_queue : mp.queues.Queue
            Multiprocessing queue that connects the output subtask to the
            environment subtask.  Each environment will put data that it wants
            the controller to generate in this queue.
        log_file_queue : VerboseMessageQueue
            Queue for putting logging messages that will be read by the logging
            subtask and written to a file.
        """
        self.environment_command_queue = environment_command_queue
        self.gui_update_queue = gui_update_queue
        self.data_analysis_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Data Analysis Command Queue")
        self.signal_generation_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Signal Generation Command Queue")
        self.spectral_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Spectral Computation Command Queue")
        self.collector_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), environment_name + " Data Collector Command Queue")
        self.controller_communication_queue = controller_communication_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.data_for_spectral_computation_queue = mp.Queue()
        self.updated_spectral_quantities_queue = mp.Queue()
        self.time_history_to_generate_queue = mp.Queue()
        self.log_file_queue = log_file_queue


# region: Metadata
class SineMetadata(SysIdEnvironmentMetadata):
    """Metadata describing the Sine environment"""

    def __init__(
        self,
        *,
        environment_name: str,
        channel_list_bools: list,
        sample_rate: int,
        samples_per_frame,
        number_of_channels,
        specifications,
        ramp_time,
        buffer_blocks,
        control_convergence,
        update_drives_after_environment,
        phase_fit,
        allow_automatic_aborts,
        tracking_filter_type,
        tracking_filter_cutoff,
        tracking_filter_order,
        vk_filter_order,
        vk_filter_bandwidth,
        vk_filter_blocksize,
        vk_filter_overlap,
        control_python_script,
        control_python_class,
        control_python_parameters,
        control_channel_indices,
        output_channel_indices,
        response_transformation_matrix,
        output_transformation_matrix,
        sysid_metadata=None,
    ):
        """Creates a Metadata object defining a Sine environment

        Parameters
        ----------
        sample_rate : float
            The sample rate in Hz
        samples_per_frame : int
            The number of samples per acquisition block
        number_of_channels : int
            The number of channels in the environment
        specifications : array of SineSpecification
            One SineSpecification object for each tone in the environment
        ramp_time : float
            The time to ramp to the initial level and ramp down from the final level
        buffer_blocks : int
            The number of blocks of data to use in various buffered activities,
            such as filtering and overlapping and adding data
        control_convergence : float
            A value between 0 and 1 that controls the proportional correction
            factor in the on-line control.
        update_drives_after_environment : bool
            If True, the control class will call the finalize_control method to
            update the preshaped drives based on the results of the previous test
        phase_fit : bool
            If True, the achieved phases will be best-fit to the specification
            phase.  This should handle time delays between acquisition and output
        allow_automatic_aborts : bool
            If True, the test will shut down if an abort level is reached
        tracking_filter_type : int
            0 for digital tracking filter, 1 for vold kalman filter
        tracking_filter_cutoff : float
            Filter cutoff for the digital tracking filter.
        tracking_filter_order : int
            Filter order for the digital tracking filter.
        vk_filter_order : int
            Order of the Vold-Kalman filter.
        vk_filter_bandwidth : float
            Bandwidth of the Vold-Kalman filter
        vk_filter_blocksize : int
            Number of samples in each Vold-Kalman filter segment
        vk_filter_overlap : float
            Overlap fraction between 0 and 0.5 for the Vold-Kalman filter
        control_python_script : str
            Path to a Python script containing an alternative Sine control law class
        control_python_class : str
            Name of the sine control law class
        control_python_parameters : str
            Extra parameters passed to the sine control law
        control_channel_indices : array of int
            Indices into the channels specifying the channels used as control channels
        output_channel_indices : array of int
            Indices into the channels specifying the channels used as drive channels
        response_transformation_matrix : ndarray
            A 2D np.ndarray consisting of a transformation matrix applied to the
            control channels
        output_transformation_matrix : _type_
            A 2D np.ndarray consisting of a transformation matrix applied to the
            drive channels
        """
        super().__init__(CONTROL_TYPE, environment_name, channel_list_bools, sample_rate, sysid_metadata)
        self.samples_per_frame = samples_per_frame
        self.number_of_channels = number_of_channels
        self.specifications = specifications
        self.ramp_time = ramp_time
        self.buffer_blocks = buffer_blocks
        self.control_convergence = control_convergence
        self.update_drives_after_environment = update_drives_after_environment
        self.phase_fit = phase_fit
        self.allow_automatic_aborts = allow_automatic_aborts
        self.tracking_filter_type = tracking_filter_type
        self.tracking_filter_cutoff = tracking_filter_cutoff
        self.tracking_filter_order = tracking_filter_order
        self.vk_filter_order = vk_filter_order
        self.vk_filter_bandwidth = vk_filter_bandwidth
        self.vk_filter_blocksize = vk_filter_blocksize
        self.vk_filter_overlap = vk_filter_overlap
        self.control_python_script = control_python_script
        self.control_python_class = control_python_class
        self.control_python_parameters = control_python_parameters
        self.control_channel_indices = control_channel_indices
        self.output_channel_indices = output_channel_indices
        self.response_transformation_matrix = response_transformation_matrix
        self.reference_transformation_matrix = output_transformation_matrix

    @property
    def sample_rate(self):
        """Sample rate of the data acquisition system"""
        return self._sample_rate

    @sample_rate.setter
    def sample_rate(self, value):
        """Sets the stored sample rate parameter"""
        self._sample_rate = value

    @property
    def ramp_samples(self):
        """Number of samples in the ramp time"""
        return int(self.ramp_time * self.sample_rate)

    @property
    def number_of_channels(self):
        """Total number of channels in the environment"""
        return self._number_of_channels

    @number_of_channels.setter
    def number_of_channels(self, value):
        """Sets the stored value of the number of channels in the environment"""
        self._number_of_channels = value

    @property
    def reference_channel_indices(self):
        """Indices corresponding to the drive channels"""
        return self.output_channel_indices

    @property
    def response_channel_indices(self):
        """Indices corresponding to the control channels"""
        return self.control_channel_indices

    @property
    def response_transformation_matrix(self):
        """Transformation matrix applied to the control channels"""
        return self._response_transformation_matrix

    @response_transformation_matrix.setter
    def response_transformation_matrix(self, value):
        """Sets the transformation matrix applied to the control channels"""
        self._response_transformation_matrix = value

    @property
    def reference_transformation_matrix(self):
        """Transformation matrix applied to the drive channels"""
        return self._reference_transformation_matrix

    @reference_transformation_matrix.setter
    def reference_transformation_matrix(self, value):
        """Sets the transformation matrix applied to the drive channels"""
        self._reference_transformation_matrix = value

    def validate(self, hardware_metadata):
        return super().validate(hardware_metadata)

    # region: Loading
    def store_to_netcdf(
        self,
        netcdf_group_handle: nc4._netCDF4.Group,  # pylint: disable=c-extension-no-member
    ):
        """Store parameters to a group in a netCDF streaming file.

        This function stores parameters from the environment into the netCDF
        file in a group with the environment's name as its name.  The function
        will receive a reference to the group within the dataset and should
        store the environment's parameters into that group in the form of
        attributes, dimensions, or variables.

        This function is the "write" counterpart to the retrieve_metadata
        function in the AbstractUI class, which will read parameters from
        the netCDF file to populate the parameters in the user interface.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A reference to the Group within the netCDF dataset where the
            environment's metadata is stored.

        """
        super().store_to_netcdf(netcdf_group_handle)

        netcdf_group_handle.sample_rate = self.sample_rate
        netcdf_group_handle.samples_per_frame = self.samples_per_frame
        netcdf_group_handle.ramp_time = self.ramp_time
        netcdf_group_handle.number_of_channels = self.number_of_channels
        netcdf_group_handle.update_drives_after_environment = 1 if self.update_drives_after_environment else 0
        netcdf_group_handle.phase_fit = 1 if self.phase_fit else 0
        netcdf_group_handle.control_convergence = self.control_convergence
        netcdf_group_handle.allow_automatic_aborts = 1 if self.allow_automatic_aborts else 0
        netcdf_group_handle.control_python_script = "" if self.control_python_script is None else self.control_python_script
        netcdf_group_handle.control_python_class = "" if self.control_python_script is None else self.control_python_class
        netcdf_group_handle.control_python_parameters = "" if self.control_python_script is None else self.control_python_parameters
        netcdf_group_handle.tracking_filter_type = self.tracking_filter_type
        netcdf_group_handle.tracking_filter_cutoff = self.tracking_filter_cutoff
        netcdf_group_handle.tracking_filter_order = self.tracking_filter_order
        netcdf_group_handle.vk_filter_order = self.vk_filter_order
        netcdf_group_handle.vk_filter_bandwidth = self.vk_filter_bandwidth
        netcdf_group_handle.vk_filter_blocksize = self.vk_filter_blocksize
        netcdf_group_handle.vk_filter_overlap = self.vk_filter_overlap
        netcdf_group_handle.buffer_blocks = self.buffer_blocks

        # Control channels
        netcdf_group_handle.createDimension("control_channels", len(self.control_channel_indices))
        var = netcdf_group_handle.createVariable("control_channel_indices", "i4", ("control_channels"))
        var[...] = self.control_channel_indices
        # Transformation matrices
        if self.response_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "response_transformation_rows",
                self.response_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "response_transformation_cols",
                self.response_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "response_transformation_matrix",
                "f8",
                ("response_transformation_rows", "response_transformation_cols"),
            )
            var[...] = self.response_transformation_matrix
        if self.reference_transformation_matrix is not None:
            netcdf_group_handle.createDimension(
                "reference_transformation_rows",
                self.reference_transformation_matrix.shape[0],
            )
            netcdf_group_handle.createDimension(
                "reference_transformation_cols",
                self.reference_transformation_matrix.shape[1],
            )
            var = netcdf_group_handle.createVariable(
                "reference_transformation_matrix",
                "f8",
                ("reference_transformation_rows", "reference_transformation_cols"),
            )
            var[...] = self.reference_transformation_matrix
        # Specification
        spec_group = netcdf_group_handle.createGroup("specifications")
        for specification in self.specifications:
            specification: SineSpecification
            grp = spec_group.createGroup(specification.name)
            grp.start_time = specification.start_time
            grp.createDimension("num_breakpoints", len(specification.breakpoint_table))
            grp.createDimension(
                "specification_channels",
                specification.breakpoint_table["amplitude"].shape[-1],
            )
            grp.createDimension("two", 2)
            var = grp.createVariable("spec_frequency", "f8", ("num_breakpoints"))
            var[...] = specification.breakpoint_table["frequency"]
            var = grp.createVariable("spec_amplitude", "f8", ("num_breakpoints", "specification_channels"))
            var[...] = specification.breakpoint_table["amplitude"]
            var = grp.createVariable("spec_phase", "f8", ("num_breakpoints", "specification_channels"))
            var[...] = specification.breakpoint_table["phase"]
            var = grp.createVariable("spec_sweep_type", "i1", ("num_breakpoints"))
            var[...] = specification.breakpoint_table["sweep_type"]
            var = grp.createVariable("spec_sweep_rate", "f8", ("num_breakpoints"))
            var[...] = specification.breakpoint_table["sweep_rate"]
            var = grp.createVariable(
                "spec_warning",
                "f8",
                ("num_breakpoints", "two", "two", "specification_channels"),
            )
            var[...] = specification.breakpoint_table["warning"]
            var = grp.createVariable(
                "spec_abort",
                "f8",
                ("num_breakpoints", "two", "two", "specification_channels"),
            )
            var[...] = specification.breakpoint_table["abort"]

    @classmethod
    def retrieve_metadata_from_netcdf(
        cls,
        netcdf_group_handle: nc4._netCDF4.Group,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf
        function in the ModalMetadata class, which will write parameters to
        the netCDF file to document the metadata.

        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.

        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``

        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset :
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.

        """

        sysid_metadata = SysIdMetadata.retrieve_metadata_from_netcdf(netcdf_group_handle, hardware_metadata)
        # Get the group
        sample_rate = hardware_metadata.sample_rate
        samples_per_frame = netcdf_group_handle.samples_per_frame
        number_of_channels = netcdf_group_handle.number_of_channels
        ramp_time = netcdf_group_handle.ramp_time
        buffer_blocks = netcdf_group_handle.buffer_blocks
        control_convergence = netcdf_group_handle.control_convergence
        update_drives_after_environment = bool(netcdf_group_handle.update_drives_after_environment)
        phase_fit = bool(netcdf_group_handle.phase_fit)
        allow_automatic_aborts = bool(netcdf_group_handle.allow_automatic_aborts)
        tracking_filter_type = netcdf_group_handle.tracking_filter_type
        tracking_filter_cutoff = netcdf_group_handle.tracking_filter_cutoff
        tracking_filter_order = netcdf_group_handle.tracking_filter_order
        vk_filter_order = netcdf_group_handle.vk_filter_order
        vk_filter_bandwidth = netcdf_group_handle.vk_filter_bandwidth
        vk_filter_blocksize = netcdf_group_handle.vk_filter_blocksize
        vk_filter_overlap = netcdf_group_handle.vk_filter_overlap
        control_python_script = netcdf_group_handle.control_python_script
        control_python_class = netcdf_group_handle.control_python_class
        control_python_parameters = netcdf_group_handle.control_python_parameters
        control_channel_indices = netcdf_group_handle.variables["control_channel_indices"][...]
        environment_channel_list = [channel for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools) if channel_bool]
        output_channel_indices = [index for index, channel in enumerate(environment_channel_list) if channel.feedback_device is not None]
        response_transformation_matrix = None
        if "response_transformation_matrix" in netcdf_group_handle.variables:
            response_transformation_matrix = netcdf_group_handle.variables["response_transformation_matrix"][...]
        output_transformation_matrix = None
        if "output_transformation_matrix" in netcdf_group_handle.variables:
            output_transformation_matrix = netcdf_group_handle.variables["output_transformation_matrix"][...]
        specifications = []
        if "specifications" in netcdf_group_handle.groups:
            spec_group = netcdf_group_handle.groups["specifications"]
            for spec_name, grp in spec_group.groups.items():
                start_time = grp.start_time
                frequency = grp.variables["spec_frequency"][...]
                amplitude = grp.variables["spec_amplitude"][...]
                phase = grp.variables["spec_phase"][...]
                sweep_type = grp.variables["spec_sweep_type"][...]
                sweep_rate = grp.variables["spec_sweep_rate"][...]
                warning = grp.variables["spec_warning"][...]
                abort = grp.variables["spec_abort"][...]
                num_control = amplitude.shape[-1]
                spec = SineSpecification(
                    name=spec_name,
                    start_time=start_time,
                    num_control=num_control,
                    frequency_breakpoints=frequency,
                    amplitude_breakpoints=amplitude,
                    phase_breakpoints=phase,
                    sweep_type_breakpoints=sweep_type[:-1],  # matches constructor logic
                    sweep_rate_breakpoints=sweep_rate[:-1],  # matches constructor logic
                    warning_breakpoints=warning,
                    abort_breakpoints=abort,
                )
                specifications.append(spec)
        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            samples_per_frame=samples_per_frame,
            number_of_channels=number_of_channels,
            specifications=specifications,
            ramp_time=ramp_time,
            buffer_blocks=buffer_blocks,
            control_convergence=control_convergence,
            update_drives_after_environment=update_drives_after_environment,
            phase_fit=phase_fit,
            allow_automatic_aborts=allow_automatic_aborts,
            tracking_filter_type=tracking_filter_type,
            tracking_filter_cutoff=tracking_filter_cutoff,
            tracking_filter_order=tracking_filter_order,
            vk_filter_order=vk_filter_order,
            vk_filter_bandwidth=vk_filter_bandwidth,
            vk_filter_blocksize=vk_filter_blocksize,
            vk_filter_overlap=vk_filter_overlap,
            control_python_script=control_python_script,
            control_python_class=control_python_class,
            control_python_parameters=control_python_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=output_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

    @staticmethod
    def create_blank_worksheet_template(worksheet):
        worksheet.cell(1, 1, "Control Type")
        worksheet.cell(1, 2, "Sine")
        worksheet.cell(
            1,
            4,
            "Note: Replace cells with hash marks (#) to provide the requested parameters.",
        )
        worksheet.cell(2, 1, "Test Ramp Time")
        worksheet.cell(2, 3, "# Time for the test to ramp up or down when starting or stopping")
        worksheet.cell(3, 1, "Control Convergence")
        worksheet.cell(
            3,
            3,
            "# A scale factor on the closed-loop update to " "balance stability with speed of convergence",
        )
        worksheet.cell(4, 1, "Update Drives after Environment:")
        worksheet.cell(
            4,
            3,
            "# If Y, then a control calculation will be performed after the " "environment finishes to update the next drive signal (Y/N)",
        )
        worksheet.cell(5, 1, "Fit Phases")
        worksheet.cell(
            5,
            3,
            "# If Y, perform a best fit to phase quantities to accommodate time delays (Y/N)",
        )
        worksheet.cell(6, 1, "Allow Automatic Aborts")
        worksheet.cell(
            6,
            3,
            "# Shut down the test automatically if an abort level is reached (Y/N)",
        )
        worksheet.cell(7, 1, "Buffer Blocks")
        worksheet.cell(
            7,
            3,
            "# Number of write blocks to keep in the buffer to " "guard against running out of samples to generate",
        )
        worksheet.cell(8, 1, "Tracking Filter Type")
        worksheet.cell(
            8,
            3,
            "# Select the tracking filter type to use " "(VK - Vold-Kalman / DFT - Digital Tracking Filter)",
        )
        worksheet.cell(9, 1, "Digital Tracking Filter Cutoff Percent:")
        worksheet.cell(
            9,
            3,
            "# Tracking filter cutoff frequency compared to the instantaneous frequency",
        )
        worksheet.cell(10, 1, "Digital Tracking Filter Order")
        worksheet.cell(10, 3, "# Order of the Butterworth filter used in the tracking filter")
        worksheet.cell(11, 1, "VK Filter Order")
        worksheet.cell(11, 3, "# Order of the Vold-Kalman Filter (1, 2, or 3)")
        worksheet.cell(12, 1, "VK Filter Bandwidth")
        worksheet.cell(12, 3, "# Bandwidth of the Vold-Kalman Filter")
        worksheet.cell(13, 1, "VK Filter Block Size")
        worksheet.cell(13, 3, "# Number of samples in the filter blocks for the Vold-Kalman Filter")
        worksheet.cell(14, 1, "VK Filter Overlap")
        worksheet.cell(14, 3, "Overlap between frames in the VK filter as a fraction (0.5, not 50)")
        worksheet.cell(15, 1, "Custom Control Python Script:")
        worksheet.cell(15, 3, "# Path to the Python script containing the control law")
        worksheet.cell(16, 1, "Custom Control Python Class:")
        worksheet.cell(
            16,
            3,
            "# Class name within the Python Script that will serve as the control law",
        )
        worksheet.cell(17, 1, "Control Parameters:")
        worksheet.cell(17, 3, "# Extra parameters used in the control law")
        worksheet.cell(18, 1, "Control Channels (1-based):")
        worksheet.cell(18, 3, "# List of channels, one per cell on this row")
        SysIdMetadata.create_blank_worksheet_template(worksheet, start_row=19)
        worksheet.cell(33, 1, "Specification File:")
        worksheet.cell(33, 3, "# Path to the file containing the Specification. Can specify multiple by using multiple columns")
        worksheet.cell(34, 1, "Response Transformation Matrix:")
        worksheet.cell(
            34,
            2,
            (
                "# Transformation matrix to apply to the response channels.  Type None if there is "
                "none.  Otherwise, make this a 2D array in the spreadsheet and move the Output "
                "Transformation Matrix line down so it will fit.  The number of columns should be "
                "the number of physical control channels."
            ),
        )
        worksheet.cell(35, 1, "Output Transformation Matrix:")
        worksheet.cell(
            35,
            2,
            "# Transformation matrix to apply to the outputs.  Type None if there is none.  "
            "Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be "
            "the number of physical output channels in the environment.",
        )

    def store_to_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        super().store_to_worksheet(worksheet)

        if self.ramp_time is not None:
            worksheet.cell(2, 2, self.ramp_time)
        if self.control_convergence is not None:
            worksheet.cell(3, 2, self.control_convergence)
        if self.update_drives_after_environment is not None:
            worksheet.cell(4, 2, "Y" if self.update_drives_after_environment else "N")
        if self.phase_fit is not None:
            worksheet.cell(5, 2, "Y" if self.phase_fit else "N")
        if self.allow_automatic_aborts is not None:
            worksheet.cell(6, 2, "Y" if self.allow_automatic_aborts else "N")
        if self.buffer_blocks is not None:
            worksheet.cell(7, 2, self.buffer_blocks)
        if self.tracking_filter_type is not None:
            worksheet.cell(8, 2, "DFT" if self.tracking_filter_type == 0 else "VK")
        if self.tracking_filter_cutoff is not None:
            worksheet.cell(9, 2, self.tracking_filter_cutoff)
        if self.tracking_filter_order is not None:
            worksheet.cell(10, 2, self.tracking_filter_order)
        if self.vk_filter_order is not None:
            worksheet.cell(11, 2, self.vk_filter_order)
        if self.vk_filter_bandwidth is not None:
            worksheet.cell(12, 2, self.vk_filter_bandwidth)
        if self.vk_filter_blocksize is not None:
            worksheet.cell(13, 2, self.vk_filter_blocksize)
        if self.vk_filter_overlap is not None:
            worksheet.cell(14, 2, self.vk_filter_overlap)
        if self.control_python_script is not None:
            worksheet.cell(15, 2, self.control_python_script)
        if self.control_python_class is not None:
            worksheet.cell(16, 2, self.control_python_class)
        if self.control_python_parameters is not None:
            worksheet.cell(17, 2, self.control_python_parameters)
        if self.control_channel_indices is not None:
            for idx, channel_ind in enumerate(self.control_channel_indices):
                col_idx = idx + 2
                worksheet.cell(18, col_idx, channel_ind + 1)
        self.sysid_metadata.store_to_worksheet(worksheet, start_row=19)
        response_row = 34
        output_row = 35
        if self.response_transformation_matrix is not None:
            worksheet.cell(35, 1, None)
            worksheet.cell(35, 2, None)
            for i, row in enumerate(self.response_transformation_matrix):
                for j, value in enumerate(row):
                    worksheet.cell(i + response_row, j + 2, value)
            # Shift output transfomation matrix down
            output_row = i + 1
            worksheet.cell(i + 1, 1, "Output Transformation Matrix:")
            worksheet.cell(
                i + 1,
                2,
                "# Transformation matrix to apply to the outputs.  Type None if there is none.  "
                "Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be "
                "the number of physical output channels in the environment.",
            )
        if self.response_transformation_matrix is not None:
            for i, row in enumerate(self.response_transformation_matrix):
                for j, value in enumerate(row):
                    worksheet.cell(i + output_row, j + 2, value)

    @classmethod
    def retrieve_metadata_from_worksheet(
        cls,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        environment_name: str,
        channel_list_bools: List[bool],
        hardware_metadata: HardwareMetadata,
    ):
        sample_rate = hardware_metadata.sample_rate
        number_of_channels = sum(channel_list_bools)
        samples_per_frame = hardware_metadata.samples_per_read
        environment_channel_list = [channel for channel, channel_bool in zip(hardware_metadata.channel_list, channel_list_bools) if channel_bool]
        output_channel_indices = [index for index, channel in enumerate(environment_channel_list) if channel.feedback_device is not None]

        ramp_time = float(worksheet.cell(2, 2).value)
        control_convergence = float(worksheet.cell(3, 2).value)
        update_drives_after_environment = worksheet.cell(4, 2).value.upper() == "Y"
        phase_fit = worksheet.cell(5, 2).value.upper() == "Y"
        allow_automatic_aborts = worksheet.cell(6, 2).value.upper() == "Y"
        buffer_blocks = int(worksheet.cell(7, 2).value)
        tracking_filter_type = 1 if worksheet.cell(8, 2).value.upper() == "VK" else 0
        tracking_filter_cutoff = float(worksheet.cell(9, 2).value)
        tracking_filter_order = int(worksheet.cell(10, 2).value)
        vk_filter_order = int(worksheet.cell(11, 2).value)
        vk_filter_bandwidth = float(worksheet.cell(12, 2).value)
        vk_filter_blocksize = int(worksheet.cell(13, 2).value)
        vk_filter_overlap = float(worksheet.cell(14, 2).value)
        control_python_script = worksheet.cell(15, 2).value if worksheet.cell(15, 2).value is not None else ""
        control_python_class = worksheet.cell(16, 2).value if worksheet.cell(16, 2).value is not None else ""
        control_python_parameters = worksheet.cell(17, 2).value if worksheet.cell(17, 2).value is not None else ""
        control_channel_indices = []
        column_index = 2
        while True:
            channel_ind = worksheet.cell(18, column_index).value
            if channel_ind is None or (isinstance(channel_ind, str) and channel_ind.strip() == ""):
                break
            try:
                control_channel_indices.append(int(channel_ind) - 1)
            except:
                break  # This is incase it cant be converted to int
            column_index += 1
        sysid_metadata = SysIdMetadata.retrieve_metadata_from_worksheet(worksheet, hardware_metadata, start_row=19)

        # Now we need to find the transformation matrices' sizes
        start_response_row = 34
        num_response_row = 1
        if isinstance(worksheet.cell(start_response_row, 2).value, str) and worksheet.cell(start_response_row, 2).value.lower() == "none":
            response_transformation_matrix = None
        elif worksheet.cell(start_response_row, 2).value.lower().startswith("# transformation matrix"):
            response_transformation_matrix = None
        else:

            while True:
                first_col_value = worksheet.cell(start_response_row + num_response_row, 2).value
                if worksheet.cell(start_response_row + num_response_row, 1).value == "Output Transformation Matrix:" or (
                    first_col_value is None or (isinstance(first_col_value, str) and first_col_value.strip() == "")
                ):
                    break
                num_response_row += 1
            response_transformation_matrix = read_transformation_matrix_from_worksheet(
                worksheet, start_row=start_response_row, num_rows=num_response_row, start_col=2
            )
        # Output transformation matrix
        start_output_row = start_response_row + num_response_row
        num_output_row = 1
        if isinstance(worksheet.cell(start_output_row, 2).value, str) and worksheet.cell(start_output_row, 2).value.lower() == "none":
            output_transformation_matrix = None
        elif worksheet.cell(start_output_row, 2).value.lower().startswith("# transformation matrix"):
            output_transformation_matrix = None
        else:
            while True:
                first_col_value = worksheet.cell(start_output_row + num_output_row, 2).value
                if first_col_value is None or (isinstance(first_col_value, str) and first_col_value.strip() == ""):
                    break
                num_output_row += 1
            output_transformation_matrix = read_transformation_matrix_from_worksheet(
                worksheet, start_row=start_output_row, num_rows=num_output_row, start_col=2
            )

        # Specification Files
        specification_files = []
        column_index = 2
        while True:
            filename = worksheet.cell(33, column_index).value
            if filename is None or (isinstance(filename, str) and filename.strip() == ""):
                break
            specification_files.append(str(filename))
            column_index += 1
        specifications = []
        for filename in specification_files:
            (
                frequencies,
                amplitudes,
                phases,  # Degrees
                sweep_types,
                sweep_rates,
                warnings,
                aborts,
                start_time,
                name,
            ) = load_specification(filename)
            spec = SineSpecification(
                name=name,
                start_time=start_time,
                num_control=len(control_channel_indices),
                frequency_breakpoints=frequencies,
                amplitude_breakpoints=amplitudes,
                phase_breakpoints=phases,
                sweep_type_breakpoints=sweep_types,
                sweep_rate_breakpoints=sweep_rates,
                warning_breakpoints=warnings,
                abort_breakpoints=aborts,
            )
            specifications.append(spec)

        return cls(
            environment_name=environment_name,
            channel_list_bools=channel_list_bools,
            sample_rate=sample_rate,
            samples_per_frame=samples_per_frame,
            number_of_channels=number_of_channels,
            specifications=specifications,
            ramp_time=ramp_time,
            buffer_blocks=buffer_blocks,
            control_convergence=control_convergence,
            update_drives_after_environment=update_drives_after_environment,
            phase_fit=phase_fit,
            allow_automatic_aborts=allow_automatic_aborts,
            tracking_filter_type=tracking_filter_type,
            tracking_filter_cutoff=tracking_filter_cutoff,
            tracking_filter_order=tracking_filter_order,
            vk_filter_order=vk_filter_order,
            vk_filter_bandwidth=vk_filter_bandwidth,
            vk_filter_blocksize=vk_filter_blocksize,
            vk_filter_overlap=vk_filter_overlap,
            control_python_script=control_python_script,
            control_python_class=control_python_class,
            control_python_parameters=control_python_parameters,
            control_channel_indices=control_channel_indices,
            output_channel_indices=output_channel_indices,
            response_transformation_matrix=response_transformation_matrix,
            output_transformation_matrix=output_transformation_matrix,
            sysid_metadata=sysid_metadata,
        )

    def set_parameters_from_template(self, worksheet):

        # Now we need to find the transformation matrices' sizes
        response_channels = self.definition_widget.control_channels_display.value()
        output_channels = self.definition_widget.output_channels_display.value()
        output_transform_row = 35
        if isinstance(worksheet.cell(34, 2).value, str) and worksheet.cell(34, 2).value.lower() == "none":
            self.response_transformation_matrix = None
        else:
            while True:
                if worksheet.cell(output_transform_row, 1).value == "Output Transformation Matrix:":
                    break
                output_transform_row += 1
            response_size = output_transform_row - 34
            response_transformation = []
            for i in range(response_size):
                response_transformation.append([])
                for j in range(response_channels):
                    response_transformation[-1].append(float(worksheet.cell(34 + i, 2 + j).value))
            self.response_transformation_matrix = np.array(response_transformation)
        if isinstance(worksheet.cell(output_transform_row, 2).value, str) and worksheet.cell(output_transform_row, 2).value.lower() == "none":
            self.output_transformation_matrix = None
        else:
            output_transformation = []
            i = 0
            while True:
                if worksheet.cell(output_transform_row + i, 2).value is None or (
                    isinstance(worksheet.cell(output_transform_row + i, 2).value, str)
                    and worksheet.cell(output_transform_row + i, 2).value.strip() == ""
                ):
                    break
                output_transformation.append([])
                for j in range(output_channels):
                    output_transformation[-1].append(float(worksheet.cell(output_transform_row + i, 2 + j).value))
                i += 1
            self.output_transformation_matrix = np.array(output_transformation)
        self.define_transformation_matrices(None, dialog=False)

        # Load in the specification
        if worksheet.cell(33, 2).value:
            self.sine_tables[0].load_specification(None, worksheet.cell(33, 2).value)
        column_index = 3
        while True:
            if worksheet.cell(33, column_index).value:
                self.add_sine_table_tab()
                self.sine_tables[-1].load_specification(None, worksheet.cell(33, column_index).value)
                column_index += 1
            else:
                break


# region: Instructions
class SineInstructions(EnvironmentInstructions):
    def __init__(self, environment_name, control_test_level, control_tones, control_start_time, control_end_time):
        super().__init__(CONTROL_TYPE, environment_name)
        self.control_test_level = control_test_level
        self.control_tones = control_tones
        self.control_start_time = control_start_time
        self.control_end_time = control_end_time

    def validate(self):
        return super().validate()


# region: Environment
class SineEnvironment(SysIdEnvironmentProcess):
    """Class representing the environment computations on a separate process from the main UI"""

    def __init__(
        self,
        environment_name: str,
        queue_name: str,
        queue_container: SineQueues,
        acquisition_active_event: mp.synchronize.Event,
        output_active_event: mp.synchronize.Event,
        active_event: mp.synchronize.Event,
        ready_event: mp.synchronize.Event,
        sysid_event: mp.synchronize.Event,
    ):
        """Initializes the sine environment computation class

        Parameters
        ----------
        environment_name : str
            Name of the environment
        queue_container : SineQueues
            A container containing all the queues that the Sine environment needs to
            pass information between its parts
        acquisition_active : int
            A multiprocessing shared value that is used to tell all processes whether or not
            the acquisition is currently running
        output_active : int
            A multiprocessing shared value that is used to tell all processes whether or not
            the output is currently running
        """
        super().__init__(
            environment_name,
            queue_name,
            queue_container.environment_command_queue,
            queue_container.gui_update_queue,
            queue_container.controller_communication_queue,
            queue_container.log_file_queue,
            queue_container.collector_command_queue,
            queue_container.signal_generation_command_queue,
            queue_container.spectral_command_queue,
            queue_container.data_analysis_command_queue,
            queue_container.data_in_queue,
            queue_container.data_out_queue,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
            sysid_event,
        )
        self.map_command(SineCommands.PERFORM_CONTROL_PREDICTION, self.perform_control_prediction)
        self.map_command(GlobalCommands.START_ENVIRONMENT, self.start_control)
        self.map_command(SineCommands.START_CONTROL, self.start_control)
        self.map_command(SineCommands.STOP_CONTROL, self.stop_environment)
        self.map_command(SineCommands.SAVE_CONTROL_DATA, self.save_control_data)
        self.map_command(SineCommands.SEND_RESPONSE_PREDICTION, self.send_response_prediction)
        self.map_command(SineCommands.SEND_EXCITATION_PREDICTION, self.send_excitation_prediction)
        self.map_command(SineCommands.SET_TEST_LEVEL, self.set_test_level)
        # Persistent data
        self.hardware_metadata = None
        self.environment_metadata = None
        self.queue_container = queue_container
        self.plot_downsample = None
        # Control data
        self.sysid_frequencies = None
        self.sysid_frf = None
        self.sysid_coherence = None
        self.sysid_response_cpsd = None
        self.sysid_reference_cpsd = None
        self.sysid_condition = None
        self.sysid_response_noise = None
        self.sysid_reference_noise = None
        self.sysid_frames = None
        self.control_class = None
        self.extra_control_parameters = None
        # Specification data
        self.specification_signals_combined = None
        self.specification_signals = None
        self.specification_frequencies = None
        self.specification_arguments = None
        self.specification_amplitudes = None
        self.specification_phases = None
        self.specification_start_indices = None
        self.specification_end_indices = None
        self.ramp_samples = None
        # Excitation Signal Data
        self.excitation_signals = None
        self.excitation_signals_combined = None
        self.excitation_signal_frequencies = None
        self.excitation_signal_arguments = None
        self.excitation_signal_amplitudes = None
        self.excitation_signal_phases = None
        self.peak_voltages = None
        # Predicted Response Data
        self.predicted_response_signals_combined = None
        self.predicted_response_signals = None
        self.predicted_response_amplitudes = None
        self.predicted_response_phases = None
        self.predicted_warning_matrix = None
        self.predicted_abort_matrix = None
        self.predicted_amplitude_error = None
        # Running data
        self.control_test_level = 0
        self.control_tones = None
        self.control_tone_indices = None
        self.control_start_time = None
        self.control_end_time = None
        self.control_time_delay = None
        self.control_write_index = 0
        self.control_read_index = 0
        self.control_analysis_index = 0
        self.control_finished = False
        self.control_analysis_finished = False
        self.control_startup = True
        self.control_first_signal = None
        self.control_response_signals_combined = None
        self.control_response_amplitudes = None
        self.control_response_phases = None
        self.control_response_frequencies = None
        self.control_response_arguments = None
        self.control_target_phases = None
        self.control_target_amplitudes = None
        self.control_specification_arguments = None
        self.control_drive_modifications = None
        self.control_block_size = None
        self.control_filters = None
        self.control_warning_flags = None
        self.control_abort_flags = None
        self.control_amplitude_errors = None
        self.control_start_index = None
        self.control_end_index = None
        self.good_line_threshold = 0.25

    # region: Initialize
    def initialize_hardware(self, hardware_metadata):
        return super().initialize_hardware(hardware_metadata)

    def initialize_environment(self, environment_metadata: SineMetadata):
        # Check if all specifications are equal
        if (
            self.environment_metadata is None
            or not np.array_equal(
                self.environment_metadata.control_channel_indices,
                environment_metadata.control_channel_indices,
            )
            or not (
                all(
                    [
                        spec1 == spec2
                        for spec1, spec2 in zip(
                            self.environment_metadata.specifications,
                            environment_metadata.specifications,
                        )
                    ]
                )
                and (len(self.environment_metadata.specifications) == len(environment_metadata.specifications))
            )
        ):
            self.sysid_frequencies = None
            self.sysid_frf = None
            self.sysid_coherence = None
            self.sysid_response_cpsd = None
            self.sysid_reference_cpsd = None
            self.sysid_condition = None
            self.sysid_response_noise = None
            self.sysid_reference_noise = None
            self.sysid_frames = None
            self.control_class = None
            self.extra_control_parameters = None
            self.excitation_signals_combined = None
            self.excitation_signals = None
            self.excitation_signal_frequencies = None
            self.excitation_signal_arguments = None
            self.excitation_signal_amplitudes = None
            self.excitation_signal_phases = None
            self.predicted_response_signals_combined = None
            self.predicted_response_signals = None
            self.predicted_response_amplitudes = None
            self.predicted_response_phases = None
            self.ramp_samples = None
        super().initialize_environment(environment_metadata)
        self.environment_metadata: SineMetadata
        if not environment_metadata.control_python_script:
            control_class = DefaultSineControlLaw
            self.extra_control_parameters = environment_metadata.control_python_parameters
        else:
            _, file = os.path.split(environment_metadata.control_python_script)
            file, _ = os.path.splitext(file)
            spec = importlib.util.spec_from_file_location(file, environment_metadata.control_python_script)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            self.extra_control_parameters = environment_metadata.control_python_parameters
            control_class = getattr(module, environment_metadata.control_python_class)
        self.control_class = control_class(
            self.environment_metadata.sample_rate,
            self.environment_metadata.specifications,
            self.hardware_metadata.output_oversample,
            self.environment_metadata.ramp_time,
            self.environment_metadata.control_convergence,
            self.hardware_metadata.samples_per_write,
            self.environment_metadata.buffer_blocks,
            self.extra_control_parameters,  # Required parameters
            self.environment_metadata.sysid_metadata.sysid_frequency_spacing,  # Frequency Spacing
            self.sysid_frf,  # Transfer Functions
            self.sysid_response_noise,  # Noise levels and correlation
            self.sysid_reference_noise,  # from the system identification
            self.sysid_response_cpsd,  # Response levels and correlation
            self.sysid_reference_cpsd,  # from the system identification
            self.sysid_coherence,  # Coherence from the system identification
            self.sysid_frames,  # Number of frames in the FRF matrices
        )
        self.log("Creating Specification Signals...")
        (
            self.specification_signals_combined,
            self.specification_signals,
            self.specification_frequencies,
            self.specification_arguments,
            self.specification_amplitudes,
            self.specification_phases,
            self.specification_start_indices,
            self.specification_end_indices,
        ) = SineSpecification.create_combined_signals(
            self.environment_metadata.specifications,
            self.hardware_metadata.sample_rate * self.hardware_metadata.output_oversample,
            self.environment_metadata.ramp_samples * self.hardware_metadata.output_oversample,
        )
        self.ramp_samples = self.environment_metadata.ramp_samples * self.hardware_metadata.output_oversample
        self.plot_downsample = (
            self.specification_signals_combined.shape[-1] // self.hardware_metadata.output_oversample // MAXIMUM_SAMPLES_TO_PLOT + 1
        )
        self.gui_update_queue.put(
            (
                self.environment_name,
                (
                    SineUICommands.SPECIFICATION_FOR_PLOTTING,
                    (
                        self.specification_signals_combined[
                            ...,
                            :: self.hardware_metadata.output_oversample * self.plot_downsample,
                        ],
                        self.specification_signals[
                            ...,
                            :: self.hardware_metadata.output_oversample * self.plot_downsample,
                        ],
                        self.specification_frequencies[
                            ...,
                            :: self.hardware_metadata.output_oversample * self.plot_downsample,
                        ],
                        self.specification_arguments[
                            ...,
                            :: self.hardware_metadata.output_oversample * self.plot_downsample,
                        ],
                        self.specification_amplitudes[
                            ...,
                            :: self.hardware_metadata.output_oversample * self.plot_downsample,
                        ],
                        wrap(
                            self.specification_phases[
                                ...,
                                :: self.hardware_metadata.output_oversample * self.plot_downsample,
                            ]
                        )
                        * 180
                        / np.pi,
                        self.plot_downsample,
                    ),
                ),
            )
        )
        self.log("Done!")

    def initialize_sysid(self, sysid_metadata):
        return super().initialize_sysid(sysid_metadata)

    # region: System Identification
    def system_id_complete(self, data):
        # print('Finished System Identification')
        self.log("Finished System Identification")
        super().system_id_complete(data)
        (
            self.sysid_frames,
            _,
            self.sysid_frequencies,
            self.sysid_frf,
            self.sysid_coherence,
            self.sysid_response_cpsd,
            self.sysid_reference_cpsd,
            self.sysid_condition,
            self.sysid_response_noise,
            self.sysid_reference_noise,
        ) = data
        # Perform the control prediction
        self.perform_control_prediction(True)

    def filter_predicted_signal(self):
        """Extract amplitude and phase information from predicted signals"""
        # print('Filtering Predicted Signal')
        predicted_signals = []
        predicted_amplitudes = []
        predicted_phases = []
        arguments = self.excitation_signal_arguments[..., :: self.hardware_metadata.output_oversample]
        frequencies = self.excitation_signal_frequencies[..., :: self.hardware_metadata.output_oversample]
        for signal in self.predicted_response_signals_combined:
            if self.environment_metadata.tracking_filter_type == 0:
                block_size = self.hardware_metadata.samples_per_read
                generator = [
                    digital_tracking_filter_generator(
                        dt=1 / self.environment_metadata.sample_rate,
                        cutoff_frequency_ratio=self.environment_metadata.tracking_filter_cutoff,
                        filter_order=self.environment_metadata.tracking_filter_order,
                    )
                    for tone in self.excitation_signals
                ]
                for gen in generator:
                    gen.send(None)
            else:
                block_size = self.environment_metadata.vk_filter_blocksize
                generator = vold_kalman_filter_generator(
                    sample_rate=self.environment_metadata.sample_rate,
                    num_orders=self.excitation_signals.shape[0],
                    block_size=block_size,
                    overlap=self.environment_metadata.vk_filter_overlap,
                    bandwidth=self.environment_metadata.vk_filter_bandwidth,
                    filter_order=self.environment_metadata.vk_filter_order,
                    buffer_size_factor=self.environment_metadata.buffer_blocks + 1,
                )
                generator.send(None)

            # print(f"{self.signal.shape=}")
            start_index = 0
            reconstructed_signals = []
            reconstructed_amplitudes = []
            reconstructed_phases = []

            last_data = False
            while not last_data:
                end_index = start_index + block_size
                block = signal[start_index:end_index]
                block_arguments = arguments[:, start_index:end_index]
                block_frequencies = frequencies[:, start_index:end_index]
                last_data = end_index >= signal.size
                # print(f"{block.shape=}, {block_arguments.shape=}, "
                # f"{block_frequencies.shape=}, {last_data=}, {block_size=}")
                if self.environment_metadata.tracking_filter_type == 0:
                    amps = []
                    phss = []
                    for arg, freq, gen in zip(block_arguments, block_frequencies, generator):
                        amp, phs = gen.send((block, freq, arg))
                        amps.append(amp)
                        phss.append(phs)
                    reconstructed_amplitudes.append(np.array(amps))
                    reconstructed_phases.append(np.array(phss))
                    reconstructed_signals.append(np.array(amps) * np.cos(block_arguments + np.array(phss)))
                else:
                    vk_signals, vk_amplitudes, vk_phases = generator.send((block, block_arguments, last_data))
                    if vk_signals is not None:
                        reconstructed_signals.append(vk_signals)
                        reconstructed_amplitudes.append(vk_amplitudes)
                        reconstructed_phases.append(vk_phases)
                start_index += block_size

            predicted_signals.append(np.concatenate(reconstructed_signals, axis=-1))
            predicted_amplitudes.append(np.concatenate(reconstructed_amplitudes, axis=-1))
            predicted_phases.append(np.concatenate(reconstructed_phases, axis=-1))

        self.predicted_response_signals = np.array(predicted_signals).transpose(1, 0, 2)
        self.predicted_response_amplitudes = np.array(predicted_amplitudes).transpose(1, 0, 2)
        self.predicted_response_phases = np.array(predicted_phases).transpose(1, 0, 2)
        # Pull out the data to compute maximum amplitude error
        self.predicted_amplitude_error = np.zeros(self.specification_amplitudes.shape[:2])
        for tone_index, (specs, preds, start_index, end_index) in enumerate(
            zip(
                self.specification_amplitudes,
                self.predicted_response_amplitudes,
                self.specification_start_indices,
                self.specification_end_indices,
            )
        ):
            for channel_index, (spec, pred) in enumerate(zip(specs, preds)):
                spec = spec[start_index : end_index : self.hardware_metadata.output_oversample]
                pred = pred[
                    start_index // self.hardware_metadata.output_oversample : start_index // self.hardware_metadata.output_oversample + spec.size
                ]
                max_error = np.max(np.abs(scale2db(pred / spec)))
                self.predicted_amplitude_error[tone_index, channel_index] = max_error

    def compare_predictions_to_warning_and_abort(self):
        """Compares the extracted prediction information to abort and warning levels"""
        specs = self.environment_metadata.specifications
        amps = self.predicted_response_amplitudes
        warning_matrix = np.zeros(amps.shape[:2], dtype=bool)
        abort_matrix = np.zeros(amps.shape[:2], dtype=bool)
        for tone_index in range(amps.shape[0]):
            freqs = self.excitation_signal_frequencies[
                tone_index,
                self.specification_start_indices[tone_index] : self.specification_end_indices[tone_index] : self.hardware_metadata.output_oversample,
            ]
            for channel_index in range(amps.shape[1]):
                warning_levels = specs[tone_index].interpolate_warning(channel_index, freqs)
                abort_levels = specs[tone_index].interpolate_abort(channel_index, freqs)
                predicted = amps[
                    tone_index,
                    channel_index,
                    self.specification_start_indices[tone_index]
                    // self.hardware_metadata.output_oversample : self.specification_start_indices[tone_index]
                    // self.hardware_metadata.output_oversample
                    + freqs.size,
                ]
                warning_ratio = predicted / warning_levels
                if np.any(warning_ratio[0] < 1.0):
                    warning_matrix[tone_index, channel_index] = True
                if np.any(warning_ratio[1] > 1.0):
                    warning_matrix[tone_index, channel_index] = True
                abort_ratio = predicted / abort_levels
                if np.any(abort_ratio[0] < 1.0):
                    abort_matrix[tone_index, channel_index] = True
                if np.any(abort_ratio[1] > 1.0):
                    abort_matrix[tone_index, channel_index] = True
        self.predicted_warning_matrix = warning_matrix
        self.predicted_abort_matrix = abort_matrix

    def perform_control_prediction(self, sysid_update):
        """Compute the prediction from the test by convolving with the transfer functions"""
        # print('Performing Control Prediction')
        if self.sysid_frf is None:
            self.gui_update_queue.put(
                (
                    UICommands.ERROR,
                    (
                        "Perform System Identification",
                        "Perform System ID before performing test predictions",
                    ),
                )
            )
            return
        # print('Computing Drive Signal')
        if sysid_update:
            self.log("Updating System Identification...")
            (
                self.excitation_signals,
                self.excitation_signal_frequencies,
                self.excitation_signal_arguments,
                self.excitation_signal_amplitudes,
                self.excitation_signal_phases,
            ) = self.control_class.system_id_update(
                self.environment_metadata.sysid_metadata.sysid_frequency_spacing,
                self.sysid_frf,  # Transfer Functions
                self.sysid_response_noise,  # Noise levels and correlation
                self.sysid_reference_noise,  # from the system identification
                self.sysid_response_cpsd,  # Response levels and correlation
                self.sysid_reference_cpsd,  # from the system identification
                self.sysid_coherence,  # Coherence from the system identification
                self.sysid_frames,  # Number of frames in the CPSD and FRF matrices
            )
            self.excitation_signals_combined = np.sum(self.excitation_signals, axis=0)
            self.peak_voltages = np.max(np.abs(self.excitation_signals_combined), axis=-1)
            self.log("Done!")
        # print('Performing Response Prediction')
        # print('Drive Signals {:}'.format(self.next_drive.shape))
        drive_signals = self.excitation_signals_combined[:, :: self.hardware_metadata.output_oversample]
        impulse_responses = np.moveaxis(np.fft.irfft(self.sysid_frf, axis=0), 0, -1)

        self.log("Predicting Test Response...")
        self.predicted_response_signals_combined = np.zeros((impulse_responses.shape[0], drive_signals.shape[-1]))

        for i, impulse_response_row in enumerate(impulse_responses):
            for impulse, drive in zip(impulse_response_row, drive_signals):
                # print('Convolving {:},{:}'.format(i,j))
                self.predicted_response_signals_combined[i, :] += sig.convolve(drive, impulse, "full")[: drive_signals.shape[-1]]
        self.log("Done!")

        self.log("Filtering Predicted Signals...")
        self.filter_predicted_signal()
        self.log("Done!")

        # print('From Performing Control Predictions')
        # print(f'{self.excitation_signals_combined.shape=}')
        # print(f'{self.excitation_signals.shape=}')
        # print(f'{self.excitation_signal_frequencies.shape=}')
        # print(f'{self.excitation_signal_arguments.shape=}')
        # print(f'{self.excitation_signal_amplitudes.shape=}')
        # print(f'{self.excitation_signal_phases.shape=}')
        # print(f'{self.ramp_samples=}')
        # print(f'{self.predicted_response_signals_combined.shape=}')
        # print(f'{self.predicted_response_signals.shape=}')
        # print(f'{self.predicted_response_amplitudes.shape=}')
        # print(f'{self.predicted_response_phases.shape=}')

        self.log("Comparing to Warning and Abort Curves")
        self.compare_predictions_to_warning_and_abort()
        self.log("Done!")

        self.log("Showing Test Predictions...")
        self.show_test_prediction()
        self.log("Done!")

    def show_test_prediction(self):
        """Starts the process to show the predictions by requesting the current plot choices"""
        self.gui_update_queue.put((self.environment_name, (SineUICommands.REQUEST_PREDICTION_PLOT_CHOICES, None)))
        self.gui_update_queue.put((self.environment_name, (SineUICommands.EXCITATION_VOLTAGE_LIST, self.peak_voltages)))
        self.gui_update_queue.put(
            (
                self.environment_name,
                (
                    SineUICommands.RESPONSE_ERROR_MATRIX,
                    (
                        self.predicted_amplitude_error,
                        self.predicted_warning_matrix,
                        self.predicted_abort_matrix,
                    ),
                ),
            )
        )

    def send_excitation_prediction(self, excitation_plot_choices):
        """Sends the predicted excitation for the channel, tone, and data type requested"""
        channel_index, type_index, tone_index = excitation_plot_choices
        # print(f'Excitation Predictions: {channel_index=}, {type_index=}, {tone_index}')
        if type_index == 0:  # Time histories
            if tone_index == -1:
                ordinate = self.excitation_signals_combined[
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ]
                abscissa = np.arange(ordinate.shape[-1]) / self.hardware_metadata.sample_rate * self.plot_downsample
            else:
                ordinate = self.excitation_signals[
                    tone_index,
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ]
                abscissa = np.arange(ordinate.shape[-1]) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 1:  # Amplitude Vs Time
            ordinate = self.excitation_signal_amplitudes[
                tone_index,
                channel_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
            abscissa = np.arange(ordinate.shape[-1]) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 2:  # Phase Vs Time
            ordinate = (
                wrap(
                    self.excitation_signal_phases[
                        tone_index,
                        channel_index,
                        :: self.plot_downsample * self.hardware_metadata.output_oversample,
                    ]
                )
                * 180
                / np.pi
            )
            abscissa = np.arange(ordinate.shape[-1]) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 3:  # Amplitude Vs Frequency
            ordinate = self.excitation_signal_amplitudes[
                tone_index,
                channel_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
            abscissa = self.specification_frequencies[
                tone_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
        elif type_index == 4:  # Phase Vs Frequency
            ordinate = (
                wrap(
                    self.excitation_signal_phases[
                        tone_index,
                        channel_index,
                        :: self.plot_downsample * self.hardware_metadata.output_oversample,
                    ]
                )
                * 180
                / np.pi
            )
            abscissa = self.specification_frequencies[
                tone_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
        else:
            raise ValueError(f"Undefined type_index {type_index}")
        # print(f'{ordinate.shape=}, {abscissa.shape=}')
        # print(f'{abscissa.min()=}, {abscissa.max()=}')
        self.gui_update_queue.put((self.environment_name, (SineUICommands.EXCITATION_PRECDICTION, (abscissa, ordinate))))

    def send_response_prediction(self, response_plot_choices):
        """Sends the response predictions at the requested channel, tone, and data type"""
        channel_index, type_index, tone_index = response_plot_choices
        # print(f'Response Predictions: {channel_index=}, {type_index=}, {tone_index}')
        if type_index == 0:  # Time histories
            if tone_index == -1:
                ordinate = [
                    self.specification_signals_combined[
                        channel_index,
                        :: self.plot_downsample * self.hardware_metadata.output_oversample,
                    ],
                    self.predicted_response_signals_combined[channel_index, :: self.plot_downsample],
                ]
                abscissa = np.arange(max(v.shape[-1] for v in ordinate)) / self.hardware_metadata.sample_rate * self.plot_downsample
            else:
                ordinate = [
                    self.specification_signals[
                        tone_index,
                        channel_index,
                        :: self.plot_downsample * self.hardware_metadata.output_oversample,
                    ],
                    self.predicted_response_signals[tone_index, channel_index, :: self.plot_downsample],
                ]
                abscissa = np.arange(max(v.shape[-1] for v in ordinate)) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 1:  # Amplitude Vs Time
            ordinate = [
                self.specification_amplitudes[
                    tone_index,
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ],
                self.predicted_response_amplitudes[tone_index, channel_index, :: self.plot_downsample],
            ]
            abscissa = np.arange(max(v.shape[-1] for v in ordinate)) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 2:  # Phase Vs Time
            ordinate = [
                self.specification_phases[
                    tone_index,
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ]
                * 180
                / np.pi,
                self.predicted_response_phases[tone_index, channel_index, :: self.plot_downsample] * 180 / np.pi,
            ]
            abscissa = np.arange(max(v.shape[-1] for v in ordinate)) / self.hardware_metadata.sample_rate * self.plot_downsample
        elif type_index == 3:  # Amplitude Vs Frequency
            ordinate = [
                self.specification_amplitudes[
                    tone_index,
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ],
                self.predicted_response_amplitudes[tone_index, channel_index, :: self.plot_downsample],
            ]
            abscissa = self.specification_frequencies[
                tone_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
        elif type_index == 4:  # Phase Vs Frequency
            ordinate = [
                self.specification_phases[
                    tone_index,
                    channel_index,
                    :: self.plot_downsample * self.hardware_metadata.output_oversample,
                ]
                * 180
                / np.pi,
                self.predicted_response_phases[tone_index, channel_index, :: self.plot_downsample] * 180 / np.pi,
            ]
            abscissa = self.specification_frequencies[
                tone_index,
                :: self.plot_downsample * self.hardware_metadata.output_oversample,
            ]
        else:
            raise ValueError(f"Undefined type_index {type_index}")
        # print(f'{ordinate[0].shape=}, {ordinate[1].shape=}, {abscissa.shape=}')
        # print(f'{abscissa.min()=}, {abscissa.max()=}')
        self.gui_update_queue.put((self.environment_name, (SineUICommands.RESPONSE_PREDICTION, (abscissa, ordinate))))

    def compute_spec_amplitudes_and_phases(self):
        """Computes amplitude and phase information from the specification"""
        spec_ordinates = []
        spec_amplitudes = []
        spec_phases = []
        spec_frequencies = []
        spec_arguments = []

        for channel_index in range(len(self.environment_metadata.control_channel_indices)):
            spec = self.environment_metadata.specification
            # Convert octave per min to octave per second
            sweep_rates = spec["sweep_rate"].copy()
            sweep_rates[spec["sweep_type"] == 1] = sweep_rates[spec["sweep_type"] == 1] / 60
            # Create the sweep types array
            sweep_types = ["lin" if sweep_type == 0 else "log" for sweep_type in spec["sweep_type"][:-1]]
            spec_ordinate, spec_argument, spec_frequency, spec_amplitude, spec_phase = sine_sweep(
                1 / self.hardware_metadata.sample_rate,
                spec["frequency"],
                sweep_rates,
                sweep_types,
                spec["amplitude"][:, channel_index],
                spec["phase"][:, channel_index],
                return_frequency=True,
                return_argument=True,
                return_amplitude=True,
                return_phase=True,
            )
            spec_ordinates.append(spec_ordinate)
            spec_amplitudes.append(spec_amplitude)
            spec_phases.append(spec_phase)
            spec_arguments.append(spec_argument)
            spec_frequencies.append(spec_frequency)

        return (
            np.array(spec_ordinates),
            np.array(spec_arguments),
            np.array(spec_frequencies),
            np.array(spec_amplitudes),
            np.array(spec_phases),
        )

    # region: Control Loop
    def get_signal_generation_metadata(self):
        """Gets a SignalGenerationMetadata object for the current environment"""
        return SignalGenerationMetadata(
            samples_per_write=self.hardware_metadata.samples_per_write,
            level_ramp_samples=self.environment_metadata.ramp_time * self.environment_metadata.sample_rate * self.hardware_metadata.output_oversample,
            output_transformation_matrix=self.environment_metadata.reference_transformation_matrix,
        )

    def start_control(self, data: SineInstructions):
        """
        Starts up and runs the control with the specified test level,
        tones, start and end times
        """
        if self.control_startup:
            self.log("Starting Environment")
            # Read in the starting parameters
            self.control_test_level = db2scale(data.control_test_level)
            self.control_tones = data.control_tones
            self.control_start_time = data.control_start_time
            self.control_end_time = data.control_end_time
            if self.control_tones is not None and len(self.control_tones) == 0:
                self.control_tones = None
            if self.control_tones is None:
                self.control_tones = slice(None)
                self.control_tone_indices = np.arange(self.excitation_signal_arguments.shape[0])
            else:
                self.control_tone_indices = self.control_tones
            # Precompute the number of channels for convenience
            n_control_channels = (
                len(self.environment_metadata.control_channel_indices)
                if self.environment_metadata.response_transformation_matrix is None
                else self.environment_metadata.response_transformation_matrix.shape[0]
            )
            n_output_channels = (
                len(self.environment_metadata.output_channel_indices)
                if self.environment_metadata.reference_transformation_matrix is None
                else self.environment_metadata.reference_transformation_matrix.shape[0]
            )

            self.control_time_delay = None  # We will need to compute this when we get our first data point
            if DEBUG:
                num_files = len(glob(FILE_OUTPUT.format("*")))
                np.savez(
                    FILE_OUTPUT.format(num_files),
                    excitation_signal=self.control_first_signal,
                    done_controlling=False,
                )
            # print('Parsing Indices')
            # Parse out frequency information to get the indices into the full
            # arrays that we are controlling to
            if self.control_start_time is None:
                self.control_start_index = 0
            else:
                self.control_start_index = int(
                    self.control_start_time * self.hardware_metadata.sample_rate * self.hardware_metadata.output_oversample
                )
                if self.control_start_index < 0:
                    self.control_start_index = 0
            if self.control_end_time is None:
                self.control_end_index = None
            else:
                self.control_end_index = (
                    int(self.control_end_time * self.hardware_metadata.sample_rate * self.hardware_metadata.output_oversample) + 2 * self.ramp_samples
                )
            if self.control_end_index is None or self.control_end_index > self.specification_arguments.shape[-1]:
                self.control_end_index = self.specification_arguments.shape[-1]
            # Check to make sure if we are controlling to a subset of the tones that we
            # aren't starting with a bunch of zeros
            if self.control_tones is not None:
                first_nonzero_index = np.min(
                    np.argmax(
                        self.excitation_signal_amplitudes[self.control_tones] != 0,
                        axis=-1,
                    )
                )
                if first_nonzero_index > self.control_start_index:
                    self.control_start_index = first_nonzero_index
            control_slice = slice(self.control_start_index, self.control_end_index)

            # print('Initializing Control')
            # Call the control law to get the first of its signals
            self.control_first_signal = self.control_class.initialize_control(self.control_tones, self.control_start_index, self.control_end_index)

            # print('Constructing Control Arrays')
            # Construct the full arrays that we are controlling to
            self.control_specification_arguments = self.excitation_signal_arguments[self.control_tones, control_slice]

            # print('Setting Up Tracking Filters')
            # Set up the tracking filters to track amplitude and phase information
            self.control_filters = []
            if self.environment_metadata.tracking_filter_type == 0:
                self.control_block_size = self.hardware_metadata.samples_per_read
            else:
                self.control_block_size = self.environment_metadata.vk_filter_blocksize
            for signal in self.predicted_response_signals_combined:
                if self.environment_metadata.tracking_filter_type == 0:
                    generator = [
                        digital_tracking_filter_generator(
                            dt=1 / self.environment_metadata.sample_rate,
                            cutoff_frequency_ratio=self.environment_metadata.tracking_filter_cutoff,
                            filter_order=self.environment_metadata.tracking_filter_order,
                        )
                        for tone in self.control_specification_arguments
                    ]
                    for gen in generator:
                        gen.send(None)
                    self.control_filters.append(generator)
                else:
                    generator = vold_kalman_filter_generator(
                        sample_rate=self.environment_metadata.sample_rate,
                        num_orders=self.control_specification_arguments.shape[0],
                        block_size=self.control_block_size,
                        overlap=self.environment_metadata.vk_filter_overlap,
                        bandwidth=self.environment_metadata.vk_filter_bandwidth,
                        filter_order=self.environment_metadata.vk_filter_order,
                        buffer_size_factor=self.environment_metadata.buffer_blocks + 1,
                    )
                    generator.send(None)
                    self.control_filters.append(generator)

            # Set up empty warning and abort flags
            self.control_warning_flags = np.zeros(
                (self.control_specification_arguments.shape[0], n_control_channels),
                dtype=bool,
            )
            self.control_abort_flags = np.zeros(
                (self.control_specification_arguments.shape[0], n_control_channels),
                dtype=bool,
            )
            self.control_amplitude_errors = np.zeros((self.control_specification_arguments.shape[0], n_control_channels))

            # print('Setting up Signal Generation')
            # Set up the signal generation
            self.siggen_shutdown_achieved = False
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (
                    SignalGenerationCommands.INITIALIZE_PARAMETERS,
                    self.get_signal_generation_metadata(),
                ),
            )
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (
                    SignalGenerationCommands.INITIALIZE_SIGNAL_GENERATOR,
                    ContinuousTransientSignalGenerator(
                        num_samples_per_frame=self.hardware_metadata.samples_per_write,
                        num_signals=n_output_channels,
                        signal=self.control_first_signal,
                        last_signal=False,
                    ),
                ),
            )
            self.queue_container.signal_generation_command_queue.put(
                self.environment_name,
                (SignalGenerationCommands.SET_TEST_LEVEL, self.control_test_level),
            )
            # Tell the signal generation to start generating signals
            self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.GENERATE_SIGNALS, None))

            # print('Setting up last arguments')
            self.control_write_index = self.control_first_signal.shape[-1]
            self.control_read_index = 0
            self.control_analysis_index = 0
            self.control_finished = False
            self.control_analysis_finished = False
            self.control_response_signals_combined = []
            self.control_response_amplitudes = []
            self.control_response_phases = []
            self.control_drive_modifications = []
            self.control_response_frequencies = self.excitation_signal_frequencies[
                self.control_tones,
                self.control_start_index : self.control_end_index : self.hardware_metadata.output_oversample,
            ]
            self.control_response_arguments = self.excitation_signal_arguments[
                self.control_tones,
                self.control_start_index : self.control_end_index : self.hardware_metadata.output_oversample,
            ]
            self.control_target_phases = self.specification_phases[
                self.control_tones,
                :,
                self.control_start_index : self.control_end_index : self.hardware_metadata.output_oversample,
            ]
            self.control_target_amplitudes = self.specification_amplitudes[
                self.control_tones,
                :,
                self.control_start_index : self.control_end_index : self.hardware_metadata.output_oversample,
            ]
            self.control_startup = False
            self.set_active()
            self.gui_update_queue.put((self.environment_name, (UICommands.ENVIRONMENT_STARTED, None)))
        # See if any data has come in
        try:
            # print('Listening for Data')
            acquisition_data, last_acquisition = self.queue_container.data_in_queue.get_nowait()
            # print('Got Data')
            if last_acquisition:
                self.log("Acquired Last Data, Signal Generation Shutdown " f"Achieved: {self.siggen_shutdown_achieved}")
            else:
                self.log("Acquired Data")
            scale_factor = 0.0 if self.control_test_level < 1e-10 else 1 / self.control_test_level
            # print('Parsing Control and Excitation Data')
            control_data = acquisition_data[self.environment_metadata.control_channel_indices] * scale_factor
            if self.environment_metadata.response_transformation_matrix is not None:
                control_data = self.environment_metadata.response_transformation_matrix @ control_data
            excitation_data = acquisition_data[self.environment_metadata.output_channel_indices] * scale_factor
            if self.environment_metadata.reference_transformation_matrix is not None:
                excitation_data = self.environment_metadata.reference_transformation_matrix @ excitation_data
            block_size = acquisition_data.shape[-1]
            block_slice = slice(self.control_read_index, self.control_read_index + block_size)
            # print(f'Read Block Range {block_slice}')
            # Send the time data to the GUI
            # print('Sending Time Data')
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (
                        SineUICommands.TIME_DATA,
                        (
                            excitation_data[..., :: self.plot_downsample],
                            control_data[..., :: self.plot_downsample],
                        ),
                    ),
                )
            )
            self.control_response_signals_combined.append(control_data)
            # Find the time delay between the first signal and what we've
            # measured
            # print('Computing Time Delay')
            if self.control_first_signal is not None:
                first_signal = self.control_first_signal[..., :: self.hardware_metadata.output_oversample][..., : excitation_data.shape[-1]]
                reference_fft = np.fft.rfft(first_signal, axis=-1)
                this_fft = np.fft.rfft(excitation_data, axis=-1)
                freq = np.fft.rfftfreq(
                    first_signal.shape[-1],
                    1 / self.hardware_metadata.sample_rate,
                )
                good_lines = np.abs(reference_fft) / np.max(np.abs(reference_fft), axis=-1, keepdims=True) > self.good_line_threshold
                good_lines[..., 0] = False
                phase_difference = np.angle(this_fft / reference_fft)
                phase_slope = np.median(phase_difference[good_lines] / np.broadcast_to(freq, phase_difference.shape)[good_lines])
                self.control_time_delay = phase_slope / (2 * np.pi)
                # print(f"Time Delay: {self.control_time_delay=}")
                if DEBUG:
                    np.savez(
                        "debug_data/first_signal_sine_control.npz",
                        first_signal=self.control_first_signal,
                        excitation_data=excitation_data,
                    )
                self.control_first_signal = None
            # Analyze the recovered data
            if not self.control_analysis_finished:
                # print('Analyzing Data')
                achieved_signals = []
                achieved_amplitudes = []
                achieved_phases = []
                block_frequencies = self.control_response_frequencies[..., block_slice]
                block_arguments = self.control_response_arguments[..., block_slice]
                # print('Block Frequencies')
                # print(block_frequencies)
                # Check if this is the last control data we will be getting
                self.control_analysis_finished = block_slice.stop >= self.control_response_frequencies.shape[-1]
                # print(f'Is Last Data? {self.control_analysis_finished=}')
                # Truncate just in case we've gotten some extra data in the acquisition
                block_signal = control_data[..., : block_arguments.shape[-1]]
                # print('Filtering Data to extract amplitude and phase')
                start_time = time.time()
                if self.environment_metadata.tracking_filter_type == 0:
                    for signal, tone_filters in zip(block_signal, self.control_filters):
                        amps = []
                        phss = []
                        for tone_argument, tone_frequency, tone_filter in zip(block_arguments, block_frequencies, tone_filters):
                            amp, phs = tone_filter.send((signal, tone_frequency, tone_argument))
                            amps.append(amp)
                            phss.append(phs)  # Radians
                        achieved_amplitudes.append(np.array(amps))
                        achieved_phases.append(np.array(phss))
                        achieved_signals.append(np.array(amps) * np.cos(block_arguments + np.array(phss)))  # Radians
                else:
                    for signal, vk_filter in zip(block_signal, self.control_filters):
                        vk_signals, vk_amplitudes, vk_phases = vk_filter.send((signal, block_arguments, self.control_analysis_finished))
                        achieved_amplitudes.append(vk_amplitudes)
                        achieved_phases.append(vk_phases)  # Radians
                        achieved_signals.append(vk_signals)

                finish_time = time.time()
                self.log(f"Signal filtering achieved in {finish_time - start_time:0.2f}s.")
                achieved_signals = np.array(achieved_signals)
                achieved_amplitudes = np.array(achieved_amplitudes)
                achieved_phases = np.array(achieved_phases)
                if not np.all(achieved_signals == None):  # noqa: E711 # pylint: disable=singleton-comparison
                    # print('Got Amplitude and Phase Data')
                    block_start = self.control_analysis_index
                    block_end = self.control_analysis_index + achieved_signals.shape[-1]
                    block_slice = slice(block_start, block_end)
                    # print(f'Analysis Block Range {block_slice}')
                    achieved_frequencies = self.control_response_frequencies[..., block_slice]
                    # print('Analysis Frequencies')
                    # print(achieved_frequencies)
                    # print(f'Achieved Frequency Size {achieved_frequencies.shape=}')
                    achieved_signals = achieved_signals.transpose(1, 0, 2)
                    self.log(f"Analyzing Dataset With Size {achieved_signals.shape}")
                    # print(f'Achieved Signals Size {achieved_signals.shape=}')
                    achieved_amplitudes = achieved_amplitudes.transpose(1, 0, 2)
                    # print(f'Achieved Amplitudes Size {achieved_signals.shape=}')
                    # Correct for time delay on the phases, newaxis to broadcast across signals
                    achieved_phases = (
                        achieved_phases.transpose(1, 0, 2) - self.control_time_delay * 2 * np.pi * achieved_frequencies[:, np.newaxis, :]
                    )
                    # print(f'Achieved Phases Size {achieved_phases.shape=}')
                    # Here I want to do the best fit to the phases, need to compare phase
                    # achieved vs phase desired
                    if self.environment_metadata.phase_fit:
                        # print('Fitting Phases')
                        target = self.control_target_amplitudes[..., block_slice] * np.exp(1j * self.control_target_phases[..., block_slice])
                        achieved = achieved_amplitudes * np.exp(1j * achieved_phases)
                        phase_change_fit = np.angle(np.sum(target * achieved.conj()))
                        achieved_phases = achieved_phases + phase_change_fit
                    # print('Computing Drive Updates')
                    drive_modification = self.control_class.update_control(
                        achieved_signals,
                        achieved_amplitudes,
                        achieved_phases,  # Radians
                        achieved_frequencies,
                        self.control_time_delay,
                    )
                    self.control_drive_modifications.append(drive_modification)
                    # Need to develop data for the table.  First we need to pick out "valid"
                    # data, which will exclude the ramp-ups and ramp-downs
                    for tone_index in range(achieved_signals.shape[0]):
                        full_tone_index = self.control_tone_indices[tone_index]
                        compare_start = max(
                            (self.specification_start_indices[full_tone_index] - self.control_start_index)
                            // self.hardware_metadata.output_oversample,
                            block_start,
                            self.ramp_samples // self.hardware_metadata.output_oversample,
                        )
                        compare_end = min(
                            (self.specification_end_indices[full_tone_index] - self.control_start_index) // self.hardware_metadata.output_oversample,
                            block_end,
                            (self.control_end_index - self.control_start_index - self.ramp_samples) // self.hardware_metadata.output_oversample,
                        )
                        if compare_start >= compare_end:
                            continue
                        block_start_offset = compare_start - block_start
                        block_end_offset = compare_end - block_start
                        amplitudes = achieved_amplitudes[tone_index, :, block_start_offset:block_end_offset]
                        compare_amplitudes = self.control_target_amplitudes[tone_index, :, compare_start:compare_end]
                        compare_frequencies = self.control_response_frequencies[tone_index, compare_start:compare_end]
                        self.control_amplitude_errors[tone_index] = np.max(np.abs(scale2db(amplitudes / compare_amplitudes)), axis=-1)
                        if np.any(np.isinf(self.control_amplitude_errors[tone_index])):
                            self.log(f"Found Infinities:\n{amplitudes.shape=} " f"{compare_amplitudes.shape=}")
                            self.log(f"Infinity Frequencies: {compare_frequencies}")
                            self.log(f"Comparison Amplitudes: " f"{compare_amplitudes[np.isinf(self.control_amplitude_errors[tone_index])]}")
                        for channel_index in range(amplitudes.shape[0]):
                            compare_warnings = self.environment_metadata.specifications[full_tone_index].interpolate_warning(
                                channel_index, compare_frequencies
                            )
                            compare_aborts = self.environment_metadata.specifications[full_tone_index].interpolate_abort(
                                channel_index, compare_frequencies
                            )
                            warning_ratio = amplitudes[channel_index] / compare_warnings
                            abort_ratio = amplitudes[channel_index] / compare_aborts
                            if np.any(warning_ratio[0] < 1.0):
                                self.control_warning_flags[tone_index, channel_index] = True
                                self.log(
                                    f"Lower Warning at Tone {full_tone_index} Channel "
                                    f"{channel_index} Frequency "
                                    f"{compare_frequencies[warning_ratio[0] < 1.0]}"
                                )
                                self.log(f"Amplitudes: {amplitudes[channel_index, warning_ratio[0] < 1.0]}")
                                self.log(f"Warning Level: {compare_warnings[0, warning_ratio[0] < 1.0]}")
                            if np.any(warning_ratio[1] > 1.0):
                                self.control_warning_flags[tone_index, channel_index] = True
                                self.log(
                                    f"Upper Warning at Tone {full_tone_index} Channel "
                                    f"{channel_index} Frequency "
                                    f"{compare_frequencies[warning_ratio[1] > 1.0]}"
                                )
                                self.log(f"Amplitudes: {amplitudes[channel_index, warning_ratio[1] > 1.0]}")
                                self.log(f"Warning Level: {compare_warnings[1, warning_ratio[1] > 1.0]}")
                            if np.any(abort_ratio[0] < 1.0):
                                self.control_abort_flags[tone_index, channel_index] = True
                                self.log(
                                    f"Lower Abort at Tone {full_tone_index} Channel "
                                    f"{channel_index} Frequency "
                                    f"{compare_frequencies[abort_ratio[0] < 1.0]}"
                                )
                                self.log(f"Amplitudes: {amplitudes[channel_index, abort_ratio[0] < 1.0]}")
                                self.log(f"Abort Level: {compare_aborts[0, abort_ratio[0] < 1.0]}")
                            if np.any(abort_ratio[1] > 1.0):
                                self.control_abort_flags[tone_index, channel_index] = True
                                self.log(
                                    f"Upper Abort at Tone {full_tone_index} Channel {channel_index} "
                                    f"Frequency {compare_frequencies[abort_ratio[1] > 1.0]}"
                                )
                                self.log(f"Amplitudes: {amplitudes[channel_index, abort_ratio[1] > 1.0]}")
                                self.log(f"Abort Level: {compare_aborts[1, abort_ratio[1] > 1.0]}")
                    # print('Populating Full Block Data')
                    full_achieved_signals = np.zeros((self.specification_signals.shape[0],) + achieved_signals.shape[1:])
                    full_achieved_signals[self.control_tones] = achieved_signals
                    full_achieved_amplitudes = np.zeros((self.specification_amplitudes.shape[0],) + achieved_amplitudes.shape[1:])
                    full_achieved_amplitudes[self.control_tones] = achieved_amplitudes
                    full_achieved_phases = np.zeros((self.specification_phases.shape[0],) + achieved_phases.shape[1:])
                    full_achieved_phases[self.control_tones] = achieved_phases
                    full_achieved_frequencies = np.zeros((self.specification_frequencies.shape[0],) + achieved_frequencies.shape[1:])
                    full_achieved_frequencies[self.control_tones] = achieved_frequencies
                    full_drive_modification = np.zeros(
                        (self.specification_frequencies.shape[0],) + drive_modification.shape[1:],
                        dtype=complex,
                    )
                    full_drive_modification[self.control_tones] = drive_modification
                    full_achieved_amplitude_errors = np.zeros((self.specification_signals.shape[0],) + self.control_amplitude_errors.shape[1:])
                    full_achieved_amplitude_errors[self.control_tones] = self.control_amplitude_errors
                    full_achieved_warning_flags = np.zeros(
                        (self.specification_signals.shape[0],) + self.control_warning_flags.shape[1:],
                        dtype=bool,
                    )
                    full_achieved_warning_flags[self.control_tones] = self.control_warning_flags
                    full_achieved_abort_flags = np.zeros(
                        (self.specification_signals.shape[0],) + self.control_abort_flags.shape[1:],
                        dtype=bool,
                    )
                    full_achieved_abort_flags[self.control_tones] = self.control_abort_flags
                    self.control_response_amplitudes.append(achieved_amplitudes)
                    self.control_response_phases.append(achieved_phases)

                    # print('Sending Block Data to GUI')
                    self.gui_update_queue.put(
                        (
                            self.environment_name,
                            (
                                SineUICommands.CONTROL_DATA,
                                (
                                    full_achieved_signals[..., :: self.plot_downsample],
                                    full_achieved_amplitudes[..., :: self.plot_downsample],
                                    full_achieved_phases[..., :: self.plot_downsample] * 180 / np.pi,
                                    full_achieved_frequencies[..., :: self.plot_downsample],
                                    full_drive_modification,
                                    full_achieved_amplitude_errors,
                                    full_achieved_warning_flags,
                                    full_achieved_abort_flags,
                                ),
                            ),
                        )
                    )

                    self.control_analysis_index += achieved_signals.shape[-1]
                # Check if we're done analyzing control outputs and generate the next outputs
                if self.control_analysis_finished:
                    # print('Analysis is Finished')
                    self.log("Analysis Finished")
                    (
                        self.excitation_signals,
                        self.excitation_signal_frequencies,
                        self.excitation_signal_arguments,
                        self.excitation_signal_amplitudes,
                        self.excitation_signal_phases,  # Degrees
                        self.ramp_samples,
                    ) = self.control_class.finalize_control()
            self.control_read_index += acquisition_data.shape[-1]
            # Now we need to see if we need to write more data to the controller
            # This will be related to our current read index and write index for the control;
            # we don't want to run out of samples being generated on the output hardware
            # print('Checking if New Data is Needed')
            if (
                self.control_write_index // self.hardware_metadata.output_oversample
                < self.control_read_index + self.environment_metadata.buffer_blocks * self.hardware_metadata.samples_per_write
            ) and not self.control_finished:
                # print('Generating New Data')
                self.log("Generating New Data")
                excitation_signal, self.control_finished = self.control_class.generate_signal()
                # print('Data Generation Complete')
                if self.control_finished:
                    # print('Generated Last Data')
                    self.log("Control Finished")
                if DEBUG:
                    print("Writing Debug File")
                    num_files = len(glob(FILE_OUTPUT.format("*")))
                    np.savez(
                        FILE_OUTPUT.format(num_files),
                        excitation_signal=excitation_signal,
                        done_controlling=self.control_finished,
                    )
                self.log(f"Excitation Size: {np.sqrt(np.mean(excitation_signal**2))=}")
                self.queue_container.time_history_to_generate_queue.put((excitation_signal, self.control_finished))
                self.control_write_index += excitation_signal.shape[-1]
        except mp.queues.Empty:
            # print("Didn't Find Data")
            last_acquisition = False
        # See if we need to keep going
        if self.siggen_shutdown_achieved and last_acquisition:
            self.shutdown()
        else:
            self.queue_container.environment_command_queue.put(self.environment_name, (SineCommands.START_CONTROL, None))

    # region: Shutdown
    def shutdown(self):
        """Handles the environment after it has shut down"""
        self.log("Environment Shut Down")
        self.log(f"Before Flush: {self.queue_container.time_history_to_generate_queue.qsize()=}")
        flush_queue(self.queue_container.time_history_to_generate_queue, timeout=0.01)
        self.log(f"After Flush: {self.queue_container.time_history_to_generate_queue.qsize()=}")
        self.clear_active()
        self.gui_update_queue.put((self.environment_name, (UICommands.ENVIRONMENT_ENDED, None)))
        self.control_startup = True

    def stop_environment(self, data):
        """Sends a signal to start the shutdown process"""
        self.queue_container.signal_generation_command_queue.put(self.environment_name, (SignalGenerationCommands.START_SHUTDOWN, None))

    # region: Environment Commands
    def save_control_data(self, filename):
        """Saves the control data to a numpy file"""
        output_dict = {}
        for label in [
            "control_response_signals_combined",
            "control_response_amplitudes",
            "control_response_phases",
            "control_drive_modifications",
        ]:
            for index, array in enumerate(getattr(self, label)):
                output_dict[f"{label}_{index}"] = array
        for label in [
            "control_response_frequencies",
            "control_response_arguments",
            "control_target_phases",
            "control_target_amplitudes",
        ]:
            output_dict[label] = getattr(self, label)
        output_dict["sample_rate"] = self.hardware_metadata.sample_rate
        output_dict["output_oversample"] = self.hardware_metadata.output_oversample
        output_dict["names"] = [spec.name for spec in self.environment_metadata.specifications]
        np.savez(filename, **output_dict)

    def set_test_level(self, data):
        level = db2scale(data)
        self.control_test_level = level
        self.queue_container.gui_update_queue.put((self.environment_name, (SineCommands.SET_TEST_LEVEL, data)))


# region: Process
def sine_process(
    environment_name: str,
    queue_name: str,
    input_queue: VerboseMessageQueue,
    gui_update_queue: mp.Queue,
    controller_command_queue: VerboseMessageQueue,
    log_file_queue: mp.Queue,
    data_in_queue: mp.Queue,
    data_out_queue: mp.Queue,
    acquisition_active_event: mp.synchronize.Event,
    output_active_event: mp.synchronize.Event,
    active_event: mp.synchronize.Event,
    ready_event: mp.synchronize.Event,
    shutdown_event: mp.synchronize.Event,
    sysid_event: mp.synchronize.Event,
    threaded: bool,
):
    """A function to be used by multiprocessing to run the Sine environment.  It sets up
    the class and kicks off the run loop.

    Parameters
    ----------
    environment_name : str
        The name of the environment
    input_queue : VerboseMessageQueue
        A queue used to provide commands to the environment
    gui_update_queue : Queue
        A queue used to provide updates to the user interface from the environment
    controller_communication_queue : VerboseMessageQueue
        A queue used to communicate with the larger controller
    log_file_queue : Queue
        A queue used to handle logging
    data_in_queue : Queue
        A queue used to send data to the environment from the acqusition process
    data_out_queue : Queue
        A queue used to send data to the output process from the environment
    acquisition_active : int
        A multiprocessing value used as a flag to show when the acquisition is running
    output_active : int
        A multiprocessing value used as a flag to show when the output is running
    """
    try:
        if threaded:
            new_process = threading.Thread  # worker threads
        else:
            new_process = mp.Process  # worker processes

        # Create vibration queues
        queue_container = SineQueues(
            environment_name,
            input_queue,
            gui_update_queue,
            controller_command_queue,
            data_in_queue,
            data_out_queue,
            log_file_queue,
        )

        spectral_proc = new_process(
            target=spectral_processing_process,
            args=(
                environment_name,
                queue_container.spectral_command_queue,
                queue_container.data_for_spectral_computation_queue,
                queue_container.updated_spectral_quantities_queue,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.log_file_queue,
            ),
        )
        spectral_proc.start()
        analysis_proc = new_process(
            target=sysid_data_analysis_process,
            args=(
                environment_name,
                queue_container.data_analysis_command_queue,
                queue_container.updated_spectral_quantities_queue,
                queue_container.time_history_to_generate_queue,
                queue_container.environment_command_queue,
                queue_container.gui_update_queue,
                queue_container.log_file_queue,
            ),
        )
        analysis_proc.start()
        siggen_proc = new_process(
            target=signal_generation_process,
            args=(
                environment_name,
                queue_container.signal_generation_command_queue,
                queue_container.time_history_to_generate_queue,
                queue_container.data_out_queue,
                queue_container.environment_command_queue,
                queue_container.log_file_queue,
                queue_container.gui_update_queue,
            ),
        )
        siggen_proc.start()
        collection_proc = new_process(
            target=data_collector_process,
            args=(
                environment_name,
                queue_container.collector_command_queue,
                queue_container.data_in_queue,
                [queue_container.data_for_spectral_computation_queue],
                queue_container.environment_command_queue,
                queue_container.log_file_queue,
                queue_container.gui_update_queue,
            ),
        )
        collection_proc.start()

        process_class = SineEnvironment(
            environment_name,
            queue_name,
            queue_container,
            acquisition_active_event,
            output_active_event,
            active_event,
            ready_event,
            sysid_event,
        )
        process_class.run(shutdown_event)

        # Rejoin all the processes
        process_class.log("Joining Subprocesses")
        process_class.log("Joining Spectral Computation")
        spectral_proc.join()
        process_class.log("Joining Data Analysis")
        analysis_proc.join()
        process_class.log("Joining Signal Generation")
        siggen_proc.join()
        process_class.log("Joining Data Collection")
        collection_proc.join()
    except Exception:  # pylint: disable=broad-exception-caught
        print(traceback.format_exc())
