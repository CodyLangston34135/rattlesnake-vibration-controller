from rattlesnake.utilities import RattlesnakeError, GlobalCommands
from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.hardware.hardware_registry import HARDWARE_METADATA
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.environment.environment_registry import ENVIRONMENT_METADATA, ENVIRONMENT_COMMANDS
from rattlesnake.profile_manager import ProfileEvent
import netCDF4
import openpyxl


# region: Channel Table
def load_channel_table_from_netcdf(filepath):
    dataset = netCDF4.Dataset(filepath)
    channel_table = dataset["channels"]
    channel_list = []
    num_channels = dataset.dimensions["response_channels"].size
    channel_attr_list = Channel().channel_attr_list
    for row_idx in range(num_channels):
        channel = Channel()
        for attr in channel_attr_list:
            value = channel_table[attr][row_idx]
            value = None if isinstance(value, str) and not value.strip() else value
            setattr(channel, attr, value)

        if not channel.is_empty:  # optional safety check
            channel_list.append(channel)

    return channel_list


def load_channel_table_from_worksheet(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheets = workbook.sheetnames

    if len(sheets) > 1:
        sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
    if len(sheets) > 1:
        raise RattlesnakeError("Multiple channel table sheets located in Excel Spreadsheet")
    if len(sheets) == 0:
        raise RattlesnakeError("Excel Spreadsheet does not contain a channel table sheet")

    worksheet = workbook[sheets[0]]

    channel_list = []
    channel_attr_list = Channel().channel_attr_list
    for row in worksheet.iter_rows(min_row=3, min_col=2, max_col=23):
        channel = Channel()
        for col, cell in enumerate(row):
            value = cell.value
            value = None if isinstance(value, str) and not value.strip() else value
            setattr(channel, channel_attr_list[col], cell.value)
        if channel.is_empty:
            break
        channel_list.append(channel)
    workbook.close()

    return channel_list


def save_channel_table_worksheet(filepath, channel_list):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Channel Table"
    worksheet.cell(row=1, column=2, value="Test Article Definition")
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    worksheet.cell(row=1, column=5, value="Instrument Definition")
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    worksheet.cell(row=1, column=12, value="Channel Definition")
    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
    worksheet.cell(row=1, column=20, value="Output Feedback")
    worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
    worksheet.cell(row=1, column=22, value="Limits")
    worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
    for col_idx, val in enumerate(
        [
            "Channel Index",
            "Node Number",
            "Node Direction",
            "Comment",
            "Serial Number",
            "Triax DoF",
            "Sensitivity (mV/EU)",
            "Engineering Unit",
            "Make",
            "Model",
            "Calibration Exp Date",
            "Physical Device",
            "Physical Channel",
            "Type",
            "Minimum Value (V)",
            "Maximum Value (V)",
            "Coupling",
            "Current Excitation Source",
            "Current Excitation Value",
            "Physical Device",
            "Physical Channel",
            "Warning Level (EU)",
            "Abort Level (EU)",
        ]
    ):
        worksheet.cell(row=2, column=1 + col_idx, value=val)
    # Fill out values
    channel_attr_list = Channel().channel_attr_list
    for row, channel in enumerate(channel_list):
        row_idx = row + 3
        worksheet.cell(row=row_idx, column=1, value=row)
        for col, attr in enumerate(channel_attr_list):
            col_idx = col + 2
            val = getattr(channel, attr)
            val = str(val) if val is not None else ""
            worksheet.cell(row=row_idx, column=col_idx, value=val)

    workbook.save(filepath)


# region: Profiles
def load_profile_from_worksheet(filepath, environment_types):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    profile_sheet = workbook["Test Profile"]
    index = 2
    profile_event_list = []
    while True:
        timestamp = profile_sheet.cell(index, 1).value
        if timestamp is None or (isinstance(timestamp, str) and timestamp.strip() == ""):
            break
        timestamp = float(timestamp)

        environment_name = profile_sheet.cell(index, 2).value
        environment_type = environment_types[environment_name]

        # I have to conver the command string to an actual command
        command = profile_sheet.cell(index, 3).value
        command = str(command).upper().strip().replace(" ", "_")
        if command in GlobalCommands.__members__:
            command = GlobalCommands[command]
        elif command in ENVIRONMENT_COMMANDS[environment_type].__members__:
            command = ENVIRONMENT_COMMANDS[environment_type][command]
        else:
            raise RattlesnakeError(f"Invalid command: {command} for {environment_name} | {environment_type}")

        data = profile_sheet.cell(index, 4).value
        data = None if isinstance(data, str) and not data.strip() else data

        event = ProfileEvent(timestamp, environment_name, command, data)
        profile_event_list.append(event)
        index += 1
    workbook.close()

    return profile_event_list


def save_profile_worksheet(filepath, profile_event_list):
    workbook = openpyxl.Workbook()
    profile_sheet = workbook.active
    profile_sheet.title = "Test Profile"
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")
    # Fill out values
    if profile_event_list:
        for row, event in enumerate(profile_event_list):
            row_idx = row + 2
            profile_sheet.cell(row_idx, 1, str(event.timestamp))
            profile_sheet.cell(row_idx, 2, event.environment_name)
            profile_sheet.cell(row_idx, 3, event.command.label)
            profile_sheet.cell(row_idx, 4, str(event.data))
    workbook.save(filepath)


# region: Full Template
def load_metadata_from_netcdf(filepath):
    """Loads a test file using a file dialog"""
    dataset = netCDF4.Dataset(filepath)

    # Channel Table
    channel_list = load_channel_table_from_netcdf(filepath)

    # Hardware

    hardware_type = HardwareType(dataset.hardware)
    channel_list = channel_list
    sample_rate = int(dataset.sample_rate)
    time_per_read = float(dataset.time_per_read)
    time_per_write = float(dataset.time_per_write)
    output_oversample = int(dataset.output_oversample)

    hardware_metadata_class = HARDWARE_METADATA[hardware_type]
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            hardware_file = dataset.hardware_file
            hardware_metadata = hardware_metadata_class(channel_list, sample_rate, time_per_read, time_per_write, output_oversample, hardware_file)

    # Environments
    environment_metadata_list = []
    for environment_index, environment_name in enumerate(
        dataset.variables["environment_names"][...],
    ):
        environment_active_channels = dataset.variables["environment_active_channels"][:, environment_index]

        # Discover environment type
        environment_group = dataset.groups[environment_name]

        try:
            environment_type_int = dataset.variables["environment_types"][environment_index]
            environment_type = ControlTypes(environment_type_int)
        except:
            environment_type = discover_environment_type_in_old_netcdf(environment_group)

        environment_metadata_class = ENVIRONMENT_METADATA[environment_type]
        channel_list_bools = environment_active_channels
        environment_metadata = environment_metadata_class.retrieve_metadata_from_netcdf(
            environment_group, environment_name, channel_list_bools, sample_rate
        )
        environment_metadata_list.append(environment_metadata)

    return (hardware_metadata, environment_metadata_list)


def discover_environment_type_in_old_netcdf(environment_group):
    if hasattr(environment_group, "cancel_rampdown_time"):
        return ControlTypes.TIME
    else:
        raise RattlesnakeError("Invalid netcdf4 file")


def load_metadata_from_worksheet(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)

    # Channel table
    channel_list = load_channel_table_from_worksheet(filepath)

    # Hardware
    hardware_sheet = workbook["Hardware"]
    for row in hardware_sheet.rows:
        name = str(row[0].value).lower().strip().replace(" ", "_")
        value = row[1].value
        if value is None or value == "":
            continue
        match name:
            case "hardware_type":
                hardware_type_int = int(value)
            case "hardware_file":
                hardware_file = value
            case "sample_rate":
                sample_rate = value
            case "time_per_read":
                time_per_read = value
            case "time_per_write":
                time_per_write = value
            case "integration_oversampling":
                output_oversample = int(value)
            case "task_trigger":
                task_trigger = int(value)
            case "task_trigger_output_channel":
                task_output = str(value)
            case "maximum_acquisition_processes":
                maximum_acquisition_processes = int(value)
            case "":
                continue
            case _:
                print(f"Hardware sheet entry {row[0].value} not recognized")

    hardware_type = HardwareType(hardware_type_int)

    channel_list = channel_list
    sample_rate = int(sample_rate)
    time_per_read = float(time_per_read)
    time_per_write = float(time_per_write)

    hardware_metadata_class = HARDWARE_METADATA[hardware_type]
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            hardware_file = hardware_file
            output_oversample = output_oversample

            hardware_metadata = hardware_metadata_class(channel_list, sample_rate, time_per_read, time_per_write, output_oversample, hardware_file)

    # Environment
    environment_names = []
    environment_channel_list_bools = {}
    sheets = workbook.sheetnames
    if len(sheets) > 1:
        sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
    channel_sheet = workbook[sheets[0]]
    col = 24
    num_channels = len(channel_list)
    while True:
        environment_name = channel_sheet.cell(row=2, column=col).value

        # Stop if empty or None
        if environment_name is None or str(environment_name).strip() == "":
            break

        # Build environment channel list
        environment_active_channels = [False] * num_channels
        for i in range(num_channels):
            row = 3 + i
            value = channel_sheet.cell(row=row, column=col).value

            if value is not None and str(value).strip() != "":
                environment_active_channels[i] = True

        environment_names.append(environment_name)
        environment_channel_list_bools[environment_name] = environment_active_channels
        col += 1

    environment_metadata_list = []
    environment_types = {"Global": "Global"}
    for environment_name in environment_names:
        environment_sheet = workbook[environment_name]
        environment_type_name = environment_sheet.cell(row=1, column=2).value
        environment_type_name = str(environment_type_name).upper()
        environment_type = ControlTypes[environment_type_name]
        environment_types[environment_name] = environment_type

        environment_metadata_class = ENVIRONMENT_METADATA[environment_type]
        channel_list_bools = environment_channel_list_bools[environment_name]
        environment_metadata = environment_metadata_class.retrieve_metadata_from_worksheet(
            environment_sheet, environment_name, channel_list_bools, sample_rate
        )
        environment_metadata_list.append(environment_metadata)

    profile_sheet = workbook["Test Profile"]
    index = 2
    profile_event_list = []
    while True:
        timestamp = profile_sheet.cell(index, 1).value
        if timestamp is None or (isinstance(timestamp, str) and timestamp.strip() == ""):
            break
        timestamp = float(timestamp)

        environment_name = profile_sheet.cell(index, 2).value
        environment_type = environment_types[environment_name]

        # I have to conver the command string to an actual command
        command = profile_sheet.cell(index, 3).value
        command = str(command).upper().strip().replace(" ", "_")
        if command in GlobalCommands.__members__:
            command = GlobalCommands[command]
        elif command in ENVIRONMENT_COMMANDS[environment_type].__members__:
            command = ENVIRONMENT_COMMANDS[environment_type][command]
        else:
            raise RattlesnakeError(f"Invalid command: {command} for {environment_name} | {environment_type}")

        data = profile_sheet.cell(index, 4).value
        data = None if isinstance(data, str) and not data.strip() else data

        event = ProfileEvent(timestamp, environment_name, command, data)
        profile_event_list.append(event)
        index += 1

    workbook.close()

    return (hardware_metadata, environment_metadata_list, profile_event_list)


def save_rattlesnake_template(filepath, hardware_metadata=None, environment_metadata_list=None, profile_event_list=None):
    workbook = openpyxl.Workbook()
    # Channel Table
    worksheet = workbook.active
    worksheet.title = "Channel Table"
    worksheet.cell(row=1, column=2, value="Test Article Definition")
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    worksheet.cell(row=1, column=5, value="Instrument Definition")
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    worksheet.cell(row=1, column=12, value="Channel Definition")
    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
    worksheet.cell(row=1, column=20, value="Output Feedback")
    worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
    worksheet.cell(row=1, column=22, value="Limits")
    worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
    for col_idx, val in enumerate(
        [
            "Channel Index",
            "Node Number",
            "Node Direction",
            "Comment",
            "Serial Number",
            "Triax DoF",
            "Sensitivity (mV/EU)",
            "Engineering Unit",
            "Make",
            "Model",
            "Calibration Exp Date",
            "Physical Device",
            "Physical Channel",
            "Type",
            "Minimum Value (V)",
            "Maximum Value (V)",
            "Coupling",
            "Current Excitation Source",
            "Current Excitation Value",
            "Physical Device",
            "Physical Channel",
            "Warning Level (EU)",
            "Abort Level (EU)",
        ]
    ):
        worksheet.cell(row=2, column=1 + col_idx, value=val)
    # Fill out values
    channel_list = []
    if hardware_metadata:
        channel_list = hardware_metadata.channel_list
        channel_attr_list = Channel().channel_attr_list
        for row, channel in enumerate(channel_list):
            row_idx = row + 3
            worksheet.cell(row=row_idx, column=1, value=row)
            for col, attr in enumerate(channel_attr_list):
                col_idx = col + 2
                val = getattr(channel, attr)
                val = str(val) if val is not None else ""
                worksheet.cell(row=row_idx, column=col_idx, value=val)

    # Hardware
    hardware_worksheet = workbook.create_sheet("Hardware")
    hardware_worksheet.cell(1, 1, "Hardware Type")
    hardware_worksheet.cell(1, 2, "# Enter hardware index here")
    hardware_worksheet.cell(
        1,
        3,
        "Hardware Indices: 0 - NI DAQmx; 1 - LAN XI; 2 - Data Physics Quattro; "
        "3 - Data Physics 900 Series; 4 - Exodus Modal Solution; 5 - State Space Integration; "
        "6 - SDynPy System Integration",
    )
    hardware_worksheet.cell(2, 1, "Hardware File")
    hardware_worksheet.cell(
        2,
        3,
        "# Path to Hardware File (Depending on Hardware Device: 0 - Not Used; 1 - Not Used; "
        "2 - Path to DpQuattro.dll library file; 3 - Not Used; 4 - Path to Exodus Eigensolution; "
        "5 - Path to State Space File; 6 - Path to SDynPy system file)",
    )
    hardware_worksheet.cell(3, 1, "Sample Rate")
    hardware_worksheet.cell(3, 3, "# Sample Rate of Data Acquisition System")
    hardware_worksheet.cell(4, 1, "Time Per Read")
    hardware_worksheet.cell(4, 3, "# Number of seconds per Read from the Data Acquisition System")
    hardware_worksheet.cell(5, 1, "Time Per Write")
    hardware_worksheet.cell(5, 3, "# Number of seconds per Write to the Data Acquisition System")
    hardware_worksheet.cell(6, 1, "Maximum Acquisition Processes")
    hardware_worksheet.cell(
        6,
        3,
        "# Maximum Number of Acquisition Processes to start to pull data from hardware",
    )
    hardware_worksheet.cell(
        6,
        4,
        "Only Used by LAN-XI Hardware.  This row can be deleted if LAN-XI is not used",
    )
    hardware_worksheet.cell(7, 1, "Integration Oversampling")
    hardware_worksheet.cell(7, 3, "# For virtual control, an integration oversampling can be specified")
    hardware_worksheet.cell(
        7,
        3,
        "Only used for virtual control (Exodus, State Space, or SDynPy).  " "This row can be deleted if these are not used.",
    )
    hardware_worksheet.cell(8, 1, "Task Trigger")
    hardware_worksheet.cell(8, 3, "# Start trigger type")
    hardware_worksheet.cell(
        8,
        3,
        "Task Triggers: 0 - Internal, 1 - PFI0 with external trigger, 2 - PFI0 with Analog Output "
        "trigger.  Only used for NI hardware.  This row can be deleted if NI is not used.",
    )
    hardware_worksheet.cell(9, 1, "Task Trigger Output Channel")
    hardware_worksheet.cell(9, 3, "# Physical device and channel that generates a trigger signal")
    hardware_worksheet.cell(
        9,
        4,
        "Only used if Task Triggers is 2.  Only used for NI hardware.  " "This row can be deleted if it is not used.",
    )
    # Fill out values
    if hardware_metadata is not None and hardware_metadata.hardware_type != "Select":
        hardware_type = hardware_metadata.hardware_type
        hardware_worksheet.cell(1, 2, str(hardware_type.value))
        hardware_worksheet.cell(3, 2, str(hardware_metadata.sample_rate))
        hardware_worksheet.cell(4, 2, str(hardware_metadata.time_per_read))
        hardware_worksheet.cell(5, 2, str(hardware_metadata.time_per_write))
        match hardware_type:
            case HardwareType.SDYNPY_SYSTEM:
                hardware_worksheet.cell(2, 2, hardware_metadata.hardware_file)
                hardware_worksheet.cell(7, 2, hardware_metadata.output_oversample)

    # Environments
    worksheet.cell(row=1, column=24, value="Environments")
    # Fill out values
    if environment_metadata_list:
        for col, environment_metadata in enumerate(environment_metadata_list):
            col_idx = col + 24
            environment_name = environment_metadata.environment_name
            worksheet.cell(row=2, column=col_idx, value=environment_name)
            bool_indices = environment_metadata.map_channel_indices(channel_list)
            for row in bool_indices:
                row_idx = row + 3
                worksheet.cell(row=row_idx, column=col_idx, value="x")
            environment_worksheet = workbook.create_sheet(environment_name)
            environment_metadata.store_to_worksheet(environment_worksheet)

    # Profiles
    profile_sheet = workbook.create_sheet("Test Profile")
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")
    # Fill out values
    if profile_event_list:
        for row, event in enumerate(profile_event_list):
            row_idx = row + 2
            profile_sheet.cell(row_idx, 1, str(event.timestamp))
            profile_sheet.cell(row_idx, 2, event.environment_name)
            profile_sheet.cell(row_idx, 3, event.command.label)
            profile_sheet.cell(row_idx, 4, str(event.data))

    workbook.save(filepath)
