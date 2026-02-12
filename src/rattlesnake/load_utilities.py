from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.environment.environment_utilities import ControlTypes
import os
import openpyxl
import netCDF4
import numpy as np
import scipy.signal as sig
from scipy.interpolate import interp1d
from scipy.io import loadmat


def load_time_history(signal_path, sample_rate):
    """Loads a time history from a given file

    The signal can be loaded from numpy files (.npz, .npy) or matlab files (.mat).
    For .mat and .npz files, the time data can be included in the file in the
    't' field, or it can be excluded and the sample_rate input argument will
    be used.  If time data is specified, it will be linearly interpolated to the
    sample rate of the controller.
    For these file types, the signal should be stored in the 'signal'
    field.  For .npy files, only one array is stored, so it is treated as the
    signal, and the sample_rate input argument is used to construct the time
    data.

    Parameters
    ----------
    signal_path : str:
        Path to the file from which to load the time history

    sample_rate : str:
        The sample rate of the loaded signal.

    Returns
    -------
    signal : np.ndarray:
        A signal loaded from the file

    """
    _, extension = os.path.splitext(signal_path)
    if extension.lower() == ".npy":
        signal = np.load(signal_path)
    elif extension.lower() == ".npz":
        data = np.load(signal_path)
        signal = data["signal"]
        try:
            times = data["t"].squeeze()
            fn = interp1d(times, signal)
            abscissa = np.arange(0, max(times) + 1 / sample_rate - 1e-10, 1 / sample_rate)
            abscissa = abscissa[abscissa <= max(times)]
            signal = fn(abscissa)
        except KeyError:
            pass
    elif extension.lower() == ".mat":
        data = loadmat(signal_path)
        signal = data["signal"]
        try:
            times = data["t"].squeeze()
            fn = interp1d(times, signal)
            abscissa = np.arange(0, max(times) + 1 / sample_rate - 1e-10, 1 / sample_rate)
            abscissa = abscissa[abscissa <= max(times)]
            signal = fn(abscissa)
        except KeyError:
            pass
    else:
        raise ValueError(f"Could Not Determine the file type from the filename {signal_path}: {extension}")
    if signal.shape[-1] % 2 == 1:
        signal = signal[..., :-1]
    return signal


def load_channel_list_from_worksheet(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)
    sheets = workbook.sheetnames

    if len(sheets) > 1:
        sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
    if len(sheets) > 1:
        raise ValueError("Multiple channel table sheets located in Excel Spreadsheet")
    if len(sheets) == 0:
        raise ValueError("Excel Spreadsheet does not contain a channel table sheet")

    worksheet = workbook[sheets[0]]

    channel_list = []
    channel_attr_list = Channel().channel_attr_list
    for row in worksheet.iter_rows(min_row=3, max_col=23):
        channel = Channel()
        for col, cell in enumerate(row):
            setattr(channel, channel_attr_list[col], cell.value)
        if channel.is_empty:
            break
        channel_list.append(Channel)
    workbook.close()

    return channel_list


def load_metadata_from_netcdf4(filepath):
    """Loads a test file using a file dialog"""
    dataset = netCDF4.Dataset(filepath)  # pylint: disable=no-member

    # Channel Table
    channel_table = dataset["channels"]
    channel_list = []
    num_channels = dataset.dimensions["response_channels"].size
    channel_attr_list = Channel().channel_attr_list
    for row_idx in range(num_channels):
        channel = Channel()
        for attr in channel_attr_list:
            value = channel_table[attr][row_idx]
            setattr(channel, attr, value)

        if not channel.is_empty:  # optional safety check
            channel_list.append(channel)

    # Hardware
    hardware_type = HardwareType(dataset.hardware)
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            from rattlesnake.hardware.sdynpy_system import SDynPySystemMetadata

            hardware_metadata = SDynPySystemMetadata()
            hardware_metadata.hardware_file = dataset.hardware_file

        case _:
            raise ValueError(f"{hardware_type} has not been implemented yet")

    hardware_metadata.channel_list = channel_list
    hardware_metadata.sample_rate = dataset.sample_rate
    hardware_metadata.time_per_read = dataset.time_per_read
    hardware_metadata.time_per_write = dataset.time_per_write
    hardware_metadata.output_oversample = dataset.output_oversample

    # Environments
    environment_metadata_list = []
    for environment_index, environment_name in enumerate(
        dataset.variables["environment_names"][...],
    ):
        environment_active_channels = dataset.variables["environment_active_channels"][:, environment_index]
        environment_channel_list = [channel for channel, channel_bool in zip(channel_list, environment_active_channels) if channel_bool == 1]
        environment_type_int = dataset.variables["environment_types"][environment_index]
        environment_type = ControlTypes(environment_type_int)
        environment_group = dataset.groups[environment_name]

        match environment_type:
            case ControlTypes.TIME:
                from rattlesnake.environment.time_environment import TimeMetadata

                environment_metadata = TimeMetadata(environment_name)
                environment_metadata.sample_rate = hardware_metadata.sample_rate  # This is rough
            case _:
                raise TypeError(f"{environment_type} has not been implemented yet")

        environment_metadata.channel_list = environment_channel_list
        environment_metadata.retrieve_metadata(environment_group)

        environment_metadata_list.append(environment_metadata)

    return (hardware_metadata, environment_metadata)


def load_metadata_from_worksheet(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)

    channel_list = load_channel_list_from_worksheet(filepath)

    hardware_sheet = workbook["Hardware"]
    hardware_type_int = int(hardware_sheet.rows[0][1].value)
    hardware_type = HardwareType(hardware_type_int)
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            from rattlesnake.hardware.sdynpy_system import SDynPySystemMetadata

            hardware_metadata = SDynPySystemMetadata()

    hardware_metadata.channel_list = channel_list
    for row in enumerate(hardware_sheet.rows[4:]):
        name = str(row[0].value).lower().strip().replace(" ", "_")
        value = row[1].value
        match name:
            case "hardware_file":
                hardware_metadata.hardware_file = value
            case "sample_rate":
                hardware_metadata.sample_rate = value
            case "time_per_read":
                hardware_metadata.time_per_read = value
            case "time_per_write":
                hardware_metadata.time_per_write = value
            case "integration_oversampling":
                hardware_metadata.output_oversample = int(value)
            case "task_trigger":
                hardware_metadata.task_trigger = int(value)
            case "task_trigger_output_channel":
                hardware_metadata.task_output = str(value)
            case "maximum_acquisition_processes":
                hardware_metadata.maximum_acquisition_processes = int(value)
            case "":
                continue
            case _:
                print(f"Hardware sheet entry {row[0].value} not recognized")
