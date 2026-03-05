"""
Defines data analysis performed for environments that use system identification

Abstract environment that can be used to create new environment control strategies
in the controller that use system identification.

Rattlesnake Vibration Control Software
Copyright (C) 2021  National Technology & Engineering Solutions of Sandia, LLC
(NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.process.abstract_message_process import AbstractMessageProcess
from rattlesnake.utilities import VerboseMessageQueue, flush_queue
import openpyxl
import multiprocessing as mp
import numpy as np
import netCDF4 as nc4
from enum import Enum


class SysIdDataAnalysisCommands(Enum):
    """Valid commands to send to the data analysis process of an environment using system id"""

    INITIALIZE_PARAMETERS = 0
    RUN_NOISE = 1
    RUN_TRANSFER_FUNCTION = 2
    START_SHUTDOWN = 3
    STOP_SYSTEM_ID = 4
    SHUTDOWN_ACHIEVED = 5
    SYSTEM_ID_NOISE_COMPLETE = 6
    SYSTEM_ID_COMPLETE = 7
    LOAD_TRANSFER_FUNCTION = 8
    LOAD_NOISE = 9


class SysIdDataAnalysisUICommands(Enum):
    NOISE_UPDATE = 1
    SYSID_UPDATE = 2
    NOISE_COMPLETED = 3
    TRANSFER_COMPLETED = 4


# region: Metadata
class SysIdMetadata:
    """Abstract class for storing metadata for an environment.

    This class is used as a storage container for parameters used by an
    environment.  It is returned by the environment UI's
    ``collect_environment_definition_parameters`` function as well as its
    ``initialize_environment`` function.  Various parts of the controller and
    environment will query the class's data members for parameter values.

    Classes inheriting from AbstractMetadata must define:
      1. store_to_netcdf - A function defining the way the parameters are
         stored to a netCDF file saved during streaming operations.
    """

    def __init__(
        self,
        sample_rate=None,
        sysid_frame_size=None,
        sysid_averaging_type=None,
        sysid_noise_averages=None,
        sysid_averages=None,
        sysid_exponential_averaging_coefficient=None,
        sysid_estimator=None,
        sysid_level=None,
        sysid_level_ramp_time=None,
        sysid_signal_type=None,
        sysid_window=None,
        sysid_overlap=None,
        sysid_burst_on=None,
        sysid_pretrigger=None,
        sysid_burst_ramp_fraction=None,
        sysid_low_frequency_cutoff=None,
        sysid_high_frequency_cutoff=None,
        stream_file=None,
        auto_shutdown=False,
    ):
        self.sample_rate = sample_rate
        self.sysid_frame_size = sysid_frame_size
        self.sysid_averaging_type = sysid_averaging_type
        self.sysid_noise_averages = sysid_noise_averages
        self.sysid_averages = sysid_averages
        self.sysid_exponential_averaging_coefficient = sysid_exponential_averaging_coefficient
        self.sysid_estimator = sysid_estimator
        self.sysid_level = sysid_level
        self.sysid_level_ramp_time = sysid_level_ramp_time
        self.sysid_signal_type = sysid_signal_type
        self.sysid_window = sysid_window
        self.sysid_overlap = sysid_overlap
        self.sysid_burst_on = sysid_burst_on
        self.sysid_pretrigger = sysid_pretrigger
        self.sysid_burst_ramp_fraction = sysid_burst_ramp_fraction
        self.sysid_low_frequency_cutoff = sysid_low_frequency_cutoff
        self.sysid_high_frequency_cutoff = sysid_high_frequency_cutoff
        self.stream_file = stream_file
        self.auto_shutdown = auto_shutdown

    @property
    def sysid_frequency_spacing(self):
        """Frequency spacing in spectral quantities computed by system identification"""
        return self.sample_rate / self.sysid_frame_size

    @property
    def sysid_fft_lines(self):
        """Number of frequency lines in the FFT"""
        return self.sysid_frame_size // 2 + 1

    @property
    def sysid_skip_frames(self):
        """Number of frames to skip in the time stream due to ramp time"""
        return int(np.ceil(self.sysid_level_ramp_time * self.sample_rate / (self.sysid_frame_size * (1 - self.sysid_overlap))))

    def store_to_netcdf(self, netcdf_group_handle: nc4._netCDF4.Group):  # pylint: disable=c-extension-no-member
        """Store parameters to a group in a netCDF streaming file.

        This function stores parameters from the environment into the netCDF
        file in a group with the environment's name as its name.  The function
        will receive a reference to the group within the dataset and should
        store the environment's parameters into that group in the form of
        attributes, dimensions, or variables.

        This function is the "write" counterpart to the retrieve_metadata
        function in the AbstractUI class, which will read parameters from
        the netCDF file to populate the parameters in the user interface.

        Parameters
        ----------
        netcdf_group_handle : nc4._netCDF4.Group
            A reference to the Group within the netCDF dataset where the
            environment's metadata is stored.
        """
        netcdf_group_handle.sysid_frame_size = self.sysid_frame_size
        netcdf_group_handle.sysid_averaging_type = self.sysid_averaging_type
        netcdf_group_handle.sysid_noise_averages = self.sysid_noise_averages
        netcdf_group_handle.sysid_averages = self.sysid_averages
        netcdf_group_handle.sysid_exponential_averaging_coefficient = self.sysid_exponential_averaging_coefficient
        netcdf_group_handle.sysid_estimator = self.sysid_estimator
        netcdf_group_handle.sysid_level = self.sysid_level
        netcdf_group_handle.sysid_level_ramp_time = self.sysid_level_ramp_time
        netcdf_group_handle.sysid_signal_type = self.sysid_signal_type
        netcdf_group_handle.sysid_window = self.sysid_window
        netcdf_group_handle.sysid_overlap = self.sysid_overlap
        netcdf_group_handle.sysid_burst_on = self.sysid_burst_on
        netcdf_group_handle.sysid_pretrigger = self.sysid_pretrigger
        netcdf_group_handle.sysid_burst_ramp_fraction = self.sysid_burst_ramp_fraction
        netcdf_group_handle.sysid_low_frequency_cutoff = self.sysid_low_frequency_cutoff
        netcdf_group_handle.sysid_high_frequency_cutoff = self.sysid_high_frequency_cutoff

    @classmethod
    def retrieve_metadata_from_netcdf(cls, netcdf_group_handle: nc4._netCDF4.Group, hardware_metadata: HardwareMetadata):
        sample_rate = hardware_metadata.sample_rate
        sysid_frame_size = netcdf_group_handle.sysid_frame_size
        sysid_averaging_type = netcdf_group_handle.sysid_averaging_type
        sysid_noise_averages = netcdf_group_handle.sysid_noise_averages
        sysid_averages = netcdf_group_handle.sysid_averages
        sysid_exponential_averaging_coefficient = netcdf_group_handle.sysid_exponential_averaging_coefficient
        sysid_estimator = netcdf_group_handle.sysid_estimator
        sysid_level = netcdf_group_handle.sysid_level
        sysid_level_ramp_time = netcdf_group_handle.sysid_level_ramp_time
        sysid_signal_type = netcdf_group_handle.sysid_signal_type
        sysid_window = netcdf_group_handle.sysid_window
        sysid_overlap = netcdf_group_handle.sysid_overlap
        sysid_burst_on = netcdf_group_handle.sysid_burst_on
        sysid_pretrigger = netcdf_group_handle.sysid_pretrigger
        sysid_burst_ramp_fraction = netcdf_group_handle.sysid_burst_ramp_fraction
        sysid_low_frequency_cutoff = netcdf_group_handle.sysid_low_frequency_cutoff
        sysid_high_frequency_cutoff = netcdf_group_handle.sysid_high_frequency_cutoff
        return cls(
            sample_rate,
            sysid_frame_size,
            sysid_averaging_type,
            sysid_noise_averages,
            sysid_averages,
            sysid_exponential_averaging_coefficient,
            sysid_estimator,
            sysid_level,
            sysid_level_ramp_time,
            sysid_signal_type,
            sysid_window,
            sysid_overlap,
            sysid_burst_on,
            sysid_pretrigger,
            sysid_burst_ramp_fraction,
            sysid_low_frequency_cutoff,
            sysid_high_frequency_cutoff,
        )

    @staticmethod
    def create_blank_worksheet_template(worksheet: openpyxl.worksheet.worksheet.Worksheet, start_row):
        worksheet.cell(start_row, 1, "System ID Samples per Frame")
        worksheet.cell(
            start_row,
            3,
            "# Number of Samples per Measurement Frame in the System Identification",
        )
        worksheet.cell(start_row + 1, 1, "System ID Averaging:")
        worksheet.cell(start_row + 1, 3, "# Averaging Type, should be Linear or Exponential")
        worksheet.cell(start_row + 2, 1, "Noise Averages:")
        worksheet.cell(start_row + 2, 3, "# Number of Averages used when characterizing noise")
        worksheet.cell(start_row + 3, 1, "System ID Averages:")
        worksheet.cell(start_row + 3, 3, "# Number of Averages used when computing the FRF")
        worksheet.cell(start_row + 4, 1, "Exponential Averaging Coefficient:")
        worksheet.cell(start_row + 4, 3, "# Averaging Coefficient for Exponential Averaging (if used)")
        worksheet.cell(start_row + 5, 1, "System ID Estimator:")
        worksheet.cell(
            start_row + 5,
            3,
            "# Technique used to compute system ID.  Should be one of H1, H2, H3, or Hv.",
        )
        worksheet.cell(start_row + 6, 1, "System ID Level (V RMS):")
        worksheet.cell(
            start_row + 6,
            3,
            "# RMS Value of Flat Voltage Spectrum used for System Identification.",
        )
        worksheet.cell(start_row + 7, 1, "System ID Ramp Time")
        worksheet.cell(
            start_row + 7,
            3,
            "# Time for the system identification to ramp between levels or from start or to stop.",
        )
        worksheet.cell(start_row + 8, 1, "System ID Signal Type:")
        worksheet.cell(start_row + 8, 3, "# Signal to use for the system identification")
        worksheet.cell(start_row + 9, 1, "System ID Window:")
        worksheet.cell(
            start_row + 9,
            2,
            "# Window used to compute FRFs during system ID.  Should be one of Hann or None",
        )
        worksheet.cell(start_row + 10, 1, "System ID Overlap %:")
        worksheet.cell(start_row + 10, 3, "# Overlap to use in the system identification")
        worksheet.cell(start_row + 11, 1, "System ID Burst On %:")
        worksheet.cell(start_row + 11, 3, "# Percentage of a frame that the burst random is on for")
        worksheet.cell(start_row + 12, 1, "System ID Burst Pretrigger %:")
        worksheet.cell(
            start_row + 12,
            2,
            "# Percentage of a frame that occurs before the burst starts in a burst random signal",
        )
        worksheet.cell(start_row + 13, 1, "System ID Ramp Fraction %:")
        worksheet.cell(
            start_row + 13,
            3,
            '# Percentage of the "System ID Burst On %" that will be used to ramp up to full level',
        )

    def store_to_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int):
        if self.sysid_frame_size is not None:
            worksheet.cell(start_row, 2, self.sysid_frame_size)
        if self.sysid_averaging_type is not None:
            worksheet.cell(start_row + 1, 2, self.sysid_averaging_type)
        if self.sysid_noise_averages is not None:
            worksheet.cell(start_row + 2, 2, self.sysid_noise_averages)
        if self.sysid_averages is not None:
            worksheet.cell(start_row + 3, 2, self.sysid_averages)
        if self.sysid_exponential_averaging_coefficient is not None:
            worksheet.cell(start_row + 4, 2, self.sysid_exponential_averaging_coefficient)
        if self.sysid_estimator is not None:
            worksheet.cell(start_row + 5, 2, self.sysid_estimator)
        if self.sysid_level is not None:
            worksheet.cell(start_row + 6, 2, self.sysid_level)
        if self.sysid_level_ramp_time is not None:
            worksheet.cell(start_row + 7, 2, self.sysid_level_ramp_time)
        if self.sysid_signal_type is not None:
            worksheet.cell(start_row + 8, 2, self.sysid_signal_type)
        if self.sysid_window is not None:
            worksheet.cell(start_row + 9, 2, self.sysid_window)
        if self.sysid_overlap is not None:
            worksheet.cell(start_row + 10, 2, self.sysid_overlap * 100)
        if self.sysid_burst_on is not None:
            worksheet.cell(start_row + 11, 2, self.sysid_burst_on * 100)
        if self.sysid_pretrigger is not None:
            worksheet.cell(start_row + 12, 2, self.sysid_pretrigger * 100)
        if self.sysid_burst_ramp_fraction is not None:
            worksheet.cell(start_row + 13, 2, self.sysid_burst_ramp_fraction * 100)

    @classmethod
    def retrieve_metadata_from_worksheet(cls, worksheet: openpyxl.worksheet.worksheet.Worksheet, hardware_metadata: HardwareMetadata, start_row: int):
        sysid_frame_size = int(worksheet.cell(start_row, 2).value)
        sysid_averaging_type = worksheet.cell(start_row + 1, 2).value
        sysid_noise_averages = int(worksheet.cell(start_row + 2, 2).value)
        sysid_averages = int(worksheet.cell(start_row + 3, 2).value)
        sysid_exponential_averaging_coefficient = float(worksheet.cell(start_row + 4, 2).value)
        sysid_estimator = worksheet.cell(start_row + 5, 2).value
        sysid_level = float(worksheet.cell(start_row + 6, 2).value)
        sysid_level_ramp_time = float(worksheet.cell(start_row + 7, 2).value)
        sysid_signal_type = worksheet.cell(start_row + 8, 2).value
        sysid_window = worksheet.cell(start_row + 9, 2).value
        sysid_overlap = float(worksheet.cell(start_row + 10, 2).value) / 100
        sysid_burst_on = float(worksheet.cell(start_row + 11, 2).value) / 100
        sysid_pretrigger = float(worksheet.cell(start_row + 12, 2).value) / 100
        sysid_burst_ramp_fraction = float(worksheet.cell(start_row + 13, 2).value) / 100

        return cls(
            sample_rate=hardware_metadata.sample_rate,
            sysid_frame_size=sysid_frame_size,
            sysid_averaging_type=sysid_averaging_type,
            sysid_noise_averages=sysid_noise_averages,
            sysid_averages=sysid_averages,
            sysid_exponential_averaging_coefficient=sysid_exponential_averaging_coefficient,
            sysid_estimator=sysid_estimator,
            sysid_level=sysid_level,
            sysid_level_ramp_time=sysid_level_ramp_time,
            sysid_signal_type=sysid_signal_type,
            sysid_window=sysid_window,
            sysid_overlap=sysid_overlap,
            sysid_burst_on=sysid_burst_on,
            sysid_pretrigger=sysid_pretrigger,
            sysid_burst_ramp_fraction=sysid_burst_ramp_fraction,
            sysid_low_frequency_cutoff=0,
            sysid_high_frequency_cutoff=int(hardware_metadata.sample_rate / 2),
            stream_file=None,
            auto_shutdown=False,
        )

    def validate(self):
        return True

    @classmethod
    def default_metadata(cls, sample_rate):
        return cls(
            sample_rate=sample_rate,
            sysid_frame_size=sample_rate,
            sysid_averaging_type="Linear",
            sysid_noise_averages=20,
            sysid_averages=20,
            sysid_exponential_averaging_coefficient=0.01,
            sysid_estimator="H1",
            sysid_level=0.01,
            sysid_level_ramp_time=0.5,
            sysid_signal_type="Random",
            sysid_window="Hann",
            sysid_overlap=0.5,
            sysid_burst_on=0.5,
            sysid_pretrigger=0.05,
            sysid_burst_ramp_fraction=0.05,
            sysid_low_frequency_cutoff=0,
            sysid_high_frequency_cutoff=int(sample_rate / 2),
            stream_file=None,
            auto_shutdown=False,
        )

    def __eq__(self, other):
        try:
            return np.all([np.all(value == other.__dict__[field]) for field, value in self.__dict__.items()])
        except (AttributeError, KeyError):
            return False


class SysIdDataPackage:
    def __init__(
        self,
        frequencies,
        sysid_frf,
        sysid_coherence,
        sysid_response_cpsd,
        sysid_reference_cpsd,
        sysid_condition,
        sysid_response_noise,
        sysid_reference_noise,
    ):
        self.frequencies = frequencies
        self.sysid_frf = sysid_frf
        self.sysid_coherence = sysid_coherence
        self.sysid_response_cpsd = sysid_response_cpsd
        self.sysid_reference_cpsd = sysid_reference_cpsd
        self.sysid_condition = sysid_condition
        self.sysid_response_noise = sysid_response_noise
        self.sysid_reference_noise = sysid_reference_noise

    def store_to_netcdf(self, netcdf_group_handle: nc4._netCDF4.Group):
        netcdf_group_handle.createDimension("sysid_control_channels", self.sysid_frf.shape[1])
        netcdf_group_handle.createDimension("sysid_output_channels", self.sysid_frf.shape[2])
        netcdf_group_handle.createDimension("sysid_fft_lines", self.sysid_frf.shape[0])
        var = netcdf_group_handle.createVariable(
            "frf_data_real",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_frf.real
        var = netcdf_group_handle.createVariable(
            "frf_data_imag",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_frf.imag
        var = netcdf_group_handle.createVariable("frf_coherence", "f8", ("sysid_fft_lines", "sysid_control_channels"))
        var[...] = self.sysid_coherence.real
        var = netcdf_group_handle.createVariable(
            "response_cpsd_real",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_control_channels"),
        )
        var[...] = self.sysid_response_cpsd.real
        var = netcdf_group_handle.createVariable(
            "response_cpsd_imag",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_control_channels"),
        )
        var[...] = self.sysid_response_cpsd.imag
        var = netcdf_group_handle.createVariable(
            "reference_cpsd_real",
            "f8",
            ("sysid_fft_lines", "sysid_output_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_reference_cpsd.real
        var = netcdf_group_handle.createVariable(
            "reference_cpsd_imag",
            "f8",
            ("sysid_fft_lines", "sysid_output_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_reference_cpsd.imag
        var = netcdf_group_handle.createVariable(
            "response_noise_cpsd_real",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_control_channels"),
        )
        var[...] = self.sysid_response_noise.real
        var = netcdf_group_handle.createVariable(
            "response_noise_cpsd_imag",
            "f8",
            ("sysid_fft_lines", "sysid_control_channels", "sysid_control_channels"),
        )
        var[...] = self.sysid_response_noise.imag
        var = netcdf_group_handle.createVariable(
            "reference_noise_cpsd_real",
            "f8",
            ("sysid_fft_lines", "sysid_output_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_reference_noise.real
        var = netcdf_group_handle.createVariable(
            "reference_noise_cpsd_imag",
            "f8",
            ("sysid_fft_lines", "sysid_output_channels", "sysid_output_channels"),
        )
        var[...] = self.sysid_reference_noise.imag

    @classmethod
    def retrieve_from_netcdf(cls, netcdf_group_handle: nc4._netCDF4.Group, sample_rate: int):
        frame_size = netcdf_group_handle.sysid_frame_size
        fft_lines = netcdf_group_handle.dimensions["fft_lines"].size
        variables = netcdf_group_handle.variables
        combine = np.vectorize(complex)
        sysid_frf = np.array(combine(variables["frf_data_real"][:], variables["frf_data_imag"][:]))
        sysid_coherence = np.array(variables["frf_coherence"][:])
        sysid_response_cpsd = np.array(
            combine(
                variables["response_cpsd_real"][:],
                variables["response_cpsd_imag"][:],
            )
        )
        sysid_reference_cpsd = np.array(
            combine(
                variables["reference_cpsd_real"][:],
                variables["reference_cpsd_imag"][:],
            )
        )
        sysid_response_noise = np.array(
            combine(
                variables["response_noise_cpsd_real"][:],
                variables["response_noise_cpsd_imag"][:],
            )
        )
        sysid_reference_noise = np.array(
            combine(
                variables["reference_noise_cpsd_real"][:],
                variables["reference_noise_cpsd_imag"][:],
            )
        )
        sysid_condition = np.linalg.cond(sysid_frf)
        frequencies = np.arange(fft_lines) * sample_rate / frame_size

        return cls(
            frequencies,
            sysid_frf,
            sysid_coherence,
            sysid_response_cpsd,
            sysid_reference_cpsd,
            sysid_condition,
            sysid_response_noise,
            sysid_reference_noise,
        )


# region: Data Analysis
class SysIDAnalysisProcess(AbstractMessageProcess):
    """Process to perform data analysis and control calculations in an environment
    using system id"""

    def __init__(
        self,
        process_name: str,
        command_queue: VerboseMessageQueue,
        data_in_queue: mp.queues.Queue,
        data_out_queue: mp.queues.Queue,
        environment_command_queue: VerboseMessageQueue,
        log_file_queue: mp.queues.Queue,
        gui_update_queue: mp.queues.Queue,
        environment_name: str,
    ):
        """Initialize the environment process

        Parameters
        ----------
        process_name : str
            The name of the process
        command_queue : VerboseMessageQueue
            A queue used to send commands to this process
        data_in_queue : mp.queues.Queue
            A queue receiving frames of data from the data collector
        data_out_queue : mp.queues.Queue
            A queue to put the next output or analysis results for the environment to use
        environment_command_queue : VerboseMessageQueue
            A queue used to send commands to the main environment process
        log_file_queue : mp.queues.Queue
            A queue used to send log file strings
        gui_update_queue : mp.queues.Queue
            A queue used to send updates back to the graphical user interface
        environment_name : str
            The name of the environment owning this process
        """
        super().__init__(process_name, log_file_queue, command_queue, gui_update_queue)
        self.map_command(SysIdDataAnalysisCommands.INITIALIZE_PARAMETERS, self.initialize_sysid_parameters)
        self.map_command(SysIdDataAnalysisCommands.RUN_NOISE, self.run_sysid_noise)
        self.map_command(SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION, self.run_sysid_transfer_function)
        self.map_command(SysIdDataAnalysisCommands.STOP_SYSTEM_ID, self.stop_sysid)
        self.map_command(SysIdDataAnalysisCommands.LOAD_NOISE, self.load_sysid_noise)
        self.map_command(SysIdDataAnalysisCommands.LOAD_TRANSFER_FUNCTION, self.load_sysid_transfer_function)
        self.environment_name = environment_name
        self.environment_command_queue = environment_command_queue
        self.data_in_queue = data_in_queue
        self.data_out_queue = data_out_queue
        self.parameters = None
        self.frames = None
        self.frequencies = None
        self.sysid_frf = None
        self.sysid_coherence = None
        self.sysid_response_cpsd = None
        self.sysid_reference_cpsd = None
        self.sysid_response_noise = None
        self.sysid_reference_noise = None
        self.sysid_condition = None
        self.startup = True

    def initialize_sysid_parameters(self, data: SysIdMetadata):
        """Stores parameters describing the system identification into the object

        Parameters
        ----------
        data : AbstractSysIdMetadata
            A metadata object containing the parameters to define the system identification
        """
        self.parameters = data

    def load_sysid_noise(self, spectral_data):
        """Loads noise data from a previous system identification

        Parameters
        ----------
        spectral_data : tuple
            A tuple containing frames, frequencies, system id FRFs, coherence, response cpsd,
            reference_cpsd and condition number
        """
        self.log("Obtained Spectral Data")
        (
            self.frames,
            self.frequencies,
            _,
            _,
            self.sysid_response_noise,
            self.sysid_reference_noise,
            _,
        ) = spectral_data

    def load_sysid_transfer_function(self, spectral_data, skip_sysid=True):
        """Loads system ID data from a previous system identification

        Parameters
        ----------
        spectral_data : tuple
            A tuple containing frames, frequencies, system id FRFs, coherence, response cpsd,
            reference_cpsd and condition number
        skip_sysid : bool, optional
            If True, send the system identification complete flag to the controller. By default True
        """
        self.log("Obtained Spectral Data")
        (
            self.frames,
            self.frequencies,
            self.sysid_frf,
            self.sysid_coherence,
            self.sysid_response_cpsd,
            self.sysid_reference_cpsd,
            self.sysid_condition,
        ) = spectral_data
        if skip_sysid:
            self.environment_command_queue.put(
                self.process_name,
                (
                    SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
                    (
                        self.frames,
                        0,
                        self.frequencies,
                        self.sysid_frf,
                        self.sysid_coherence,
                        self.sysid_response_cpsd,
                        self.sysid_reference_cpsd,
                        self.sysid_condition,
                        self.sysid_response_noise,
                        self.sysid_reference_noise,
                    ),
                ),
            )

    def run_sysid_noise(self, auto_shutdown):
        """Starts and runs the system identification noise phase.

        Parameters
        ----------
        auto_shutdown : bool
            If True, the environment will automatically shut down when the requested number of
            frames is reached.  If False, the noise characterization will run until manually
            stopped.
        """
        if self.startup:
            self.startup = False
            self.frames = 0
        spectral_data = flush_queue(self.data_in_queue)
        if len(spectral_data) > 0:
            self.load_sysid_noise(spectral_data[-1])
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (
                        SysIdDataAnalysisUICommands.NOISE_UPDATE,
                        (
                            self.frames,
                            self.parameters.sysid_noise_averages,
                            self.frequencies,
                            self.sysid_response_noise,
                            self.sysid_reference_noise,
                        ),
                    ),
                )
            )
        if auto_shutdown and self.parameters.sysid_noise_averages == self.frames:
            self.environment_command_queue.put(
                self.process_name,
                (SysIdDataAnalysisCommands.START_SHUTDOWN, False),
            )
            self.stop_sysid(None)
            self.environment_command_queue.put(self.process_name, (SysIdDataAnalysisCommands.SYSTEM_ID_NOISE_COMPLETE, None))
        else:
            self.command_queue.put(self.process_name, (SysIdDataAnalysisCommands.RUN_NOISE, auto_shutdown))

    def run_sysid_transfer_function(self, auto_shutdown):
        """Starts and runs the system identification

        Parameters
        ----------
        auto_shutdown : bool
            If True, the system identification will stop automatically upon reaching the requested
            number of measurement frames.  If False, it will run indefinitely until manually
            stopped.
        """
        if self.startup:
            self.startup = False
            self.frames = 0
        spectral_data = flush_queue(self.data_in_queue)
        if len(spectral_data) > 0:
            self.load_sysid_transfer_function(spectral_data[-1], skip_sysid=False)
            self.gui_update_queue.put(
                (
                    self.environment_name,
                    (
                        SysIdDataAnalysisUICommands.SYSID_UPDATE,
                        (
                            self.frames,
                            self.parameters.sysid_averages,
                            self.frequencies,
                            self.sysid_frf,
                            self.sysid_coherence,
                            self.sysid_response_cpsd,
                            self.sysid_reference_cpsd,
                            self.sysid_condition,
                        ),
                    ),
                )
            )
        if auto_shutdown and self.parameters.sysid_averages == self.frames:
            self.environment_command_queue.put(
                self.process_name,
                (SysIdDataAnalysisCommands.START_SHUTDOWN, False),
            )
            self.stop_sysid(None)
            self.environment_command_queue.put(
                self.process_name,
                (
                    SysIdDataAnalysisCommands.SYSTEM_ID_COMPLETE,
                    (
                        self.frames,
                        self.parameters.sysid_averages,
                        self.frequencies,
                        self.sysid_frf,
                        self.sysid_coherence,
                        self.sysid_response_cpsd,
                        self.sysid_reference_cpsd,
                        self.sysid_condition,
                        self.sysid_response_noise,
                        self.sysid_reference_noise,
                    ),
                ),
            )
        else:
            self.command_queue.put(
                self.process_name,
                (SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION, auto_shutdown),
            )

    def stop_sysid(self, data):  # pylint: disable=unused-argument
        """Stops the currently running system identification phase

        Parameters
        ----------
        data : ignored
            This argument is not used, but is required by the calling signature of functions
            that get called via the command map.
        """
        # Remove any run_transfer_function or run_control from the queue
        instructions = self.command_queue.flush(self.process_name)
        for instruction in instructions:
            if not instruction[0] in [
                SysIdDataAnalysisCommands.RUN_NOISE,
                SysIdDataAnalysisCommands.RUN_TRANSFER_FUNCTION,
            ]:
                self.command_queue.put(self.process_name, instruction)
        flush_queue(self.data_out_queue)
        self.startup = True
        self.environment_command_queue.put(self.process_name, (SysIdDataAnalysisCommands.SHUTDOWN_ACHIEVED, None))


def sysid_data_analysis_process(
    environment_name: str,
    command_queue: VerboseMessageQueue,
    data_in_queue: mp.queues.Queue,
    data_out_queue: mp.queues.Queue,
    environment_command_queue: VerboseMessageQueue,
    gui_update_queue: mp.queues.Queue,
    log_file_queue: mp.queues.Queue,
    process_name=None,
):
    """An function called by multiprocessing to start up the system identification analysis
    process.

    Some environments may override the AbstractSysIDAnalysisProcess class and therefore should
    redefine this function to call that class.

    Parameters
    ----------
    environment_name : str
        The name of the environment
    command_queue : VerboseMessageQueue
        A queue used to send commands to this process
    data_in_queue : mp.queues.Queue
        A queue used to send frames of data and spectral quantities to the data analysis process
    data_out_queue : mp.queues.Queue
        A queue used to send control and analysis results back to the environment
    environment_command_queue : VerboseMessageQueue
        A queue used to send commands to the environment
    gui_update_queue : mp.queues.Queue
        A queue used to send updates to the graphical user interface
    log_file_queue : mp.queues.Queue
        A queue used to send log file messages
    process_name : _type_, optional
        A name for the process.  If not specified, it will be the environment name appended with
        Data Analysis.
    """
    data_analysis_instance = SysIDAnalysisProcess(
        environment_name + " Data Analysis" if process_name is None else process_name,
        command_queue,
        data_in_queue,
        data_out_queue,
        environment_command_queue,
        log_file_queue,
        gui_update_queue,
        environment_name,
    )

    data_analysis_instance.run()
