from rattlesnake.utilities import GlobalCommands, VerboseMessageQueue, QueueContainer, EventContainer, log_file_task, flush_queue
from rattlesnake.process.controller import controller_process
from rattlesnake.process.acquisition import acquisition_process
from rattlesnake.process.output import output_process
from rattlesnake.process.streaming import StreamMetadata, streaming_process
from rattlesnake.profile_manager import ProfileManager, ProfileEvent
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.abstract_environment import EnvironmentMetadata, EnvironmentInstructions
from rattlesnake.environment_manager import EnvironmentManager
from rattlesnake.load_utilities import load_channel_list_from_worksheet, load_channel_list_from_netcdf
from rattlesnake.hardware.hardware_utilities import Channel, HardwareType
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.environment.environment_registry import ENVIRONMENT_COMMANDS
import openpyxl
import netCDF4
import os
import time
import multiprocessing as mp
import threading
import queue as thqueue
from enum import Enum
from datetime import datetime
from typing import List

TASK_NAME = "Rattlesnake"
CLOSE_TIMEOUT = 5  # Number of seconds to wait for process to join
THREADING = True


# region: State
class RattlesnakeState(Enum):
    # We don't check for stored stream/profiles because they can be left blank
    INIT = 0  # Nothing is stored yet
    HARDWARE_STORE = 1  # Hardware has been stored
    ENVIRONMENT_STORE = 2  # Environments have been stored
    HARDWARE_ACTIVE = 3  # Acquisition is running
    ENVIRONMENT_ACTIVE = 4  # Environment output is running


# region: Rattlesnake
class Rattlesnake:
    def __init__(self, *, threaded: bool = THREADING, timeout: float = 30):
        # Initialize values for checking state
        self._threaded = threaded
        self._blocking = True  # Wait for ready events?, True for IDE, False for UI
        self._timeout = timeout  # Timeout while waiting for ready_events

        if self.threaded:
            new_queue = thqueue.Queue  # threading-safe in-memory queue
            new_process = threading.Thread  # worker threads
            new_event = threading.Event  # optional stop flag
        else:
            new_queue = mp.Queue  # multiprocessing queue
            new_process = mp.Process  # worker processes
            new_event = mp.Event  # optional stop flag

        # Start up log file process
        log_file_queue = mp.Queue()
        log_close_event = mp.Event()
        self.log_file_process = mp.Process(
            target=log_file_task,
            args=(log_file_queue, log_close_event),
        )
        self.log_file_process.start()

        # Start up command queues and processes
        self.controller_queue_name_manager = mp.Manager()  # Adds minor overhead which is reasonable for COMMAND queues only
        controller_close_event = new_event()
        controller_ready_event = new_event()
        controller_command_queue = VerboseMessageQueue(log_file_queue, new_queue(), self.controller_queue_name_manager, "Controller Command Queue")
        acquisition_close_event = new_event()
        acquisition_ready_event = new_event()
        acquisition_active_event = new_event()
        acquisition_active_event.clear()
        acquisition_command_queue = VerboseMessageQueue(log_file_queue, new_queue(), self.controller_queue_name_manager, "Acquisition Command Queue")
        output_close_event = new_event()
        output_ready_event = new_event()
        output_active_event = new_event()
        output_active_event.clear()
        output_command_queue = VerboseMessageQueue(log_file_queue, mp.Queue(), self.controller_queue_name_manager, "Output Command Queue")
        streaming_close_event = new_event()
        streaming_ready_event = new_event()
        streaming_active_event = new_event()
        streaming_active_event.clear()
        streaming_command_queue = VerboseMessageQueue(log_file_queue, new_queue(), self.controller_queue_name_manager, "Streaming Command Queue")

        # Set up data queue
        input_output_sync_queue = new_queue()
        single_process_hardware_queue = new_queue()

        # Set up environment queues
        max_environments = 16
        environment_close_events = {}
        environment_ready_events = {}
        environment_active_events = {}
        environment_command_queues = {}
        environment_data_in_queues = {}
        environment_data_out_queues = {}
        for env_idx in range(max_environments):
            queue_name = "Environment {:}".format(env_idx)
            environment_close_events[queue_name] = new_event()
            environment_ready_events[queue_name] = new_event()
            environment_active_events[queue_name] = new_event()
            environment_active_events[queue_name].clear()
            environment_command_queues[queue_name] = VerboseMessageQueue(
                log_file_queue, mp.Queue(), self.controller_queue_name_manager, queue_name + " Command Queue"
            )
            environment_data_in_queues[queue_name] = new_queue()
            environment_data_out_queues[queue_name] = new_queue()

        # Set up output queue
        gui_update_queue = new_queue()

        # Build queue container
        self.queue_container = QueueContainer(
            controller_command_queue,
            acquisition_command_queue,
            output_command_queue,
            streaming_command_queue,
            log_file_queue,
            input_output_sync_queue,
            single_process_hardware_queue,
            gui_update_queue,
            environment_command_queues,
            environment_data_in_queues,
            environment_data_out_queues,
        )
        self.event_container = EventContainer(
            controller_ready_event,
            acquisition_ready_event,
            output_ready_event,
            streaming_ready_event,
            environment_ready_events,
            log_close_event,
            controller_close_event,
            acquisition_close_event,
            output_close_event,
            streaming_close_event,
            environment_close_events,
            acquisition_active_event,
            output_active_event,
            streaming_active_event,
            environment_active_events,
        )

        # Controller
        self.controller_proc = new_process(
            target=controller_process,
            args=(
                self.queue_container,
                self.event_container.acquisition_active_event,
                self.event_container.output_active_event,
                self.event_container.streaming_active_event,
                self.event_container.environment_active_events,
                self.event_container.controller_ready_event,
                self.event_container.controller_close_event,
            ),
        )
        self.controller_proc.start()
        # Acquisition
        self.acquisition_proc = new_process(
            target=acquisition_process,
            args=(
                self.queue_container,
                self.event_container.acquisition_active_event,
                self.event_container.streaming_active_event,
                self.event_container.acquisition_ready_event,
                self.event_container.acquisition_close_event,
            ),
        )
        self.acquisition_proc.start()
        # Output
        self.output_proc = new_process(
            target=output_process,
            args=(
                self.queue_container,
                self.event_container.output_active_event,
                self.event_container.output_ready_event,
                self.event_container.output_close_event,
            ),
        )
        self.output_proc.start()
        # Streaming
        self.streaming_proc = new_process(
            target=streaming_process,
            args=(
                self.queue_container,
                self.event_container.streaming_ready_event,
                self.event_container.streaming_close_event,
            ),
        )
        self.streaming_proc.start()

        # Set up managers that will setup processes and store metadata
        self.environment_manager = EnvironmentManager(  # Contains hardware/environment metadata
            self.queue_container,
            self.event_container,
            self.threaded,
        )
        self.profile_manager = ProfileManager(self.queue_container)  # Contains instructions/profile events
        self.hardware_metadata = None
        self.environment_metadata = {}
        # These are only used for UI to pull from if they have already been set to the controller. These
        # are not used for any logic in this controller
        self.last_stream_metadata = None
        self.last_profile_event_list = []

        if self.blocking:
            ready_event_list = [
                self.event_container.controller_ready_event,
                self.event_container.acquisition_ready_event,
                self.event_container.output_ready_event,
                self.event_container.streaming_ready_event,
                *self.environment_manager.ready_event_list,
            ]
            active_event_list = []
            self.wait_for_events(ready_event_list, active_event_list)

    # region: Properties
    @property
    def state(self) -> RattlesnakeState:
        hardware_store = self.hardware_metadata is not None
        environment_store = self.environment_metadata != {}
        acquisition_active = self.event_container.acquisition_active_event.is_set()
        output_active = self.event_container.output_active_event.is_set()
        environment_active = any(event.is_set() for event in self.event_container.environment_active_events.values())

        if hardware_store and environment_store and acquisition_active and output_active and environment_active:
            return RattlesnakeState.ENVIRONMENT_ACTIVE

        if hardware_store and environment_store and acquisition_active and output_active:
            return RattlesnakeState.HARDWARE_ACTIVE

        if hardware_store and environment_store:
            return RattlesnakeState.ENVIRONMENT_STORE

        if hardware_store:
            return RattlesnakeState.HARDWARE_STORE

        return RattlesnakeState.INIT

    @property
    def streaming(self):
        return self.event_container.streaming_active_event.is_set()

    @property
    def threaded(self):
        return self._threaded

    @property
    def blocking(self):
        return self._blocking

    @property
    def timeout(self):
        return self._timeout

    @property
    def has_streamed(self):
        if self.last_stream_metadata:
            return True
        return False

    @property
    def has_profile(self):
        if self.last_profile_event_list:
            return True
        return False

    def set_blocking(self):
        self._blocking = True

    def clear_blocking(self):
        self._blocking = False

    def wait_for_events(
        self, ready_event_list: List[mp.synchronize.Event], active_event_list: List[mp.synchronize.Event], *, active_event_check: bool = None
    ):
        start_time = time.time()

        while True:
            ready_ok = all(event.is_set() for event in ready_event_list)
            active_ok = all(event.is_set() == active_event_check for event in active_event_list)

            if ready_ok and active_ok:
                return

            if self.timeout is not None and (time.time() - start_time) >= self.timeout:
                for event in ready_event_list:
                    event.set()
                raise TimeoutError("Timeout waiting for all events to be ready")

    # region: Loading
    def load_data_from_file(self, filepath):
        filename, filetype = os.path.splitext(filepath)

        if not os.access(filepath, os.R_OK):
            raise PermissionError("You do not have permissions to open {filepath}")

        # I force blocking on this
        if not self.blocking:
            initial_blocking = False
            self.set_blocking()

        match filetype:
            case ".nc4":
                hardware_metadata, environment_metadata_list = load_metadata_from_netcdf(filepath)
                self.set_hardware(hardware_metadata)
                self.set_environments(environment_metadata_list)
            case ".xlsx":
                hardware_metadata, environment_metadata_list, profile_event_list = load_metadata_from_worksheet(filepath)
                self.set_hardware(hardware_metadata)
                self.set_environments(environment_metadata_list)
                self.set_profile_event_list(profile_event_list)

        if not initial_blocking:
            self.clear_blocking()

    def save_template(self, filepath):
        pass

    # region: Hardware
    def set_hardware(self, hardware_metadata: HardwareMetadata) -> None:
        """Validates hardware_metadata and sends data to relevant processes"""
        # Validate Rattlesnake State
        if self.state not in (RattlesnakeState.INIT, RattlesnakeState.HARDWARE_STORE, RattlesnakeState.ENVIRONMENT_STORE):
            raise RuntimeError(f"Invalid state for this setting hardware: {self.state}")
        # Validate hardware
        if not isinstance(hardware_metadata, HardwareMetadata):
            raise TypeError("Rattlesnake.set_hardware requires a valid HardwareMetadata class")
        hardware_metadata.validate()

        # Send hardware metadata to the correct processes
        self.log("Setting Hardware")
        self.environment_manager.initialize_hardware(hardware_metadata)
        self.event_container.acquisition_ready_event.clear()
        self.queue_container.acquisition_command_queue.put(TASK_NAME, (GlobalCommands.INITIALIZE_HARDWARE, hardware_metadata))
        self.event_container.output_ready_event.clear()
        self.queue_container.output_command_queue.put(TASK_NAME, (GlobalCommands.INITIALIZE_HARDWARE, hardware_metadata))

        if self.blocking:
            ready_event_list = [
                self.event_container.acquisition_ready_event,
                self.event_container.output_ready_event,
                *self.environment_manager.ready_event_list,
            ]
            active_event_list = []
            self.wait_for_events(ready_event_list, active_event_list)

            # Update state
            self.hardware_metadata = hardware_metadata

    # region: Environments
    def set_environments(self, environment_metadata_list: List[EnvironmentMetadata]):
        """Validates environment_metadata, starts up environment processes, assigns queues,
        and sends data to relevant processes"""
        # Validate Rattlesnake State
        if self.state not in (RattlesnakeState.HARDWARE_STORE, RattlesnakeState.ENVIRONMENT_STORE):
            raise RuntimeError(f"Invalid state for setting environment: {self.state}")
        # Validate environment metadata list
        self.environment_manager.validate_environment_metadata(environment_metadata_list)

        # Send environment meetadata to correct processes
        self.log("Setting Environment")
        self.environment_manager.initialize_environments(environment_metadata_list)
        self.event_container.acquisition_ready_event.clear()
        self.queue_container.acquisition_command_queue.put(
            TASK_NAME, (GlobalCommands.INITIALIZE_ENVIRONMENT, self.environment_manager.environment_metadata)
        )
        self.event_container.output_ready_event.clear()
        self.queue_container.output_command_queue.put(
            TASK_NAME, (GlobalCommands.INITIALIZE_ENVIRONMENT, self.environment_manager.environment_metadata)
        )

        if self.blocking:
            ready_event_list = [
                self.event_container.acquisition_ready_event,
                self.event_container.output_ready_event,
                *self.environment_manager.ready_event_list,
            ]
            active_event_list = []
            self.wait_for_events(ready_event_list, active_event_list)

            # Update state
            self.environment_metadata = self.environment_manager.environment_metadata

    # region: Acquisition
    def set_stream_metadata(self, stream_metadata: StreamMetadata):
        """
        This is only used to load a stream_metadata to the controller for UI purposes. Start_acquisition
        still requirs a stream_metadata object so the metadata stored here will never be used.
        """
        if self.state != RattlesnakeState.ENVIRONMENT_STORE:
            raise RuntimeError(f"Invalid state for starting acquisition: {self.state}")
        if not isinstance(stream_metadata, StreamMetadata):
            raise TypeError("Rattlesnake.set_stream requires a valid StreamMetadata class")
        stream_metadata.validate()

        self.last_stream_metadata = stream_metadata

    def start_acquisition(self, stream_metadata: StreamMetadata):
        # Validate Rattlesnake State
        if self.state != RattlesnakeState.ENVIRONMENT_STORE:
            raise RuntimeError(f"Invalid state for starting acquisition: {self.state}")
        # Validate stream metadata
        if not isinstance(stream_metadata, StreamMetadata):
            raise TypeError("Rattlesnake.set_stream requires a valid StreamMetadata class")
        stream_metadata.validate()

        # Store streaming metadata to controller (side note: ControllerProcess decides when/why to stream not StreamingProcess)
        self.log("Setting Stream Metadata")
        self.event_container.streaming_ready_event.clear()
        self.queue_container.streaming_command_queue.put(
            TASK_NAME,
            (GlobalCommands.INITIALIZE_STREAMING, (stream_metadata, self.hardware_metadata, self.environment_metadata)),
        )

        # Tell controller to start up the hardware, controller takes over logic from here
        self.log("Arming Test Hardware")
        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.RUN_HARDWARE, stream_metadata))

        if self.blocking:
            ready_event_list = [
                self.event_container.streaming_ready_event,
            ]
            active_event_list = [
                self.event_container.acquisition_active_event,
                self.event_container.output_active_event,
            ]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=True)

        # Update stream_metadata
        self.last_stream_metadata = stream_metadata

    def stop_acquisition(self):
        # Validate rattlesnake state (rattlesnake was acquiring data)
        if self.state not in (RattlesnakeState.HARDWARE_ACTIVE, RattlesnakeState.ENVIRONMENT_ACTIVE):
            raise RuntimeError(f"Invalid state for stopping acquisition: {self.state}")

        self.log("Disarming Test Hardware")
        self.event_container.acquisition_ready_event.clear()
        self.event_container.output_ready_event.clear()
        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.STOP_HARDWARE, None))

        if self.blocking:
            ready_event_list = [
                self.event_container.controller_ready_event,
            ]
            active_event_list = [
                self.event_container.acquisition_active_event,
                self.event_container.output_active_event,
                *self.environment_manager.active_event_list,
            ]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=False)

    # region: Active
    def start_environment(self, instructions):
        if self.state not in (RattlesnakeState.HARDWARE_ACTIVE, RattlesnakeState.ENVIRONMENT_ACTIVE):
            raise RuntimeError(f"Invalid state for starting environment: {self.state}")
        if not isinstance(instructions, EnvironmentInstructions):
            raise TypeError("Start_environment must be contain a valid EnvironmentInstructions object")

        # Validate instructions
        queue_name = self.environment_manager.validate_environment_instructions(instructions)

        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.START_ENVIRONMENT, (queue_name, instructions)))

        if self.blocking:
            ready_event_list = []
            active_event_list = [self.event_container.environment_active_events[queue_name]]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=True)

    def stop_environment(self, environment_name: str):
        if self.state not in (RattlesnakeState.ENVIRONMENT_ACTIVE,):
            raise RuntimeError(f"Invalid state for stopping environment: {self.state}")
        try:
            queue_name = self.environment_manager.queue_names_dict[environment_name]
        except KeyError:
            raise KeyError(f"No environments exist for {environment_name} name")

        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.STOP_ENVIRONMENT, queue_name))

        if self.blocking:
            ready_event_list = []
            active_event_list = [self.event_container.environment_active_events[queue_name]]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=False)

    def environment_at_target_level(self, environment_name: str):
        if self.state not in (RattlesnakeState.ENVIRONMENT_ACTIVE,):
            raise RuntimeError(f"Invalid state for streaming at target level: {self.state}")
        try:
            self.environment_manager.queue_names_dict[environment_name]
        except KeyError:
            raise KeyError(f"No environments exist for {environment_name} name")

        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.STREAM_AT_TARGET_LEVEL, environment_name))

    # region: Streaming
    def start_streaming(self):
        if self.state not in (RattlesnakeState.HARDWARE_ACTIVE, RattlesnakeState.ENVIRONMENT_ACTIVE):
            raise RuntimeError(f"Invalid state for starting streaming: {self.state}")
        if self.streaming:
            raise RuntimeError(f"Rattlesnake is currently streaming")

        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.STREAM_MANUAL, None))

        if self.blocking:
            ready_event_list = []
            active_event_list = [self.event_container.streaming_active_event]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=True)

    def stop_streaming(self):
        if not self.streaming:
            raise RuntimeError(f"Rattlesnake is not currently streaming")

        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.STOP_STREAMING, None))

        if self.blocking:
            ready_event_list = []
            active_event_list = [self.event_container.streaming_active_event]
            self.wait_for_events(ready_event_list, active_event_list, active_event_check=False)

    # region: Profile
    def set_profile_event_list(self, profile_event_list: List[ProfileEvent]):
        """
        This is mainly to preload profile event list for UI purposes. You
        still have to give a profile event list to start_profile so this is
        not useful for headless opperation
        """
        self.log("Settting Profile Event List")

        if self.state not in (RattlesnakeState.ENVIRONMENT_STORE, RattlesnakeState.HARDWARE_ACTIVE):
            raise RuntimeError(f"Invalid state for storing profile: {self.state}")

        self.environment_manager.validate_profile_events(profile_event_list)
        self.last_profile_event_list = profile_event_list

    def start_profile(self, profile_event_list: List[ProfileEvent]):
        self.log("Starting Profile")
        if self.state not in (RattlesnakeState.HARDWARE_ACTIVE,):
            raise RuntimeError(f"Invalid state to start profile: {self.state}")

        # Validate and assign queue_names to events
        self.environment_manager.validate_profile_events(profile_event_list)
        self.profile_manager.validate_profile_list(profile_event_list)

        # Start profile
        self.log("Starting Profile")
        self.event_container.controller_ready_event.clear()
        self.profile_manager.start_profile(profile_event_list)

        if self.blocking:
            ready_event_list = [self.event_container.controller_ready_event]
            active_event_list = []
            self.wait_for_events(ready_event_list, active_event_list)

        # Update profile event list
        self.last_profile_event_list = profile_event_list

    def stop_profile(self):
        self.log("Stopping Profile")
        self.event_container.controller_ready_event.clear()
        self.profile_manager.stop_profile()

        if self.blocking:
            ready_event_list = [self.event_container.controller_ready_event]
            active_event_list = []
            self.wait_for_events(ready_event_list, active_event_list)

    # region: Shutdown
    def shutdown(self):
        if self.state in (RattlesnakeState.HARDWARE_ACTIVE, RattlesnakeState.ENVIRONMENT_ACTIVE):
            self.stop_acquisition()
        # Close out of acquisition, output, streaming process
        self.queue_container.log_file_queue.put(f"{datetime.now()}: Joining Controller Process\n")
        flush_queue(self.queue_container.gui_update_queue, timeout=CLOSE_TIMEOUT)
        self.queue_container.controller_command_queue.put(TASK_NAME, (GlobalCommands.QUIT, None))
        self.controller_proc.join(timeout=CLOSE_TIMEOUT)
        if self.controller_proc.is_alive():
            self.queue_container.log_file_queue.put(f"{datetime.now()}: Force Closing Controller Process\n")
            self.event_container.controller_close_event.set()
            self.controller_proc.join(timeout=CLOSE_TIMEOUT)
            if self.controller_proc.is_alive() and not self.threaded:
                self.controller_proc.terminate()
                self.controller_proc.join()
        self.queue_container.log_file_queue.put(f"{datetime.now()}: Joining Acquisition Process\n")
        self.queue_container.acquisition_command_queue.put(TASK_NAME, (GlobalCommands.QUIT, None))
        self.acquisition_proc.join(timeout=CLOSE_TIMEOUT)
        if self.acquisition_proc.is_alive():
            self.queue_container.log_file_queue.put(f"{datetime.now()}: Force Closing Acquisition Process\n")
            self.event_container.acquisition_close_event.set()
            self.acquisition_proc.join(timeout=CLOSE_TIMEOUT)
            if self.acquisition_proc.is_alive() and not self.threaded:
                self.acquisition_proc.terminate()
                self.acquisition_proc.join()
        self.queue_container.log_file_queue.put(f"{datetime.now()}: Joining Output Process\n")
        self.queue_container.output_command_queue.put(TASK_NAME, (GlobalCommands.QUIT, None))
        self.output_proc.join(timeout=CLOSE_TIMEOUT)
        if self.output_proc.is_alive():
            self.queue_container.log_file_queue.put(f"{datetime.now()}: Force Closing Output Process\n")
            self.event_container.output_close_event.set()
            self.output_proc.join(timeout=CLOSE_TIMEOUT)
            if self.output_proc.is_alive() and not self.threaded:
                self.output_proc.terminate()
                self.output_proc.join()
        self.queue_container.log_file_queue.put(f"{datetime.now()}: Joining Streaming Process\n")
        self.queue_container.streaming_command_queue.put(TASK_NAME, (GlobalCommands.QUIT, None))
        self.streaming_proc.join(timeout=CLOSE_TIMEOUT)
        if self.streaming_proc.is_alive():
            self.queue_container.log_file_queue.put(f"{datetime.now()}: Force Closing Streaming Process\n")
            self.event_container.streaming_close_event.set()
            self.streaming_proc.join(timeout=CLOSE_TIMEOUT)
            if self.streaming_proc.is_alive() and not self.threaded:
                self.streaming_proc.terminate()
                self.streaming_proc.join()

        # Close out of environment processes
        self.environment_manager.close_environments()

        # Close out log file process
        self.queue_container.log_file_queue.put("{:}: Joining Log File Process\n".format(datetime.now()))
        self.queue_container.log_file_queue.put(GlobalCommands.QUIT)
        self.log_file_process.join()

    def log(self, string):
        """Pass a message to the log_file_queue along with date/time and task name

        Parameters
        ----------
        string : str
            Message that will be written to the queue

        """
        self.queue_container.log_file_queue.put(f"{datetime.now()}: {TASK_NAME} -- {string}\n")


# region: Data loading
def load_metadata_from_netcdf(filepath):
    """Loads a test file using a file dialog"""
    dataset = netCDF4.Dataset(filepath)

    # Channel Table
    channel_list = load_channel_list_from_netcdf(filepath)

    # Hardware

    hardware_type = HardwareType(dataset.hardware)
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            from rattlesnake.hardware.sdynpy_system import SDynPySystemMetadata

            hardware_metadata = SDynPySystemMetadata()
            hardware_metadata.hardware_file = dataset.hardware_file

        case _:
            raise ValueError(f"{hardware_type} has not been implemented yet")

    hardware_metadata.channel_list = channel_list
    hardware_metadata.sample_rate = int(dataset.sample_rate)
    hardware_metadata.time_per_read = float(dataset.time_per_read)
    hardware_metadata.time_per_write = float(dataset.time_per_write)
    hardware_metadata.output_oversample = int(dataset.output_oversample)

    # Environments
    environment_metadata_list = []
    for environment_index, environment_name in enumerate(
        dataset.variables["environment_names"][...],
    ):
        environment_active_channels = dataset.variables["environment_active_channels"][:, environment_index]
        environment_channel_list = [channel for channel, channel_bool in zip(channel_list, environment_active_channels) if channel_bool == 1]
        environment_type_int = dataset.variables["environment_types"][environment_index]
        environment_type = ControlTypes(environment_type_int)
        environment_group = dataset.groups[environment_name]

        match environment_type:
            case ControlTypes.TIME:
                from rattlesnake.environment.time_environment import TimeMetadata

                environment_metadata = TimeMetadata(environment_name)
                environment_metadata.sample_rate = hardware_metadata.sample_rate  # This is rough
            case _:
                raise TypeError(f"{environment_type} has not been implemented yet")

        environment_metadata.channel_list = environment_channel_list
        environment_metadata.retrieve_metadata_from_netcdf(environment_group)

        environment_metadata_list.append(environment_metadata)

    return (hardware_metadata, environment_metadata_list)


def load_metadata_from_worksheet(filepath):
    workbook = openpyxl.load_workbook(filepath, read_only=True)

    # Channel table
    channel_list = load_channel_list_from_worksheet(filepath)

    # Hardware
    hardware_sheet = workbook["Hardware"]
    for row in hardware_sheet.rows:
        name = str(row[0].value).lower().strip().replace(" ", "_")
        value = row[1].value
        match name:
            case "hardware_type":
                hardware_type_int = int(value)
            case "hardware_file":
                hardware_file = value
            case "sample_rate":
                sample_rate = value
            case "time_per_read":
                time_per_read = value
            case "time_per_write":
                time_per_write = value
            case "integration_oversampling":
                output_oversample = int(value)
            case "task_trigger":
                task_trigger = int(value)
            case "task_trigger_output_channel":
                task_output = str(value)
            case "maximum_acquisition_processes":
                maximum_acquisition_processes = int(value)
            case "":
                continue
            case _:
                print(f"Hardware sheet entry {row[0].value} not recognized")

    hardware_type = HardwareType(hardware_type_int)
    match hardware_type:
        case HardwareType.SDYNPY_SYSTEM:
            from rattlesnake.hardware.sdynpy_system import SDynPySystemMetadata

            hardware_metadata = SDynPySystemMetadata()
            hardware_metadata.hardware_file = hardware_file
            hardware_metadata.output_oversample = output_oversample
        case _:
            raise TypeError(f"{hardware_type} has not been implemented yet")

    hardware_metadata.channel_list = channel_list
    hardware_metadata.sample_rate = int(sample_rate)
    hardware_metadata.time_per_read = float(time_per_read)
    hardware_metadata.time_per_write = float(time_per_write)

    # Environment
    environment_names = []
    environment_channel_list = {}
    sheets = workbook.sheetnames
    if len(sheets) > 1:
        sheets = [sheet for sheet in sheets if "channel" in sheet.lower()]
    channel_sheet = workbook[sheets[0]]
    col = 24
    num_channels = len(channel_list)
    while True:
        environment_name = channel_sheet.cell(row=2, column=col).value

        # Stop if empty or None
        if environment_name is None or str(environment_name).strip() == "":
            break

        # Build environment channel list
        environment_active_channels = [0] * num_channels
        for i in range(num_channels):
            row = 3 + i
            value = channel_sheet.cell(row=row, column=col).value

            if value is not None and str(value).strip() != "":
                environment_active_channels[i] = 1
        environment_channels = [channel for channel, channel_bool in zip(channel_list, environment_active_channels) if channel_bool == 1]

        environment_names.append(environment_name)
        environment_channel_list[environment_name] = environment_channels
        col += 1

    environment_metadata_list = []
    environment_types = {"Global": "Global"}
    for environment_name in environment_names:
        environment_sheet = workbook[environment_name]
        environment_type_name = environment_sheet.cell(row=1, column=2).value
        environment_type_name = str(environment_type_name).upper()
        environment_type = ControlTypes[environment_type_name]
        environment_types[environment_name] = environment_type

        match environment_type:
            case ControlTypes.TIME:
                from rattlesnake.environment.time_environment import TimeMetadata

                environment_metadata = TimeMetadata(environment_name)
                environment_metadata.sample_rate = sample_rate
            case _:
                raise TypeError(f"{environment_type} has not been implemented yet")

        environment_metadata.channel_list = environment_channel_list[environment_name]
        environment_metadata.retrieve_metadata_from_worksheet(environment_sheet)
        environment_metadata_list.append(environment_metadata)

    profile_sheet = workbook["Test Profile"]
    index = 2
    profile_event_list = []
    while True:
        timestamp = profile_sheet.cell(index, 1).value
        if timestamp is None or (isinstance(timestamp, str) and timestamp.strip() == ""):
            break

        environment_name = profile_sheet.cell(index, 2).value
        environment_type = environment_types[environment_name]

        # I have to conver the command string to an actual command
        command = profile_sheet.cell(index, 3).value
        command = str(command).upper().strip().replace(" ", "_")
        if command in GlobalCommands.__members__:
            command = GlobalCommands[command]
        elif command in ENVIRONMENT_COMMANDS[environment_type].__members__:
            command = ENVIRONMENT_COMMANDS[environment_type][command]
        else:
            raise TypeError(f"Invalid command: {command} for {environment_name} | {environment_type}")

        data = profile_sheet.cell(index, 4).value
        data = None if isinstance(data, str) and not data.strip() else data

        event = ProfileEvent(timestamp, environment_name, command, data)
        profile_event_list.append(event)
        index += 1

    workbook.close()

    return (hardware_metadata, environment_metadata_list, profile_event_list)


def save_combined_environments_profile_template(filename, hardware_metadata, environment_metadata_list, profile_event_list):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Channel Table"
    hardware_worksheet = workbook.create_sheet("Hardware")
    # Create the header
    worksheet.cell(row=1, column=2, value="Test Article Definition")
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    worksheet.cell(row=1, column=5, value="Instrument Definition")
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    worksheet.cell(row=1, column=12, value="Channel Definition")
    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
    worksheet.cell(row=1, column=20, value="Output Feedback")
    worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
    worksheet.cell(row=1, column=22, value="Limits")
    worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
    for col_idx, val in enumerate(
        [
            "Channel Index",
            "Node Number",
            "Node Direction",
            "Comment",
            "Serial Number",
            "Triax DoF",
            "Sensitivity  (mV/EU)",
            "Engineering Unit",
            "Make",
            "Model",
            "Calibration Exp Date",
            "Physical Device",
            "Physical Channel",
            "Type",
            "Minimum Value (V)",
            "Maximum Value (V)",
            "Coupling",
            "Current Excitation Source",
            "Current Excitation Value",
            "Physical Device",
            "Physical Channel",
            "Warning Level (EU)",
            "Abort Level (EU)",
        ]
    ):
        worksheet.cell(row=2, column=1 + col_idx, value=val)
    # Fill out the hardware worksheet
    hardware_worksheet.cell(1, 1, "Hardware Type")
    hardware_worksheet.cell(1, 2, "# Enter hardware index here")
    hardware_worksheet.cell(
        1,
        3,
        "Hardware Indices: 0 - NI DAQmx; 1 - LAN XI; 2 - Data Physics Quattro; "
        "3 - Data Physics 900 Series; 4 - Exodus Modal Solution; 5 - State Space Integration; "
        "6 - SDynPy System Integration",
    )
    hardware_worksheet.cell(2, 1, "Hardware File")
    hardware_worksheet.cell(
        2,
        2,
        "# Path to Hardware File (Depending on Hardware Device: 0 - Not Used; 1 - Not Used; "
        "2 - Path to DpQuattro.dll library file; 3 - Not Used; 4 - Path to Exodus Eigensolution; "
        "5 - Path to State Space File; 6 - Path to SDynPy system file)",
    )
    hardware_worksheet.cell(3, 1, "Sample Rate")
    hardware_worksheet.cell(3, 2, "# Sample Rate of Data Acquisition System")
    hardware_worksheet.cell(4, 1, "Time Per Read")
    hardware_worksheet.cell(4, 2, "# Number of seconds per Read from the Data Acquisition System")
    hardware_worksheet.cell(5, 1, "Time Per Write")
    hardware_worksheet.cell(5, 2, "# Number of seconds per Write to the Data Acquisition System")
    hardware_worksheet.cell(6, 1, "Maximum Acquisition Processes")
    hardware_worksheet.cell(
        6,
        2,
        "# Maximum Number of Acquisition Processes to start to pull data from hardware",
    )
    hardware_worksheet.cell(
        6,
        3,
        "Only Used by LAN-XI Hardware.  This row can be deleted if LAN-XI is not used",
    )
    hardware_worksheet.cell(7, 1, "Integration Oversampling")
    hardware_worksheet.cell(7, 2, "# For virtual control, an integration oversampling can be specified")
    hardware_worksheet.cell(
        7,
        3,
        "Only used for virtual control (Exodus, State Space, or SDynPy).  " "This row can be deleted if these are not used.",
    )
    hardware_worksheet.cell(8, 1, "Task Trigger")
    hardware_worksheet.cell(8, 2, "# Start trigger type")
    hardware_worksheet.cell(
        8,
        3,
        "Task Triggers: 0 - Internal, 1 - PFI0 with external trigger, 2 - PFI0 with Analog Output "
        "trigger.  Only used for NI hardware.  This row can be deleted if NI is not used.",
    )
    hardware_worksheet.cell(9, 1, "Task Trigger Output Channel")
    hardware_worksheet.cell(9, 2, "# Physical device and channel that generates a trigger signal")
    hardware_worksheet.cell(
        9,
        3,
        "Only used if Task Triggers is 2.  Only used for NI hardware.  " "This row can be deleted if it is not used.",
    )

    # Now do the environment
    worksheet.cell(row=1, column=24, value="Environments")
    for row, (value, name) in enumerate(environment_metadata_list):
        environment_metadata_list[value].create_environment_template(name, workbook)
        worksheet.cell(row=2, column=24 + row, value=name)
    # Now create a profile page
    profile_sheet = workbook.create_sheet("Test Profile")
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")

    workbook.save(filename)
