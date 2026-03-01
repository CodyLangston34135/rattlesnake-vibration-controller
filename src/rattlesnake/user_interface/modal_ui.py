from rattlesnake.utilities import GlobalCommands, VerboseMessageQueue, load_python_module
from rattlesnake.hardware.abstract_hardware import HardwareMetadata
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.environment.abstract_environment import EnvironmentMetadata
from rattlesnake.environment.modal_environment import ModalCommands, ModalUICommands, ModalMetadata
from rattlesnake.user_interface.abstract_user_interface import AbstractUI
from rattlesnake.user_interface.ui_utilities import multiline_plotter, error_message_qt, environment_definition_ui_paths, environment_run_ui_paths
from rattlesnake.user_interface.modal_ui_utilities import ModalMDISubWindow
import inspect
import os
import openpyxl
import numpy as np
import multiprocessing as mp
import netCDF4 as nc4
import scipy.signal as sig
from glob import glob
from qtpy import QtWidgets, QtCore, uic

CONTROL_TYPE = ControlTypes.MODAL
MAXIMUM_NAME_LENGTH = 50


class ModalUI(AbstractUI):
    """Modal User Interface class defining the interface with the controller

    This class is used to define the interface between the User Interface of the
    Modal environment in the controller and the main controller."""

    def __init__(
        self,
        environment_name: str,
        definition_tabwidget: QtWidgets.QTabWidget,
        system_id_tabwidget: QtWidgets.QTabWidget,  # pylint: disable=unused-argument
        test_predictions_tabwidget: QtWidgets.QTabWidget,  # pylint: disable=unused-argument
        run_tabwidget: QtWidgets.QTabWidget,
        environment_command_queue: VerboseMessageQueue,
        controller_communication_queue: VerboseMessageQueue,
        log_file_queue: mp.Queue,
    ):
        """
        Constructs a Modal User Interface

        Given the tab widgets from the main interface as well as communication
        queues, this class assembles the user interface components specific to
        the Modal Environment

        Parameters
        ----------
        definition_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Control
            Definition main tab
        system_id_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the System
            Identification main tab
        test_predictions_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Test Predictions
            main tab
        run_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Run
            main tab.
        environment_command_queue : VerboseMessageQueue
            Queue for sending commands to the Modal Environment
        controller_communication_queue : VerboseMessageQueue
            Queue for sending global commands to the controller
        log_file_queue : Queue
            Queue where log file messages can be written.

        """
        super().__init__(
            environment_name,
            environment_command_queue,
            controller_communication_queue,
            log_file_queue,
        )
        # Add the page to the control definition tabwidget
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[CONTROL_TYPE], self.definition_widget)
        # Add the page to the run tabwidget
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[CONTROL_TYPE], self.run_widget)

        self.trigger_widgets = [
            self.definition_widget.trigger_channel_selector,
            self.definition_widget.pretrigger_selector,
            self.definition_widget.trigger_slope_selector,
            self.definition_widget.trigger_level_selector,
            self.definition_widget.trigger_level_voltage_display,
            self.definition_widget.trigger_level_eu_display,
            self.definition_widget.hysteresis_selector,
            self.definition_widget.hysteresis_voltage_display,
            self.definition_widget.hysteresis_eu_display,
            self.definition_widget.hysteresis_length_selector,
            self.definition_widget.hysteresis_samples_display,
            self.definition_widget.hysteresis_time_display,
        ]

        self.signal_generator_widgets = [
            self.definition_widget.random_rms_selector,
            self.definition_widget.random_min_frequency_selector,
            self.definition_widget.random_max_frequency_selector,
            self.definition_widget.burst_rms_selector,
            self.definition_widget.burst_min_frequency_selector,
            self.definition_widget.burst_max_frequency_selector,
            self.definition_widget.burst_on_percentage_selector,
            self.definition_widget.pseudorandom_rms_selector,
            self.definition_widget.pseudorandom_min_frequency_selector,
            self.definition_widget.pseudorandom_max_frequency_selector,
            self.definition_widget.chirp_level_selector,
            self.definition_widget.chirp_min_frequency_selector,
            self.definition_widget.chirp_max_frequency_selector,
            self.definition_widget.square_level_selector,
            self.definition_widget.square_frequency_selector,
            self.definition_widget.square_percent_on_selector,
            self.definition_widget.sine_level_selector,
            self.definition_widget.sine_frequency_selector,
        ]

        self.window_parameter_widgets = [
            self.definition_widget.window_value_label,
            self.definition_widget.window_value_selector,
        ]

        self.definition_widget.reference_channels_selector.setColumnCount(3)
        self.definition_widget.reference_channels_selector.setVerticalHeaderLabels(["Enabled", "Reference", "Channel"])

        self.hardware_metadata = None
        self.metadata = None
        self.channel_names = None
        self.acceptance_function = None
        self.plot_data_items = {}
        self.reference_channel_indices = None
        self.all_output_channel_indices = None
        self.response_channel_indices = None
        self.last_frame = None
        self.last_frf = None
        self.last_coherence = None
        self.last_response_cpsd = None
        self.last_reference_cpsd = None
        self.last_condition = None
        self.acquiring = False
        self.netcdf_handle = None
        self.override_table = {}
        self.reciprocal_responses = []

        # Store some information into the channel display so the plots have
        # access to it
        self.run_widget.channel_display_area.time_abscissa = None
        self.run_widget.channel_display_area.frequency_abscissa = None
        self.run_widget.channel_display_area.window_function = None
        self.run_widget.channel_display_area.last_frame = None
        self.run_widget.channel_display_area.last_spectrum = None
        self.run_widget.channel_display_area.last_autospectrum = None
        self.run_widget.channel_display_area.last_frf = None
        self.run_widget.channel_display_area.last_coh = None
        self.run_widget.channel_display_area.channel_names = None
        self.run_widget.channel_display_area.reference_channel_indices = None
        self.run_widget.channel_display_area.response_channel_indices = None

        self.complete_ui()
        self.connect_callbacks()

    @property
    def reference_indices(self):
        """Returns indices corresponding to the reference channels"""
        return [
            i
            for i in range(self.definition_widget.reference_channels_selector.rowCount())
            if self.definition_widget.reference_channels_selector.cellWidget(i, 0).isChecked()
            and self.definition_widget.reference_channels_selector.cellWidget(i, 1).isChecked()
        ]

    @property
    def response_indices(self):
        """Returns indices corresponding to the response channels in a test"""
        return [
            i
            for i in range(self.definition_widget.reference_channels_selector.rowCount())
            if self.definition_widget.reference_channels_selector.cellWidget(i, 0).isChecked()
            and not self.definition_widget.reference_channels_selector.cellWidget(i, 1).isChecked()
        ]

    @property
    def output_channel_indices(self):
        """Returns indices corresponding to the output channels in a test"""
        return [i for i in self.all_output_channel_indices if self.definition_widget.reference_channels_selector.cellWidget(i, 0).isChecked()]

    @property
    def initialized_response_names(self):
        """Returns channel names corresponding to the initialized response channels"""
        return [self.channel_names[i] for i in range(len(self.channel_names)) if i not in self.metadata.response_channel_indices]

    @property
    def initialized_reference_names(self):
        """Returns channel names corresponding to the initialized reference channels"""
        return [self.channel_names[i] for i in self.metadata.reference_channel_indices]

    def complete_ui(self):
        """Applies some finishing touches to the UI"""
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        for widget in self.trigger_widgets:
            widget.setEnabled(False)

        # Set common look and feel for plots
        plot_widgets = [self.definition_widget.output_signal_plot]
        for plot_widget in plot_widgets:
            plot_item = plot_widget.getPlotItem()
            plot_item.showGrid(True, True, 0.25)
            plot_item.enableAutoRange()
            plot_item.getViewBox().enableAutoRange(enable=True)

        # Disable the currently inactive portions of the definition layout
        self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        for widget in self.window_parameter_widgets:
            widget.hide()

    def connect_callbacks(self):
        """Connects callback functions to the user interface widgets"""
        # Definition Callbacks
        self.definition_widget.samples_per_frame_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.system_id_overlap_percentage_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.triggering_type_selector.currentIndexChanged.connect(self.activate_trigger_options)
        self.definition_widget.acceptance_selector.currentIndexChanged.connect(self.select_acceptance)
        self.definition_widget.trigger_channel_selector.currentIndexChanged.connect(self.update_trigger_levels)
        self.definition_widget.trigger_level_selector.valueChanged.connect(self.update_trigger_levels)
        self.definition_widget.hysteresis_selector.valueChanged.connect(self.update_trigger_levels)
        self.definition_widget.regenerate_signal_button.clicked.connect(self.generate_signal)
        self.definition_widget.signal_generator_selector.currentChanged.connect(self.update_signal)
        for widget in self.signal_generator_widgets:
            widget.valueChanged.connect(self.update_signal)
        self.definition_widget.check_selected_button.clicked.connect(self.check_selected_reference_channels)
        self.definition_widget.uncheck_selected_button.clicked.connect(self.uncheck_selected_reference_channels)
        self.definition_widget.enable_selected_button.clicked.connect(self.enable_selected_channels)
        self.definition_widget.disable_selected_button.clicked.connect(self.disable_selected_channels)
        self.definition_widget.hysteresis_length_selector.valueChanged.connect(self.update_hysteresis_length)
        self.definition_widget.system_id_averaging_scheme_selector.currentIndexChanged.connect(self.update_averaging_type)
        self.definition_widget.system_id_transfer_function_computation_window_selector.currentIndexChanged.connect(self.update_window)
        # Run Callbacks
        self.run_widget.preview_test_button.clicked.connect(self.preview_acquisition)
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.run_widget.select_file_button.clicked.connect(self.select_file)
        self.run_widget.accept_average_button.clicked.connect(self.accept_frame)
        self.run_widget.reject_average_button.clicked.connect(self.reject_frame)
        self.run_widget.new_window_button.clicked.connect(self.new_window)
        self.run_widget.new_from_template_combobox.currentIndexChanged.connect(self.new_window_from_template)
        self.run_widget.tile_layout_button.clicked.connect(self.run_widget.channel_display_area.tileSubWindows)
        self.run_widget.close_all_button.clicked.connect(self.close_windows)
        self.run_widget.decrement_channels_button.clicked.connect(self.decrement_channels)
        self.run_widget.increment_channels_button.clicked.connect(self.increment_channels)
        self.run_widget.dof_override_table.itemChanged.connect(self.update_override_table)
        self.run_widget.add_override_button.clicked.connect(self.add_override_channel)
        self.run_widget.remove_override_button.clicked.connect(self.remove_override_channel)

    # Definition Callbacks
    def update_parameters(self):
        """Updates widget values when fundamental signal processing parameters change"""
        if self.definition_widget.samples_per_frame_selector.value() % 2 == 1:
            self.definition_widget.samples_per_frame_selector.blockSignals(True)
            self.definition_widget.samples_per_frame_selector.setValue(self.definition_widget.samples_per_frame_selector.value() + 1)
            self.definition_widget.samples_per_frame_selector.blockSignals(False)
        data = self.collect_environment_definition_parameters()
        self.definition_widget.samples_per_acquire_display.setValue(data.samples_per_acquire)
        self.definition_widget.frame_time_display.setValue(data.frame_time)
        self.definition_widget.nyquist_frequency_display.setValue(data.nyquist_frequency)
        self.definition_widget.fft_lines_display.setValue(data.fft_lines)
        self.definition_widget.frequency_spacing_display.setValue(data.frequency_spacing)
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def update_reference_channels(self):
        """Updates widgets based on changes in the selected reference channels"""
        self.definition_widget.response_channels_display.setValue(len(self.response_indices))
        self.definition_widget.reference_channels_display.setValue(len(self.reference_indices))
        self.definition_widget.output_channels_display.setValue(len(self.output_channel_indices))
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def check_selected_reference_channels(self):
        """Checks reference channels that are selected in the list widget"""
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index, 1).setChecked(True)

    def uncheck_selected_reference_channels(self):
        """Unchecks reference channels that are selected in the list widget"""
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index, 1).setChecked(False)

    def enable_selected_channels(self):
        """Enables channels that are selected in the list widget"""
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index, 0).setChecked(True)

    def disable_selected_channels(self):
        """Disables channels that are selected in the list widget"""
        select = self.definition_widget.reference_channels_selector.selectionModel()
        rows = select.selectedRows()
        for row in rows:
            index = row.row()
            self.definition_widget.reference_channels_selector.cellWidget(index, 0).setChecked(False)

    def activate_trigger_options(self):
        """Enables widgets corresponding to the trigger selection"""
        if self.definition_widget.triggering_type_selector.currentIndex() == 0:
            for widget in self.trigger_widgets:
                widget.setEnabled(False)
        else:
            for widget in self.trigger_widgets:
                widget.setEnabled(True)

    def select_acceptance(self):
        """Selects the acceptance type and opens up a file dialog if necessary"""
        if self.definition_widget.acceptance_selector.currentIndex() == 2:
            # Open up a file dialog
            filename, _ = QtWidgets.QFileDialog.getOpenFileName(
                self.definition_widget,
                "Select Python Module",
                filter="Python Modules (*.py)",
            )
            if filename == "":
                self.definition_widget.acceptance_selector.setCurrentIndex(0)
                return
            module = load_python_module(filename)
            functions = [function for function in inspect.getmembers(module) if inspect.isfunction(function[1])]
            item, ok_pressed = QtWidgets.QInputDialog.getItem(
                self.definition_widget,
                "Select Acceptance Function",
                "Function Name:",
                [function[0] for function in functions],
                0,
                False,
            )
            if ok_pressed:
                self.acceptance_function = [filename, item]
            else:
                self.definition_widget.acceptance_selector.setCurrentIndex(0)
                return
        else:
            self.acceptance_function = None

    def update_trigger_levels(self):
        """Updates trigger levels based on selected widget values"""
        data = self.collect_environment_definition_parameters()
        t_v, t_eu, h_v, h_eu = data.get_trigger_levels(self.hardware_metadata.channel_list)
        self.definition_widget.trigger_level_voltage_display.setValue(t_v)
        self.definition_widget.trigger_level_eu_display.setValue(t_eu)
        self.definition_widget.hysteresis_voltage_display.setValue(h_v)
        self.definition_widget.hysteresis_eu_display.setValue(h_eu)
        eu_suffix = self.hardware_metadata.channel_list[data.trigger_channel].unit
        self.definition_widget.hysteresis_eu_display.setSuffix((" " + eu_suffix) if not (eu_suffix == "" or eu_suffix is None) else "")
        self.definition_widget.trigger_level_eu_display.setSuffix((" " + eu_suffix) if not (eu_suffix == "" or eu_suffix is None) else "")

    def update_hysteresis_length(self):
        """Updates hysterisis length based on the selected trigger parameters"""
        data = self.collect_environment_definition_parameters()
        self.definition_widget.hysteresis_samples_display.setValue(data.hysteresis_samples)
        self.definition_widget.hysteresis_time_display.setValue(data.hysteresis_samples / data.sample_rate)

    def update_signal(self):
        """Updates the generated signal based on widget value changes"""
        if self.definition_widget.regenerate_signal_auto_checkbox.isChecked():
            self.generate_signal()

    def generate_signal(self):
        """Generates an example signal to show in the definition widget"""
        if self.hardware_metadata is None:
            return
        output_oversample = self.hardware_metadata.output_oversample
        output_rate = self.hardware_metadata.output_sample_rate
        data = self.collect_environment_definition_parameters()
        frame_output_samples = int(data.samples_per_frame * output_oversample)
        signal = data.generate_signal()
        # Reduce down to just one frame
        while signal.shape[-1] < frame_output_samples:
            signal = np.concatenate((signal, data.generate_signal()), axis=-1)
        signal = signal[..., :frame_output_samples]
        signal[data.disabled_signals] = 0
        times = np.arange(frame_output_samples) / output_rate
        for s, plot in zip(signal, self.plot_data_items["signal_representation"]):
            plot.setData(times, s)

    def update_averaging_type(self):
        """Enables exponential averaging coefficient widgets if exponential averaging is chosen"""
        if self.definition_widget.system_id_averaging_scheme_selector.currentIndex() == 0:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(False)
        else:
            self.definition_widget.system_id_averaging_coefficient_selector.setEnabled(True)

    def update_window(self):
        """Shows additional window function options based on the selected window"""
        if self.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex() == 2:
            for widget in self.window_parameter_widgets:
                widget.show()
        else:
            for widget in self.window_parameter_widgets:
                widget.hide()

    # Run Callbacks
    def preview_acquisition(self):
        """Tells the environment process to start in preview mode"""
        self.run_widget.stop_test_button.setEnabled(True)
        self.run_widget.preview_test_button.setEnabled(False)
        self.run_widget.start_test_button.setEnabled(False)
        self.run_widget.select_file_button.setEnabled(False)
        self.controller_communication_queue.put(self.log_name, (GlobalCommands.START_ENVIRONMENT, self.environment_name))
        self.environment_command_queue.put(self.log_name, (ModalCommands.START_CONTROL, None))
        self.run_widget.dof_override_table.setEnabled(False)
        self.run_widget.add_override_button.setEnabled(False)
        self.run_widget.remove_override_button.setEnabled(False)

    def start_control(self):
        """Tells the environment process to start in acquisition mode"""
        self.acquiring = True
        # Create the output file
        filename = self.run_widget.data_file_selector.text()
        if filename == "":
            error_message_qt("Invalid File", "Please select a file in which to store modal data")
            return
        if self.run_widget.autoincrement_checkbox.isChecked():
            # Add the file increment
            path, ext = os.path.splitext(filename)
            index = len(glob(path + "*" + ext))
            filename = path + f"_{index:04d}" + ext
        self.create_netcdf_file(filename)
        self.preview_acquisition()

    def stop_control(self):
        """Tells the environment process to stop the current measurement"""
        self.environment_command_queue.put(self.log_name, (ModalCommands.STOP_CONTROL, None))

    def select_file(self):
        """Brings up a file dialog box to select the save file location"""
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self.run_widget,
            "Select NetCDF File to Save Modal Data",
            filter="NetCDF File (*.nc4)",
        )
        if filename == "":
            return
        self.run_widget.data_file_selector.setText(filename)

    def accept_frame(self):
        """Sends a signal to the environment process to accept the current measurement frame"""
        self.environment_command_queue.put(self.log_name, (ModalCommands.ACCEPT_FRAME, True))
        self.run_widget.accept_average_button.setEnabled(False)
        self.run_widget.reject_average_button.setEnabled(False)

    def reject_frame(self):
        """Sends a signal to the environment process to reject the current measurement frame"""
        self.environment_command_queue.put(self.log_name, (ModalCommands.ACCEPT_FRAME, False))
        self.run_widget.accept_average_button.setEnabled(False)
        self.run_widget.reject_average_button.setEnabled(False)

    def new_window(self):
        """Creates a new window to display modal data"""
        widget = ModalMDISubWindow(self.run_widget.channel_display_area)
        self.run_widget.channel_display_area.addSubWindow(widget)
        widget.show()
        return widget
        # print('Windows: {:}'.format(self.run_widget.channel_display_area.subWindowList()))

    def new_window_from_template(self):
        """Creates a new window from the template functions"""
        if self.run_widget.new_from_template_combobox.currentIndex() == 0:
            return
        elif self.run_widget.new_from_template_combobox.currentIndex() == 6:
            # 3x3 channel grid
            for i in range(9):
                widget = self.new_window()
                widget.signal_selector.setCurrentIndex(0)
                widget.response_coordinate_selector.setCurrentIndex(i)
                widget.lock_response_checkbox.setChecked(False)
        elif self.run_widget.new_from_template_combobox.currentIndex() == 5:
            # Reference autospectra
            for index in self.run_widget.channel_display_area.reference_channel_indices:
                widget = self.new_window()
                widget.signal_selector.setCurrentIndex(3)
                widget.response_coordinate_selector.setCurrentIndex(index)
                widget.lock_response_checkbox.setChecked(True)
        else:
            corresponding_drive_responses = self.run_widget.channel_display_area.reciprocal_responses
            if self.run_widget.new_from_template_combobox.currentIndex() == 1:
                # Create drive point FRFs in magnitude
                for i, index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(4)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.data_type_selector.setCurrentIndex(0)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 2:
                # Create drive point FRFs in imaginary
                for i, index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(4)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.data_type_selector.setCurrentIndex(3)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 3:
                # Create drive point Coherence
                for i, index in enumerate(corresponding_drive_responses):
                    widget = self.new_window()
                    widget.signal_selector.setCurrentIndex(6)
                    widget.response_coordinate_selector.setCurrentIndex(index)
                    widget.reference_coordinate_selector.setCurrentIndex(i)
                    widget.lock_response_checkbox.setChecked(True)
            elif self.run_widget.new_from_template_combobox.currentIndex() == 4:
                # Create drive point Coherence
                for i, index in enumerate(corresponding_drive_responses):
                    for j, index in enumerate(corresponding_drive_responses):
                        if i <= j:
                            continue
                        widget = self.new_window()
                        widget.signal_selector.setCurrentIndex(7)
                        widget.response_coordinate_selector.setCurrentIndex(i)
                        widget.reference_coordinate_selector.setCurrentIndex(j)
                        widget.lock_response_checkbox.setChecked(True)
        self.run_widget.new_from_template_combobox.setCurrentIndex(0)

    def close_windows(self):
        """Closes all existing windows"""
        for window in self.run_widget.channel_display_area.subWindowList():
            window.close()

    def decrement_channels(self):
        """Decrements the unlocked window response channels by the specified number of channels"""
        number = -self.run_widget.increment_channels_number.value()
        for window in self.run_widget.channel_display_area.subWindowList():
            window.widget().increment_channel(number)

    def increment_channels(self):
        """Increments the unlocked window response channels by the specified number of channels"""
        number = self.run_widget.increment_channels_number.value()
        for window in self.run_widget.channel_display_area.subWindowList():
            window.widget().increment_channel(number)

    def add_override_channel(self):
        """Adds a row to the channel override table"""
        selected_row = self.run_widget.dof_override_table.blockSignals(True)
        selected_row = self.run_widget.dof_override_table.rowCount()
        self.run_widget.dof_override_table.insertRow(selected_row)
        channel_combobox = QtWidgets.QComboBox()
        for channel_name in self.channel_names:
            channel_combobox.addItem(channel_name)
        channel_combobox.currentIndexChanged.connect(self.update_override_table)
        self.run_widget.dof_override_table.setCellWidget(selected_row, 0, channel_combobox)
        data_item = QtWidgets.QTableWidgetItem()
        data_item.setText("1")
        self.run_widget.dof_override_table.setItem(selected_row, 1, data_item)
        data_item = QtWidgets.QTableWidgetItem()
        data_item.setText("X+")
        self.run_widget.dof_override_table.setItem(selected_row, 2, data_item)
        selected_row = self.run_widget.dof_override_table.blockSignals(False)
        self.update_override_table()

    def remove_override_channel(self):
        """Removes a row from the channel override table"""
        selected_row = self.run_widget.dof_override_table.currentRow()
        if selected_row >= 0:
            self.run_widget.dof_override_table.removeRow(selected_row)
        self.update_override_table()

    def update_override_table(self):
        """Updates channel information in the test based on the override table values"""
        self.override_table = {}
        for row in range(self.run_widget.dof_override_table.rowCount()):
            index = self.run_widget.dof_override_table.cellWidget(row, 0).currentIndex()
            new_node = self.run_widget.dof_override_table.item(row, 1).text()
            new_direction = self.run_widget.dof_override_table.item(row, 2).text()
            self.override_table[index] = [new_node, new_direction]
        self.update_channel_names()
        self.run_widget.channel_display_area.reciprocal_responses = self.get_reciprocal_measurements()
        # Go through and update all the existing windows in the MDI display
        for window in self.run_widget.channel_display_area.subWindowList():
            widget = window.widget()
            current_response = widget.response_coordinate_selector.currentIndex()
            current_reference = widget.reference_coordinate_selector.currentIndex()
            current_data_type = widget.data_type_selector.currentIndex()
            widget.channel_names = self.channel_names
            widget.reference_names = [self.channel_names[i] for i in self.run_widget.channel_display_area.reference_channel_indices]
            widget.response_names = [self.channel_names[i] for i in self.run_widget.channel_display_area.response_channel_indices]
            widget.reciprocal_responses = self.run_widget.channel_display_area.reciprocal_responses
            widget.update_ui()
            widget.response_coordinate_selector.setCurrentIndex(current_response)
            widget.reference_coordinate_selector.setCurrentIndex(current_reference)
            widget.data_type_selector.setCurrentIndex(current_data_type)

    def get_reciprocal_measurements(self):
        """Finds all reciprocal measurements in the test"""
        node_numbers = np.array(
            [
                (channel.node_number if i not in self.override_table else self.override_table[i][0])
                for i, channel in enumerate(self.hardware_metadata.channel_list)
            ]
        )
        node_directions = np.array(
            [
                (
                    ""
                    if channel.node_direction is None
                    else "".join(
                        [char for char in (channel.node_direction if i not in self.override_table else self.override_table[i][1]) if char not in "+-"]
                    )
                )
                for i, channel in enumerate(self.hardware_metadata.channel_list)
            ]
        )
        reference_node_numbers = node_numbers[self.metadata.reference_channel_indices]
        reference_node_directions = node_directions[self.metadata.reference_channel_indices]
        response_node_numbers = node_numbers[self.metadata.response_channel_indices]
        response_node_directions = node_directions[self.metadata.response_channel_indices]
        corresponding_drive_responses = []
        for node, direction in zip(reference_node_numbers, reference_node_directions):
            # print('Node: {:} Direction: {:}'.format(node,direction))
            # print('Response Node Numbers:')
            # print(response_node_numbers)
            # print('Response Node Directions:')
            # print(response_node_directions)
            # print('Node Match:')
            # print(response_node_numbers == node)
            # print('Direction Match:')
            # print(response_node_directions == direction)
            index = np.where((response_node_numbers == node) & (response_node_directions == direction))[0]
            # print('Index:')
            # print(index)
            if len(index) == 0:
                corresponding_drive_responses.append(None)
                print(f"Warning: No Drive Point Found for Reference {node}{direction}")
            elif len(index) > 1:
                corresponding_drive_responses.append(None)
                print(f"Warning: Multiple Drive Points Found for Reference {node}{direction}")
            else:
                corresponding_drive_responses.append(index[0])
        # print(corresponding_drive_responses)
        return corresponding_drive_responses

    def create_netcdf_file(self, filename):
        """Creates an output NetCDF4 file to save modal data to

        Parameters
        ----------
        filename : str
            The file name to which the netCDF4 file will be stored
        """
        self.netcdf_handle = nc4.Dataset(filename, "w", format="NETCDF4", clobber=True)  # pylint: disable=no-member
        # Create dimensions
        self.netcdf_handle.createDimension("response_channels", len(self.hardware_metadata.channel_list))
        self.netcdf_handle.createDimension(
            "output_channels",
            len([channel for channel in self.hardware_metadata.channel_list if channel.feedback_device is not None]),
        )
        self.netcdf_handle.createDimension("num_environments", len(self.hardware_metadata.environment_names))
        self.netcdf_handle.createDimension("time_samples", None)
        # Create attributes
        self.netcdf_handle.sample_rate = self.hardware_metadata.sample_rate
        self.netcdf_handle.time_per_write = self.hardware_metadata.samples_per_write / self.hardware_metadata.output_sample_rate
        self.netcdf_handle.time_per_read = self.hardware_metadata.samples_per_read / self.hardware_metadata.sample_rate
        self.netcdf_handle.hardware = self.hardware_metadata.hardware
        self.netcdf_handle.hardware_file = "None" if self.hardware_metadata.hardware_file is None else self.hardware_metadata.hardware_file
        self.netcdf_handle.output_oversample = self.hardware_metadata.output_oversample
        for name, value in self.hardware_metadata.extra_parameters.items():
            setattr(self.netcdf_handle, name, value)
        # Create Variables
        self.netcdf_handle.createVariable("time_data", "f8", ("response_channels", "time_samples"))
        var = self.netcdf_handle.createVariable("environment_names", str, ("num_environments",))
        this_environment_index = None
        for i, name in enumerate(self.hardware_metadata.environment_names):
            var[i] = name
            if name == self.environment_name:
                this_environment_index = i
        var = self.netcdf_handle.createVariable(
            "environment_active_channels",
            "i1",
            ("response_channels", "num_environments"),
        )
        var[...] = self.hardware_metadata.environment_active_channels.astype("int8")[
            self.hardware_metadata.environment_active_channels[:, this_environment_index],
            :,
        ]
        # Create channel table variables
        labels = [
            ["node_number", str],
            ["node_direction", str],
            ["comment", str],
            ["serial_number", str],
            ["triax_dof", str],
            ["sensitivity", str],
            ["unit", str],
            ["make", str],
            ["model", str],
            ["expiration", str],
            ["physical_device", str],
            ["physical_channel", str],
            ["channel_type", str],
            ["minimum_value", str],
            ["maximum_value", str],
            ["coupling", str],
            ["excitation_source", str],
            ["excitation", str],
            ["feedback_device", str],
            ["feedback_channel", str],
            ["warning_level", str],
            ["abort_level", str],
        ]
        for label, netcdf_datatype in labels:
            var = self.netcdf_handle.createVariable("/channels/" + label, netcdf_datatype, ("response_channels",))
            channel_data = [getattr(channel, label) for channel in self.hardware_metadata.channel_list]
            if netcdf_datatype == "i1":
                channel_data = np.array([1 if val else 0 for val in channel_data])
            else:
                channel_data = ["" if val is None else val for val in channel_data]
            for i, cd in enumerate(channel_data):
                if label == "node_number" and i in self.override_table:
                    var[i] = self.override_table[i][0]
                elif label == "node_direction" and i in self.override_table:
                    var[i] = self.override_table[i][1]
                else:
                    var[i] = cd
        group_handle = self.netcdf_handle.createGroup(self.environment_name)
        self.metadata.store_to_netcdf(group_handle)
        group_handle.createDimension("fft_lines", self.metadata.fft_lines)
        group_handle.createVariable(
            "frf_data_real",
            "f8",
            ("fft_lines", "response_channels", "reference_channels"),
        )
        group_handle.createVariable(
            "frf_data_imag",
            "f8",
            ("fft_lines", "response_channels", "reference_channels"),
        )
        group_handle.createVariable("coherence", "f8", ("fft_lines", "response_channels"))

    def update_channel_names(self):
        """Updates channel names based on the override channel table"""
        self.channel_names = []
        for i, channel in enumerate(self.hardware_metadata.channel_list):
            channel_type_str = "" if channel.channel_type is None else channel.channel_type
            node_num_str = channel.node_number if i not in self.override_table else self.override_table[i][0]
            node_dir_str = channel.node_direction if i not in self.override_table else self.override_table[i][1]
            self.channel_names.append(f"{channel_type_str} {node_num_str} {node_dir_str}"[:MAXIMUM_NAME_LENGTH])
        self.run_widget.channel_display_area.channel_names = self.channel_names

    def initialize_hardware(self, hardware_metadata: HardwareMetadata):
        """Update the user interface with data acquisition parameters

        This function is called when the Data Acquisition parameters are
        initialized.  This function should set up the environment user interface
        accordingly.

        Parameters
        ----------
        hardware_metadata : DataAcquisitionParameters :
            Container containing the data acquisition parameters, including
            channel table and sampling information.

        """
        self.hardware_metadata = hardware_metadata
        self.definition_widget.sample_rate_display.setValue(hardware_metadata.sample_rate)
        self.all_output_channel_indices = [
            index for index, channel in enumerate(self.hardware_metadata.channel_list) if channel.feedback_device is not None
        ]
        self.update_channel_names()
        self.definition_widget.reference_channels_selector.setRowCount(0)
        self.definition_widget.trigger_channel_selector.blockSignals(True)
        self.definition_widget.trigger_channel_selector.clear()
        for i, channel_name in enumerate(self.channel_names):
            self.definition_widget.trigger_channel_selector.addItem(channel_name)
            self.definition_widget.reference_channels_selector.insertRow(i)
            item = QtWidgets.QTableWidgetItem()
            item.setText(channel_name)
            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
            self.definition_widget.reference_channels_selector.setItem(i, 2, item)
            ref_checkbox = QtWidgets.QCheckBox()
            ref_checkbox.stateChanged.connect(self.update_reference_channels)
            self.definition_widget.reference_channels_selector.setCellWidget(i, 1, ref_checkbox)
            enabled_checkbox = QtWidgets.QCheckBox()
            enabled_checkbox.setChecked(True)
            enabled_checkbox.stateChanged.connect(self.update_reference_channels)
            self.definition_widget.reference_channels_selector.setCellWidget(i, 0, enabled_checkbox)
        self.definition_widget.trigger_channel_selector.blockSignals(False)
        self.update_trigger_levels()

        checked_state = self.definition_widget.regenerate_signal_auto_checkbox.isChecked()
        self.definition_widget.regenerate_signal_auto_checkbox.setChecked(False)
        self.definition_widget.signal_generator_selector.setCurrentIndex(0)
        self.definition_widget.samples_per_frame_selector.setValue(hardware_metadata.sample_rate)
        self.definition_widget.random_max_frequency_selector.setValue(hardware_metadata.sample_rate / 2)
        self.definition_widget.random_min_frequency_selector.setValue(0)
        self.definition_widget.burst_max_frequency_selector.setValue(hardware_metadata.sample_rate / 2)
        self.definition_widget.burst_min_frequency_selector.setValue(0)
        self.definition_widget.chirp_max_frequency_selector.setValue(hardware_metadata.sample_rate / 2)
        self.definition_widget.chirp_min_frequency_selector.setValue(0)
        self.definition_widget.pseudorandom_max_frequency_selector.setValue(hardware_metadata.sample_rate / 2)
        self.definition_widget.pseudorandom_min_frequency_selector.setValue(0)

        self.definition_widget.response_channels_display.setValue(len(self.channel_names))
        self.definition_widget.reference_channels_display.setValue(0)
        num_outputs = len(self.output_channel_indices)
        self.definition_widget.output_channels_display.setValue(num_outputs)
        if num_outputs == 0:
            for i in range(self.definition_widget.signal_generator_selector.count() - 1):
                self.definition_widget.signal_generator_selector.setTabEnabled(i + 1, False)

        self.definition_widget.output_signal_plot.getPlotItem().clear()
        self.plot_data_items["signal_representation"] = multiline_plotter(
            (0, 1),
            np.zeros((len(self.all_output_channel_indices), 2)),
            widget=self.definition_widget.output_signal_plot,
            other_pen_options={"width": 1},
            names=[f"Output {i + 1}" for i in range(len(self.all_output_channel_indices))],
        )
        self.definition_widget.regenerate_signal_auto_checkbox.setChecked(checked_state)
        if checked_state:
            self.generate_signal()

        for widget in [
            self.definition_widget.random_min_frequency_selector,
            self.definition_widget.random_max_frequency_selector,
            self.definition_widget.burst_min_frequency_selector,
            self.definition_widget.burst_max_frequency_selector,
            self.definition_widget.pseudorandom_min_frequency_selector,
            self.definition_widget.pseudorandom_max_frequency_selector,
            self.definition_widget.chirp_min_frequency_selector,
            self.definition_widget.chirp_max_frequency_selector,
            self.definition_widget.square_frequency_selector,
            self.definition_widget.sine_frequency_selector,
        ]:
            widget.setMaximum(self.hardware_metadata.sample_rate / 2)

    def initialize_environment(self, environment_metadata):
        return

    def get_environment_metadata(self, global_channel_list):
        if self.hardware_metadata:
            channel_list_bools = self.get_channel_list_bools(global_channel_list)

        signal_generator_level = 0
        signal_generator_min_frequency = 0
        signal_generator_max_frequency = 0
        signal_generator_on_percent = 0
        if self.definition_widget.signal_generator_selector.currentIndex() == 0:  # None
            signal_generator_type = "none"
        elif self.definition_widget.signal_generator_selector.currentIndex() == 1:  # Random
            signal_generator_type = "random"
            signal_generator_level = self.definition_widget.random_rms_selector.value()
            signal_generator_min_frequency = self.definition_widget.random_min_frequency_selector.value()
            signal_generator_max_frequency = self.definition_widget.random_max_frequency_selector.value()
        elif self.definition_widget.signal_generator_selector.currentIndex() == 2:  # Burst Random
            signal_generator_type = "burst"
            signal_generator_level = self.definition_widget.burst_rms_selector.value()
            signal_generator_min_frequency = self.definition_widget.burst_min_frequency_selector.value()
            signal_generator_max_frequency = self.definition_widget.burst_max_frequency_selector.value()
            signal_generator_on_percent = self.definition_widget.burst_on_percentage_selector.value()
        elif self.definition_widget.signal_generator_selector.currentIndex() == 3:  # Pseudorandom
            signal_generator_type = "pseudorandom"
            signal_generator_level = self.definition_widget.pseudorandom_rms_selector.value()
            signal_generator_min_frequency = self.definition_widget.pseudorandom_min_frequency_selector.value()
            signal_generator_max_frequency = self.definition_widget.pseudorandom_max_frequency_selector.value()
        elif self.definition_widget.signal_generator_selector.currentIndex() == 4:  # Chirp
            signal_generator_type = "chirp"
            signal_generator_level = self.definition_widget.chirp_level_selector.value()
            signal_generator_min_frequency = self.definition_widget.chirp_min_frequency_selector.value()
            signal_generator_max_frequency = self.definition_widget.chirp_max_frequency_selector.value()
        elif self.definition_widget.signal_generator_selector.currentIndex() == 5:  # Square
            signal_generator_type = "square"
            signal_generator_level = self.definition_widget.square_level_selector.value()
            signal_generator_min_frequency = self.definition_widget.square_frequency_selector.value()
            signal_generator_on_percent = self.definition_widget.square_percent_on_selector.value()
        elif self.definition_widget.signal_generator_selector.currentIndex() == 6:  # Sine
            signal_generator_type = "sine"
            signal_generator_level = self.definition_widget.sine_level_selector.value()
            signal_generator_min_frequency = self.definition_widget.sine_frequency_selector.value()
        else:
            index = self.definition_widget.signal_generator_selector.currentIndex()
            raise ValueError(f"Invalid Signal Generator {index} (How did you get here?)")
        return ModalMetadata(
            self.environment_name,
            channel_list_bools,
            self.definition_widget.sample_rate_display.value(),
            self.definition_widget.samples_per_frame_selector.value(),
            self.definition_widget.system_id_averaging_scheme_selector.itemText(
                self.definition_widget.system_id_averaging_scheme_selector.currentIndex()
            ),
            self.definition_widget.system_id_frames_to_average_selector.value(),
            self.definition_widget.system_id_averaging_coefficient_selector.value(),
            self.definition_widget.system_id_frf_technique_selector.itemText(self.definition_widget.system_id_frf_technique_selector.currentIndex()),
            self.definition_widget.system_id_transfer_function_computation_window_selector.itemText(
                self.definition_widget.system_id_transfer_function_computation_window_selector.currentIndex()
            ).lower(),
            self.definition_widget.system_id_overlap_percentage_selector.value(),
            self.definition_widget.triggering_type_selector.itemText(self.definition_widget.triggering_type_selector.currentIndex()),
            self.definition_widget.acceptance_selector.itemText(self.definition_widget.acceptance_selector.currentIndex()),
            self.definition_widget.wait_for_steady_selector.value(),
            self.definition_widget.trigger_channel_selector.currentIndex(),
            self.definition_widget.pretrigger_selector.value(),
            self.definition_widget.trigger_slope_selector.currentIndex() == 0,
            self.definition_widget.trigger_level_selector.value(),
            self.definition_widget.hysteresis_selector.value(),
            self.definition_widget.hysteresis_length_selector.value(),
            signal_generator_type,
            signal_generator_level,
            signal_generator_min_frequency,
            signal_generator_max_frequency,
            signal_generator_on_percent,
            self.acceptance_function,
            self.reference_indices,
            self.response_indices,
            self.all_output_channel_indices,
            self.hardware_metadata,
            self.definition_widget.window_value_selector.value() / 100,
        )

    def set_environment_metadata(self, metadata: ModalMetadata) -> EnvironmentMetadata:
        """
        Update the user interface with environment parameters

        This function is called when the Environment parameters are initialized.
        This function should set up the user interface accordingly.  It must
        return the parameters class of the environment that inherits from
        AbstractMetadata.

        Returns
        ModalMetadata
            An AbstractMetadata-inheriting object that contains the parameters
            defining the environment.

        """
        self.metadata = metadata
        self.reference_channel_indices = self.metadata.reference_channel_indices
        self.response_channel_indices = self.metadata.response_channel_indices
        self.run_widget.channel_display_area.reference_channel_indices = self.reference_channel_indices
        self.run_widget.channel_display_area.response_channel_indices = self.response_channel_indices
        for window in self.run_widget.channel_display_area.subWindowList():
            widget = window.widget()
            current_response = widget.response_coordinate_selector.currentIndex()
            current_reference = widget.reference_coordinate_selector.currentIndex()
            current_data_type = widget.data_type_selector.currentIndex()
            widget.reference_names = np.array([widget.channel_names[i] for i in self.run_widget.channel_display_area.reference_channel_indices])
            widget.response_names = np.array([widget.channel_names[i] for i in self.run_widget.channel_display_area.response_channel_indices])
            widget.update_ui()
            widget.response_coordinate_selector.setCurrentIndex(current_response)
            widget.reference_coordinate_selector.setCurrentIndex(current_reference)
            widget.data_type_selector.setCurrentIndex(current_data_type)
        self.run_widget.total_averages_display.setValue(self.metadata.num_averages)
        self.run_widget.channel_display_area.time_abscissa = np.arange(self.metadata.samples_per_frame) / self.metadata.sample_rate
        self.run_widget.channel_display_area.frequency_abscissa = np.fft.rfftfreq(
            self.metadata.samples_per_frame,
            1 / self.metadata.sample_rate,
        )
        if self.metadata.frf_window == "rectangle":
            window = 1
        elif self.metadata.frf_window == "exponential":
            window_parameter = -(self.metadata.samples_per_frame) / np.log(self.metadata.exponential_window_value_at_frame_end)
            window = sig.get_window(
                ("exponential", 0, window_parameter),
                self.metadata.samples_per_frame,
                fftbins=True,
            )
        else:
            window = sig.get_window(
                self.metadata.frf_window,
                self.metadata.samples_per_frame,
                fftbins=True,
            )
        self.run_widget.channel_display_area.window_function = window
        self.run_widget.channel_display_area.reciprocal_responses = self.get_reciprocal_measurements()

    def update_gui(self, queue_data: tuple):
        """Update the environment's graphical user interface

        This function will receive data from the gui_update_queue that
        specifies how the user interface should be updated.  Data will usually
        be received as ``(instruction,data)`` pairs, where the ``instruction`` notes
        what operation should be taken or which widget should be modified, and
        the ``data`` notes what data should be used in the update.

        Parameters
        ----------
        queue_data : tuple
            A tuple containing ``(instruction,data)`` pairs where ``instruction``
            defines and operation or widget to be modified and ``data`` contains
            the data used to perform the operation.
        """
        # print('Got GUI Update {:}'.format(queue_data[0]))
        message, data = queue_data
        if message == ModalUICommands.SPECTRAL_UPDATE:
            (
                frames,
                _,
                _,
                self.last_frf,
                self.last_coherence,
                last_response_cpsd,
                last_reference_cpsd,
                self.last_condition,
            ) = data
            self.run_widget.channel_display_area.last_frf = self.last_frf
            self.run_widget.channel_display_area.last_coh = self.last_coherence.T
            if last_response_cpsd.ndim == 3:
                self.last_response_cpsd = np.einsum("fii->fi", last_response_cpsd)
            else:
                self.last_response_cpsd = last_response_cpsd
            if last_reference_cpsd.ndim == 3:
                self.last_reference_cpsd = np.einsum("fii->fi", last_reference_cpsd)
            else:
                self.last_reference_cpsd = last_reference_cpsd
            # Assemble autospectrum
            self.run_widget.channel_display_area.last_autospectrum = np.zeros(
                (
                    len(self.hardware_metadata.channel_list),
                    self.last_response_cpsd.shape[0],
                )
            )
            for i, index in enumerate(self.metadata.reference_channel_indices):
                self.run_widget.channel_display_area.last_autospectrum[index, :] = self.last_reference_cpsd[:, i].real
            for i, index in enumerate(self.metadata.response_channel_indices):
                self.run_widget.channel_display_area.last_autospectrum[index, :] = self.last_response_cpsd[:, i].real
            self.run_widget.current_average_display.setValue(frames)
            for window in self.run_widget.channel_display_area.subWindowList():
                widget = window.widget()
                if widget.signal_selector.currentIndex() in [3, 4, 5, 6, 7]:
                    widget.update_data()
            if self.acquiring and self.netcdf_handle is not None:
                group = self.netcdf_handle.groups[self.environment_name]
                group.variables["frf_data_real"][:] = np.real(self.last_frf)
                group.variables["frf_data_imag"][:] = np.imag(self.last_frf)
                group.variables["coherence"][:] = self.last_coherence
            if self.acquiring and frames >= self.metadata.num_averages:
                self.stop_control()
                self.acquiring = False

    @staticmethod
    def create_environment_template(environment_name: str, workbook: openpyxl.workbook.workbook.Workbook):
        """Creates a template worksheet in an Excel workbook defining the
        environment.

        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the
        environment.

        This function is the "write" counterpart to the
        ``set_parameters_from_template`` function in the ``ModalUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.workbook.workbook.Workbook :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1, 1, "Control Type")
        worksheet.cell(1, 2, "Modal")
        worksheet.cell(2, 1, "Samples Per Frame:")
        worksheet.cell(2, 2, "# Number of Samples per Measurement Frame")
        worksheet.cell(3, 1, "Averaging Type:")
        worksheet.cell(3, 2, "# Averaging Type")
        worksheet.cell(4, 1, "Number of Averages:")
        worksheet.cell(4, 2, "# Number of Averages used when computing the FRF")
        worksheet.cell(5, 1, "Averaging Coefficient:")
        worksheet.cell(5, 2, "# Averaging Coefficient for Exponential Averaging")
        worksheet.cell(6, 1, "FRF Technique:")
        worksheet.cell(6, 2, "# FRF Technique")
        worksheet.cell(7, 1, "FRF Window:")
        worksheet.cell(7, 2, "# Window used to compute FRF")
        worksheet.cell(8, 1, "Exponential Window End Value:")
        worksheet.cell(
            8,
            2,
            "# Exponential Window Value at the end of the measurement frame (0.5 or 50%, not 50)",
        )
        worksheet.cell(9, 1, "FRF Overlap:")
        worksheet.cell(9, 2, "# Overlap for FRF calculations (0.5 or 50%, not 50)")
        worksheet.cell(10, 1, "Triggering Type:")
        worksheet.cell(10, 2, '# One of "Free Run", "First Frame", or "Every Frame"')
        worksheet.cell(11, 1, "Average Acceptance:")
        worksheet.cell(11, 2, '# One of "Accept All", "Manual", or "Autoreject"')
        worksheet.cell(12, 1, "Trigger Channel")
        worksheet.cell(12, 2, "# Channel number (1-based) to use for triggering")
        worksheet.cell(13, 1, "Pretrigger")
        worksheet.cell(13, 2, "# Amount of frame to use as pretrigger (0.5 or 50%, not 50)")
        worksheet.cell(14, 1, "Trigger Slope")
        worksheet.cell(14, 2, '# One of "Positive" or "Negative"')
        worksheet.cell(15, 1, "Trigger Level")
        worksheet.cell(
            15,
            2,
            "# Level to use to trigger the test as a fraction of the total range of the channel " "(0.5 or 50%, not 50)",
        )
        worksheet.cell(16, 1, "Hysteresis Level")
        worksheet.cell(
            16,
            2,
            "# Level that a channel must fall below before another trigger can be considered " "(0.5 or 50%, not 50)",
        )
        worksheet.cell(17, 1, "Hysteresis Frame Fraction")
        worksheet.cell(
            17,
            2,
            "# Fraction of the frame that a channel maintain hysteresis condition before another " "trigger can be considered (0.5 or 50%, not 50)",
        )
        worksheet.cell(18, 1, "Signal Generator Type")
        worksheet.cell(
            18,
            2,
            '# One of "None", "Random", "Burst Random", "Pseudorandom", "Chirp", "Square", or ' '"Sine"',
        )
        worksheet.cell(19, 1, "Signal Generator Level")
        worksheet.cell(
            19,
            2,
            "# RMS voltage level for random signals, Peak voltage level for chirp, sine, and " "square pulse",
        )
        worksheet.cell(20, 1, "Signal Generator Frequency 1")
        worksheet.cell(
            20,
            2,
            "# Minimum frequency for broadband signals or frequency for sine and square pulse",
        )
        worksheet.cell(21, 1, "Signal Generator Frequency 2")
        worksheet.cell(
            21,
            2,
            "# Maximum frequency for broadband signals.  Ignored for sine and square pulse",
        )
        worksheet.cell(22, 1, "Signal Generator On Fraction")
        worksheet.cell(
            22,
            2,
            "# Fraction of time that the burst or square wave is on (0.5 or 50%, not 50)",
        )
        worksheet.cell(23, 1, "Wait Time for Steady State")
        worksheet.cell(
            23,
            2,
            "# Time to wait after output starts to allow the system to reach steady state",
        )
        worksheet.cell(24, 1, "Autoaccept Script")
        worksheet.cell(24, 2, "# File in which an autoacceptance function is defined")
        worksheet.cell(25, 1, "Autoaccept Function")
        worksheet.cell(25, 2, "# Function name in which the autoacceptance function is defined")
        worksheet.cell(26, 1, "Reference Channels")
        worksheet.cell(26, 2, "# List of channels, one per cell on this row")
        worksheet.cell(27, 1, "Disabled Channels")
        worksheet.cell(27, 2, "# List of channels, one per cell on this row")

    def set_parameters_from_template(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        """
        Collects parameters for the user interface from the Excel template file

        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.

        This function is the "read" counterpart to the
        ``create_environment_template`` function in the ``ModalUI`` class,
        which writes a template file that can be filled out by a user.


        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        self.definition_widget.samples_per_frame_selector.setValue(worksheet.cell(2, 2).value)
        self.definition_widget.system_id_averaging_scheme_selector.setCurrentIndex(
            self.definition_widget.system_id_averaging_scheme_selector.findText(worksheet.cell(3, 2).value)
        )
        self.definition_widget.system_id_frames_to_average_selector.setValue(worksheet.cell(4, 2).value)
        self.definition_widget.system_id_averaging_coefficient_selector.setValue(worksheet.cell(5, 2).value)
        self.definition_widget.system_id_frf_technique_selector.setCurrentIndex(
            self.definition_widget.system_id_frf_technique_selector.findText(worksheet.cell(6, 2).value)
        )
        self.definition_widget.system_id_transfer_function_computation_window_selector.setCurrentIndex(
            self.definition_widget.system_id_transfer_function_computation_window_selector.findText(worksheet.cell(7, 2).value)
        )
        self.definition_widget.window_value_selector.setValue(worksheet.cell(8, 2).value * 100)
        self.definition_widget.system_id_overlap_percentage_selector.setValue(worksheet.cell(9, 2).value * 100)
        self.definition_widget.triggering_type_selector.setCurrentIndex(
            self.definition_widget.triggering_type_selector.findText(worksheet.cell(10, 2).value)
        )
        acceptance = worksheet.cell(11, 2).value
        self.definition_widget.acceptance_selector.blockSignals(True)
        if acceptance == "Autoreject":
            self.definition_widget.acceptance_selector.setCurrentIndex(2)
            self.acceptance_function = [
                worksheet.cell(24, 2).value,
                worksheet.cell(25, 2).value,
            ]
        else:
            self.definition_widget.acceptance_selector.setCurrentIndex(self.definition_widget.acceptance_selector.findText(acceptance))
            self.acceptance_function = None
        self.definition_widget.acceptance_selector.blockSignals(False)
        self.definition_widget.trigger_channel_selector.setCurrentIndex(worksheet.cell(12, 2).value - 1)
        self.definition_widget.pretrigger_selector.setValue(worksheet.cell(13, 2).value * 100)
        self.definition_widget.trigger_slope_selector.setCurrentIndex(
            self.definition_widget.trigger_slope_selector.findText(worksheet.cell(14, 2).value)
        )
        self.definition_widget.trigger_level_selector.setValue(worksheet.cell(15, 2).value * 100)
        self.definition_widget.hysteresis_selector.setValue(worksheet.cell(16, 2).value * 100)
        self.definition_widget.hysteresis_length_selector.setValue(worksheet.cell(17, 2).value * 100)
        signal_index = [
            "None",
            "Random",
            "Burst Random",
            "Pseudorandom",
            "Chirp",
            "Square",
            "Sine",
        ].index(worksheet.cell(18, 2).value)
        self.definition_widget.signal_generator_selector.setCurrentIndex(signal_index)
        level = worksheet.cell(19, 2).value
        freq_1 = worksheet.cell(20, 2).value
        freq_2 = worksheet.cell(21, 2).value
        sig_on = worksheet.cell(22, 2).value * 100
        for widget in [
            self.definition_widget.random_rms_selector,
            self.definition_widget.burst_rms_selector,
            self.definition_widget.pseudorandom_rms_selector,
            self.definition_widget.chirp_level_selector,
            self.definition_widget.square_level_selector,
            self.definition_widget.sine_level_selector,
        ]:
            widget.setValue(level)
        for widget in [
            self.definition_widget.random_min_frequency_selector,
            self.definition_widget.burst_min_frequency_selector,
            self.definition_widget.pseudorandom_min_frequency_selector,
            self.definition_widget.chirp_min_frequency_selector,
            self.definition_widget.square_frequency_selector,
            self.definition_widget.sine_frequency_selector,
        ]:
            widget.setValue(freq_1)
        for widget in [
            self.definition_widget.random_max_frequency_selector,
            self.definition_widget.burst_max_frequency_selector,
            self.definition_widget.pseudorandom_max_frequency_selector,
            self.definition_widget.chirp_max_frequency_selector,
        ]:
            widget.setValue(freq_2)
        for widget in [
            self.definition_widget.burst_on_percentage_selector,
            self.definition_widget.square_percent_on_selector,
        ]:
            widget.setValue(sig_on)
        self.definition_widget.wait_for_steady_selector.setValue(worksheet.cell(23, 2).value)
        column_index = 2
        while True:
            value = worksheet.cell(26, column_index).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(value) - 1, 1)
            widget.setChecked(True)
            column_index += 1
        for i in range(self.definition_widget.reference_channels_selector.rowCount()):
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(i), 0)
            widget.setChecked(True)
        column_index = 2
        while True:
            value = worksheet.cell(27, column_index).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            widget = self.definition_widget.reference_channels_selector.cellWidget(int(value) - 1, 0)
            widget.setChecked(False)
            column_index += 1
