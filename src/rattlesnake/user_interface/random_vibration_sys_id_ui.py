from rattlesnake.user_interface.abstract_sys_id_user_interface import AbstractSysIdUI
from rattlesnake.utilities import (
    DataAcquisitionParameters,
    GlobalCommands,
    VerboseMessageQueue,
    db2scale,
    error_message_qt,
    load_python_module,
)
from rattlesnake.environment.random_vibration_sys_id_environment import (
    RandomVibrationMetadata,
    RandomVibrationCommands,
)
from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.user_interface.ui_utilities import (
    environment_definition_ui_paths,
    environment_prediction_ui_paths,
    environment_run_ui_paths,
)
from rattlesnake.user_interface.random_vibration_sys_id_ui_utilities import (
    _direction_map,
    load_specification,
)
from rattlesnake.user_interface.ui_utilities import (
    PlotWindow,
    TransformationMatrixWindow,
    multiline_plotter,
)
from rattlesnake.environment.abstract_interactive_control_law import (  # noqa: E402 pylint: disable=wrong-import-position
    AbstractControlLawComputation,
)
from qtpy import QtWidgets, uic
from qtpy.QtCore import Qt, QTimer
from qtpy.QtGui import QColor
from multiprocessing.queues import Queue
import datetime
import time
import inspect
import numpy as np
import openpyxl
import netCDF4 as nc4

CONTROL_TYPE = ControlTypes.RANDOM
MAXIMUM_NAME_LENGTH = 50


# region: User Interface
class RandomVibrationUI(AbstractSysIdUI):
    """Class defining the user interface for a Random Vibration environment.

    This class will contain four main UIs, the environment definition,
    system identification, test prediction, and run.  The widgets corresponding
    to these interfaces are stored in TabWidgets in the main UI.

    This class defines all the call backs and user interface operations required
    for the Random Vibration environment."""

    def __init__(
        self,
        environment_name: str,
        definition_tabwidget: QtWidgets.QTabWidget,
        system_id_tabwidget: QtWidgets.QTabWidget,
        test_predictions_tabwidget: QtWidgets.QTabWidget,
        run_tabwidget: QtWidgets.QTabWidget,
        environment_command_queue: VerboseMessageQueue,
        controller_communication_queue: VerboseMessageQueue,
        log_file_queue: Queue,
    ):
        """
        Constructs a Random Vibration User Interface

        Given the tab widgets from the main interface as well as communication
        queues, this class assembles the user interface components specific to
        the Random Vibration Environment

        Parameters
        ----------
        definition_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Control
            Definition main tab
        system_id_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the System
            Identification main tab
        test_predictions_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Test Predictions
            main tab
        run_tabwidget : QtWidgets.QTabWidget
            QTabWidget containing the environment subtabs on the Run
            main tab.
        environment_command_queue : VerboseMessageQueue
            Queue for sending commands to the Random Vibration Environment
        controller_communication_queue : VerboseMessageQueue
            Queue for sending global commands to the controller
        log_file_queue : Queue
            Queue where log file messages can be written.

        """
        super().__init__(
            environment_name,
            environment_command_queue,
            controller_communication_queue,
            log_file_queue,
            system_id_tabwidget,
        )
        # Add the page to the control definition tabwidget
        self.definition_widget = QtWidgets.QWidget()
        uic.loadUi(environment_definition_ui_paths[CONTROL_TYPE], self.definition_widget)
        definition_tabwidget.addTab(self.definition_widget, self.environment_name)
        # Add the page to the control prediction tabwidget
        self.prediction_widget = QtWidgets.QWidget()
        uic.loadUi(environment_prediction_ui_paths[CONTROL_TYPE], self.prediction_widget)
        test_predictions_tabwidget.addTab(self.prediction_widget, self.environment_name)
        # Add the page to the run tabwidget
        self.run_widget = QtWidgets.QWidget()
        uic.loadUi(environment_run_ui_paths[CONTROL_TYPE], self.run_widget)
        run_tabwidget.addTab(self.run_widget, self.environment_name)

        self.plot_data_items = {}
        self.plot_windows = []
        self.run_start_time = None
        self.run_level_start_time = None
        self.run_timer = QTimer()
        self.response_transformation_matrix = None
        self.output_transformation_matrix = None
        self.python_control_module = None
        self.specification_frequency_lines = None
        self.specification_cpsd_matrix = None
        self.specification_warning_matrix = None
        self.specification_abort_matrix = None
        self.physical_channel_names = None
        self.physical_output_indices = None
        self.excitation_prediction = None
        self.response_prediction = None
        self.rms_voltage_prediction = None
        self.rms_db_error_prediction = None
        self.interactive_control_law_widget = None
        self.interactive_control_law_window = None
        self.control_selector_widgets = [
            self.definition_widget.specification_row_selector,
            self.definition_widget.specification_column_selector,
            self.prediction_widget.response_row_selector,
            self.prediction_widget.response_column_selector,
            self.run_widget.control_channel_1_selector,
            self.run_widget.control_channel_2_selector,
        ]
        self.output_selector_widgets = [
            self.prediction_widget.excitation_row_selector,
            self.prediction_widget.excitation_column_selector,
        ]
        self.system_id_widget.samplesPerFrameSpinBox.setReadOnly(True)
        self.system_id_widget.samplesPerFrameSpinBox.setButtonSymbols(
            QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons
        )
        self.system_id_widget.levelRampTimeDoubleSpinBox.setReadOnly(True)
        self.system_id_widget.levelRampTimeDoubleSpinBox.setButtonSymbols(
            QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons
        )

        # Set common look and feel for plots
        plot_widgets = [
            self.definition_widget.specification_single_plot,
            self.definition_widget.specification_sum_asds_plot,
            self.prediction_widget.excitation_display_plot,
            self.prediction_widget.response_display_plot,
            self.run_widget.global_test_performance_plot,
        ]
        for plot_widget in plot_widgets:
            plot_item = plot_widget.getPlotItem()
            plot_item.showGrid(True, True, 0.25)
            plot_item.enableAutoRange()
            plot_item.getViewBox().enableAutoRange(enable=True)
        logscale_plot_widgets = [
            self.definition_widget.specification_single_plot,
            self.definition_widget.specification_sum_asds_plot,
            self.prediction_widget.excitation_display_plot,
            self.prediction_widget.response_display_plot,
            self.run_widget.global_test_performance_plot,
        ]
        for plot_widget in logscale_plot_widgets:
            plot_item = plot_widget.getPlotItem()
            plot_item.setLogMode(False, True)

        self.connect_callbacks()

        # Complete the profile commands
        self.command_map["Set Test Level"] = self.change_test_level_from_profile
        self.command_map["Change Specification"] = self.change_specification_from_profile
        self.command_map["Save Control Data"] = self.save_control_data_from_profile

    def connect_callbacks(self):
        """Connects callback functions to the UI Widgets"""
        # Definition
        self.definition_widget.samples_per_frame_selector.valueChanged.connect(
            self.update_parameters_and_clear_spec
        )
        self.definition_widget.cpsd_overlap_selector.valueChanged.connect(self.update_parameters)
        self.definition_widget.cola_overlap_percentage_selector.valueChanged.connect(
            self.update_parameters
        )
        self.definition_widget.transformation_matrices_button.clicked.connect(
            self.define_transformation_matrices
        )
        self.definition_widget.control_script_load_file_button.clicked.connect(
            self.select_python_module
        )
        self.definition_widget.control_function_input.currentIndexChanged.connect(
            self.update_generator_selector
        )
        self.definition_widget.load_spec_button.clicked.connect(self.select_spec_file)
        self.definition_widget.specification_row_selector.currentIndexChanged.connect(
            self.show_specification
        )
        self.definition_widget.specification_column_selector.currentIndexChanged.connect(
            self.show_specification
        )
        self.definition_widget.control_channels_selector.itemChanged.connect(
            self.update_control_channels
        )
        self.definition_widget.check_selected_button.clicked.connect(
            self.check_selected_control_channels
        )
        self.definition_widget.uncheck_selected_button.clicked.connect(
            self.uncheck_selected_control_channels
        )
        # Prediction
        self.prediction_widget.excitation_row_selector.currentIndexChanged.connect(
            self.update_control_predictions
        )
        self.prediction_widget.excitation_column_selector.currentIndexChanged.connect(
            self.update_control_predictions
        )
        self.prediction_widget.response_row_selector.currentIndexChanged.connect(
            self.update_control_predictions
        )
        self.prediction_widget.response_column_selector.currentIndexChanged.connect(
            self.update_control_predictions
        )
        self.prediction_widget.maximum_voltage_button.clicked.connect(
            self.show_max_voltage_prediction
        )
        self.prediction_widget.minimum_voltage_button.clicked.connect(
            self.show_min_voltage_prediction
        )
        self.prediction_widget.maximum_error_button.clicked.connect(self.show_max_error_prediction)
        self.prediction_widget.minimum_error_button.clicked.connect(self.show_min_error_prediction)
        self.prediction_widget.response_error_list.itemClicked.connect(
            self.update_response_error_prediction_selector
        )
        self.prediction_widget.excitation_voltage_list.itemClicked.connect(
            self.update_excitation_prediction_selector
        )
        self.prediction_widget.recompute_prediction_button.clicked.connect(
            self.recompute_prediction
        )
        # Run Test
        self.run_widget.current_test_level_selector.valueChanged.connect(
            self.change_control_test_level
        )
        self.run_widget.start_test_button.clicked.connect(self.start_control)
        self.run_widget.stop_test_button.clicked.connect(self.stop_control)
        self.run_widget.create_window_button.clicked.connect(self.create_window)
        self.run_widget.show_all_asds_button.clicked.connect(self.show_all_asds)
        self.run_widget.show_all_csds_phscoh_button.clicked.connect(self.show_all_csds_phscoh)
        self.run_widget.show_all_csds_realimag_button.clicked.connect(self.show_all_csds_realimag)
        self.run_widget.tile_windows_button.clicked.connect(self.tile_windows)
        self.run_widget.close_windows_button.clicked.connect(self.close_windows)
        self.run_timer.timeout.connect(self.update_run_time)
        self.run_widget.test_response_error_list.itemDoubleClicked.connect(
            self.show_magnitude_window
        )
        self.run_widget.save_current_spectral_data_button.clicked.connect(self.save_spectral_data)

    # %% Initialize Data Aquisition

    def initialize_data_acquisition(self, data_acquisition_parameters: DataAcquisitionParameters):
        """Update the user interface with data acquisition parameters

        This function is called when the Data Acquisition parameters are
        initialized.  This function should set up the environment user interface
        accordingly.

        Parameters
        ----------
        data_acquisition_parameters : DataAcquisitionParameters :
            Container containing the data acquisition parameters, including
            channel table and sampling information.

        """
        super().initialize_data_acquisition(data_acquisition_parameters)
        # Initialize the plots
        # Clear plots if there is anything on them
        self.definition_widget.specification_single_plot.getPlotItem().clear()
        self.definition_widget.specification_sum_asds_plot.getPlotItem().clear()
        self.run_widget.global_test_performance_plot.getPlotItem().clear()

        # Now add initial lines that we can update later
        self.definition_widget.specification_single_plot.getPlotItem().addLegend()
        self.plot_data_items[
            "specification_real"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": "b", "width": 1},
            name="Real Part",
        )
        self.plot_data_items[
            "specification_imag"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": "r", "width": 1},
            name="Imaginary Part",
        )
        self.plot_data_items[
            "specification_warning_upper"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": PlotWindow.WARNING_COLOR, "width": 0.25},
            name="Warning",
        )
        self.plot_data_items[
            "specification_warning_lower"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": PlotWindow.WARNING_COLOR, "width": 0.25},
        )
        self.plot_data_items[
            "specification_abort_upper"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": PlotWindow.ABORT_COLOR, "width": 0.25},
            name="Abort",
        )
        self.plot_data_items[
            "specification_abort_lower"
        ] = self.definition_widget.specification_single_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": PlotWindow.ABORT_COLOR, "width": 0.25},
        )
        self.plot_data_items[
            "specification_sum"
        ] = self.definition_widget.specification_sum_asds_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": "b", "width": 1},
        )
        self.run_widget.global_test_performance_plot.getPlotItem().addLegend()
        self.plot_data_items[
            "specification_sum_control"
        ] = self.run_widget.global_test_performance_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": "b", "width": 1},
            name="Specification",
        )
        self.plot_data_items[
            "sum_asds_control"
        ] = self.run_widget.global_test_performance_plot.getPlotItem().plot(
            np.array([0, data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={"color": "r", "width": 1},
            name="Response",
        )

        # Set up channel names
        self.physical_channel_names = [
            (
                f"{'' if channel.channel_type is None else channel.channel_type} "
                f"{channel.node_number} "
                f"{'' if channel.node_direction is None else channel.node_direction}"
            )[:MAXIMUM_NAME_LENGTH]
            for channel in data_acquisition_parameters.channel_list
        ]
        self.physical_output_indices = [
            i
            for i, channel in enumerate(data_acquisition_parameters.channel_list)
            if channel.feedback_device
        ]
        # Set up widgets
        self.definition_widget.sample_rate_display.setValue(data_acquisition_parameters.sample_rate)
        self.definition_widget.samples_per_frame_selector.setValue(
            data_acquisition_parameters.sample_rate
        )
        self.definition_widget.control_channels_selector.clear()
        for channel_name in self.physical_channel_names:
            item = QtWidgets.QListWidgetItem()
            item.setText(channel_name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.definition_widget.control_channels_selector.addItem(item)
        self.definition_widget.input_channels_display.setValue(len(self.physical_channel_names))
        self.definition_widget.output_channels_display.setValue(len(self.physical_output_indices))
        self.definition_widget.control_channels_display.setValue(0)
        self.response_transformation_matrix = None
        self.output_transformation_matrix = None
        self.define_transformation_matrices(None, False)

    @property
    def physical_output_names(self):
        """Names of the physical output channels"""
        return [self.physical_channel_names[i] for i in self.physical_output_indices]

    # %% Define Environments

    @property
    def physical_control_indices(self):
        """Indices corresponding to the physical channels that are used as outputs"""
        return [
            i
            for i in range(self.definition_widget.control_channels_selector.count())
            if self.definition_widget.control_channels_selector.item(i).checkState() == Qt.Checked
        ]

    @property
    def physical_control_names(self):
        """Names of the physical control channels"""
        return [self.physical_channel_names[i] for i in self.physical_control_indices]

    @property
    def initialized_control_names(self):
        if self.environment_parameters.response_transformation_matrix is None:
            return [
                self.physical_channel_names[i]
                for i in self.environment_parameters.control_channel_indices
            ]
        else:
            return [
                f"Transformed Response {i + 1}"
                for i in range(self.environment_parameters.response_transformation_matrix.shape[0])
            ]

    @property
    def initialized_output_names(self):
        if self.environment_parameters.reference_transformation_matrix is None:
            return self.physical_output_names
        else:
            return [
                f"Transformed Drive {i + 1}"
                for i in range(self.environment_parameters.reference_transformation_matrix.shape[0])
            ]

    def select_spec_file(self, clicked, filename=None):  # pylint: disable=unused-argument
        """Loads a specification using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the specification for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename, _ = QtWidgets.QFileDialog.getOpenFileName(
                self.definition_widget,
                "Select Specification File",
                filter="Numpyz or Mat (*.npz *.mat)",
            )
            if filename == "":
                return
        self.definition_widget.specification_file_name_display.setText(filename)
        coord_dtype = np.dtype([("node", "<u8"), ("direction", "i1")])
        if self.response_transformation_matrix is not None:
            control_coordinate = None
        else:
            control_coordinate = np.array(
                [
                    (
                        self.data_acquisition_parameters.channel_list[i].node_number,
                        _direction_map[
                            self.data_acquisition_parameters.channel_list[i].node_direction
                        ],
                    )
                    for i in self.physical_control_indices
                ],
                dtype=coord_dtype,
            )
        try:
            (
                self.specification_frequency_lines,
                self.specification_cpsd_matrix,
                self.specification_warning_matrix,
                self.specification_abort_matrix,
            ) = load_specification(
                filename,
                self.definition_widget.fft_lines_display.value(),
                self.definition_widget.frequency_spacing_display.value(),
                control_coordinate,
            )
        except ValueError as e:
            error_message_qt(type(e).__name__, str(e))
            return

        if np.all(np.isnan(self.specification_abort_matrix)):
            self.definition_widget.auto_abort_checkbox.setChecked(False)
            self.definition_widget.auto_abort_checkbox.setEnabled(False)
        else:
            self.definition_widget.auto_abort_checkbox.setEnabled(True)
        self.show_specification()

    def select_python_module(self, clicked, filename=None):  # pylint: disable=unused-argument
        """Loads a Python module using a dialog or the specified filename

        Parameters
        ----------
        clicked :
            The clicked event that triggered the callback.
        filename :
            File name defining the Python module for bypassing the callback when
            loading from a file (Default value = None).

        """
        if filename is None:
            filename, _ = QtWidgets.QFileDialog.getOpenFileName(
                self.definition_widget,
                "Select Python Module",
                filter="Python Modules (*.py)",
            )
            if filename == "":
                return
        self.python_control_module = load_python_module(filename)
        functions = [
            function
            for function in inspect.getmembers(self.python_control_module)
            if (
                inspect.isfunction(function[1])
                and len(inspect.signature(function[1]).parameters) >= 12
            )
            or inspect.isgeneratorfunction(function[1])
            or (
                inspect.isclass(function[1])
                and all(
                    [
                        (
                            method in function[1].__dict__
                            and not (
                                hasattr(function[1].__dict__[method], "__isabstractmethod__")
                                and function[1].__dict__[method].__isabstractmethod__
                            )
                        )
                        for method in ["system_id_update", "control"]
                    ]
                )
            )
        ]
        self.log(
            f"Loaded module {self.python_control_module.__name__} with "
            f"functions {[function[0] for function in functions]}"
        )
        self.definition_widget.control_function_input.clear()
        self.definition_widget.control_script_file_path_input.setText(filename)
        for function in functions:
            self.definition_widget.control_function_input.addItem(function[0])

    def update_generator_selector(self):
        """Updates the function/generator selector based on the function selected"""
        if self.python_control_module is None:
            return
        try:
            function = getattr(
                self.python_control_module,
                self.definition_widget.control_function_input.itemText(
                    self.definition_widget.control_function_input.currentIndex()
                ),
            )
        except AttributeError:
            return
        if inspect.isgeneratorfunction(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(1)
        elif inspect.isclass(function) and issubclass(function, AbstractControlLawComputation):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(3)
        elif inspect.isclass(function):
            self.definition_widget.control_function_generator_selector.setCurrentIndex(2)
        else:
            self.definition_widget.control_function_generator_selector.setCurrentIndex(0)

    def show_specification(self):
        """Show the specification on the GUI"""
        if self.specification_cpsd_matrix is None:
            self.plot_data_items["specification_real"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_imag"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_sum"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_warning_upper"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_warning_lower"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_abort_upper"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["specification_abort_lower"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            # enabled_state = self.run_widget.isEnabled()
            # self.run_widget.setEnabled(True)
            self.plot_data_items["specification_sum_control"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            # self.run_widget.setEnabled(enabled_state)
        else:
            row = self.definition_widget.specification_row_selector.currentIndex()
            column = self.definition_widget.specification_column_selector.currentIndex()
            spec_real = abs(self.specification_cpsd_matrix[:, row, column].real)
            spec_imag = abs(self.specification_cpsd_matrix[:, row, column].imag)
            spec_sum = abs(
                np.nansum(
                    self.specification_cpsd_matrix[
                        :,
                        np.arange(self.specification_cpsd_matrix.shape[-1]),
                        np.arange(self.specification_cpsd_matrix.shape[-1]),
                    ],
                    axis=-1,
                )
            )
            self.plot_data_items["specification_real"].setData(
                self.specification_frequency_lines[spec_real > 0.0],
                spec_real[spec_real > 0.0],
            )
            self.plot_data_items["specification_imag"].setData(
                self.specification_frequency_lines[spec_imag > 0.0],
                spec_imag[spec_imag > 0.0],
            )
            if row == column:
                warning_upper = abs(self.specification_warning_matrix[1, :, row])
                warning_lower = abs(self.specification_warning_matrix[0, :, row])
                abort_upper = abs(self.specification_abort_matrix[1, :, row])
                abort_lower = abs(self.specification_abort_matrix[0, :, row])
                self.plot_data_items["specification_warning_upper"].setData(
                    self.specification_frequency_lines, warning_upper
                )
                self.plot_data_items["specification_warning_lower"].setData(
                    self.specification_frequency_lines, warning_lower
                )
                self.plot_data_items["specification_abort_upper"].setData(
                    self.specification_frequency_lines, abort_upper
                )
                self.plot_data_items["specification_abort_lower"].setData(
                    self.specification_frequency_lines, abort_lower
                )
            else:
                self.plot_data_items["specification_warning_upper"].setData(
                    np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                    np.zeros(2),
                )
                self.plot_data_items["specification_warning_lower"].setData(
                    np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                    np.zeros(2),
                )
                self.plot_data_items["specification_abort_upper"].setData(
                    np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                    np.zeros(2),
                )
                self.plot_data_items["specification_abort_lower"].setData(
                    np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                    np.zeros(2),
                )
            self.plot_data_items["specification_sum"].setData(
                self.specification_frequency_lines[spec_sum > 0.0],
                spec_sum[spec_sum > 0.0],
            )
            # enabled_state = self.run_widget.isEnabled()
            # self.run_widget.setEnabled(True)
            self.plot_data_items["specification_sum_control"].setData(
                self.specification_frequency_lines[spec_sum > 0.0],
                spec_sum[spec_sum > 0.0],
            )
            # self.run_widget.setEnabled(enabled_state)

    def check_selected_control_channels(self):
        """Checks the selected channels to make them control channels"""
        for item in self.definition_widget.control_channels_selector.selectedItems():
            item.setCheckState(Qt.Checked)

    def uncheck_selected_control_channels(self):
        """Unchecks the selected channels to make them no longer control channels"""
        for item in self.definition_widget.control_channels_selector.selectedItems():
            item.setCheckState(Qt.Unchecked)

    def update_control_channels(self):
        """Resets the definition UI when the number of control channels has changed"""
        self.response_transformation_matrix = None
        self.output_transformation_matrix = None
        self.specification_abort_matrix = None
        self.specification_warning_matrix = None
        self.specification_cpsd_matrix = None
        self.specification_frequency_lines = None
        self.definition_widget.control_channels_display.setValue(len(self.physical_control_indices))
        self.definition_widget.specification_row_selector.blockSignals(True)
        self.definition_widget.specification_column_selector.blockSignals(True)
        self.definition_widget.specification_row_selector.clear()
        self.definition_widget.specification_column_selector.clear()
        for i, control_name in enumerate(self.physical_control_names):
            self.definition_widget.specification_row_selector.addItem(f"{i + 1}: {control_name}")
            self.definition_widget.specification_column_selector.addItem(f"{i + 1}: {control_name}")
        self.definition_widget.specification_row_selector.blockSignals(False)
        self.definition_widget.specification_column_selector.blockSignals(False)
        self.define_transformation_matrices(None, False)
        self.show_specification()

    def define_transformation_matrices(
        self, clicked, dialog=True
    ):  # pylint: disable=unused-argument
        """Defines the transformation matrices using the dialog box"""
        if dialog:
            (response_transformation, output_transformation, result) = (
                TransformationMatrixWindow.define_transformation_matrices(
                    self.response_transformation_matrix,
                    self.definition_widget.control_channels_display.value(),
                    self.output_transformation_matrix,
                    self.definition_widget.output_channels_display.value(),
                    self.definition_widget,
                )
            )
        else:
            response_transformation = self.response_transformation_matrix
            output_transformation = self.output_transformation_matrix
            result = True
        if result:
            # Update the control names
            for widget in self.control_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if response_transformation is None:
                for i, control_name in enumerate(self.physical_control_names):
                    for widget in self.control_selector_widgets:
                        widget.addItem(f"{i + 1}: {control_name}")
                self.definition_widget.transform_channels_display.setValue(
                    len(self.physical_control_names)
                )
            else:
                for i in range(response_transformation.shape[0]):
                    for widget in self.control_selector_widgets:
                        widget.addItem(f"{i + 1}: Transformed Response")
                self.definition_widget.transform_channels_display.setValue(
                    response_transformation.shape[0]
                )
            for widget in self.control_selector_widgets:
                widget.blockSignals(False)
            # Update the output names
            for widget in self.output_selector_widgets:
                widget.blockSignals(True)
                widget.clear()
            if output_transformation is None:
                for i, drive_name in enumerate(self.physical_output_names):
                    for widget in self.output_selector_widgets:
                        widget.addItem(f"{i + 1}: {drive_name}")
                self.definition_widget.transform_outputs_display.setValue(
                    len(self.physical_output_names)
                )
            else:
                for i in range(output_transformation.shape[0]):
                    for widget in self.output_selector_widgets:
                        widget.addItem(f"{i + 1}: Transformed Drive")
                self.definition_widget.transform_outputs_display.setValue(
                    output_transformation.shape[0]
                )
            for widget in self.output_selector_widgets:
                widget.blockSignals(False)

            self.response_transformation_matrix = response_transformation
            self.output_transformation_matrix = output_transformation
            self.update_parameters_and_clear_spec()

    def update_parameters(self):
        """Recompute derived parameters from updated sampling parameters"""
        data = self.collect_environment_definition_parameters()
        self.definition_widget.samples_per_acquire_display.setValue(data.samples_per_acquire)
        self.definition_widget.frame_time_display.setValue(data.frame_time)
        self.definition_widget.nyquist_frequency_display.setValue(data.nyquist_frequency)
        self.definition_widget.fft_lines_display.setValue(data.fft_lines)
        self.definition_widget.frequency_spacing_display.setValue(data.frequency_spacing)
        self.definition_widget.samples_per_write_display.setValue(data.samples_per_output)

    def update_parameters_and_clear_spec(self):
        """Clears the specification data and updates parameters"""
        samples_per_frame = self.definition_widget.samples_per_frame_selector.value()
        if samples_per_frame % 2 != 0:
            self.definition_widget.samples_per_frame_selector.blockSignals(True)
            self.definition_widget.samples_per_frame_selector.setValue(samples_per_frame + 1)
            self.definition_widget.samples_per_frame_selector.blockSignals(False)
        self.specification_frequency_lines = None
        self.specification_cpsd_matrix = None
        self.specification_warning_matrix = None
        self.specification_abort_matrix = None
        self.definition_widget.specification_file_name_display.setText("")
        self.show_specification()
        self.update_parameters()

    def collect_environment_definition_parameters(self) -> RandomVibrationMetadata:
        """
        Collect the parameters from the user interface defining the environment

        Returns
        -------
        RandomVibrationMetadata
            A metadata or parameters object containing the parameters defining
            the corresponding environment.

        """
        if self.python_control_module is None:
            control_module = None
            control_function = None
            control_function_type = None
            control_function_parameters = None
        else:
            control_module = self.definition_widget.control_script_file_path_input.text()
            control_function = self.definition_widget.control_function_input.itemText(
                self.definition_widget.control_function_input.currentIndex()
            )
            control_function_type = (
                self.definition_widget.control_function_generator_selector.currentIndex()
            )
            control_function_parameters = (
                self.definition_widget.control_parameters_text_input.toPlainText()
            )
        return RandomVibrationMetadata(
            number_of_channels=len(self.data_acquisition_parameters.channel_list),
            sample_rate=self.definition_widget.sample_rate_display.value(),
            samples_per_frame=self.definition_widget.samples_per_frame_selector.value(),
            test_level_ramp_time=self.definition_widget.ramp_time_spinbox.value(),
            cola_window=self.definition_widget.cola_window_selector.itemText(
                self.definition_widget.cola_window_selector.currentIndex()
            ),
            cola_overlap=self.definition_widget.cola_overlap_percentage_selector.value() / 100,
            cola_window_exponent=self.definition_widget.cola_exponent_selector.value(),
            sigma_clip=self.definition_widget.sigma_clipping_selector.value(),
            update_tf_during_control=self.definition_widget.update_transfer_function_during_control_selector.isChecked(),
            frames_in_cpsd=self.definition_widget.cpsd_frames_selector.value(),
            cpsd_window=self.definition_widget.cpsd_computation_window_selector.itemText(
                self.definition_widget.cpsd_computation_window_selector.currentIndex()
            ),
            cpsd_overlap=self.definition_widget.cpsd_overlap_selector.value() / 100,
            response_transformation_matrix=self.response_transformation_matrix,
            output_transformation_matrix=self.output_transformation_matrix,
            control_python_script=control_module,
            control_python_function=control_function,
            control_python_function_type=control_function_type,
            control_python_function_parameters=control_function_parameters,
            control_channel_indices=self.physical_control_indices,
            output_channel_indices=self.physical_output_indices,
            specification_frequency_lines=self.specification_frequency_lines,
            specification_cpsd_matrix=self.specification_cpsd_matrix,
            specification_warning_matrix=self.specification_warning_matrix,
            specification_abort_matrix=self.specification_abort_matrix,
            percent_lines_out=self.definition_widget.frequency_lines_out_spinbox.value(),
            allow_automatic_aborts=self.definition_widget.auto_abort_checkbox.isChecked(),
        )

    def initialize_environment(self) -> RandomVibrationMetadata:
        """
        Update the user interface with environment parameters

        This function is called when the Environment parameters are initialized.
        This function should set up the user interface accordingly.  It must
        return the parameters class of the environment that inherits from
        AbstractMetadata.

        Returns
        -------
        AbstractMetadata
            An AbstractMetadata-inheriting object that contains the parameters
            defining the environment.

        """
        self.system_id_widget.samplesPerFrameSpinBox.setMaximum(
            self.definition_widget.samples_per_frame_selector.value()
        )
        self.system_id_widget.samplesPerFrameSpinBox.setValue(
            self.definition_widget.samples_per_frame_selector.value()
        )
        self.system_id_widget.levelRampTimeDoubleSpinBox.setValue(
            self.definition_widget.ramp_time_spinbox.value()
        )
        super().initialize_environment()
        for widget in [
            self.prediction_widget.response_row_selector,
            self.prediction_widget.response_column_selector,
            self.run_widget.control_channel_1_selector,
            self.run_widget.control_channel_2_selector,
        ]:
            widget.blockSignals(True)
            widget.clear()
            for i, control_name in enumerate(self.initialized_control_names):
                widget.addItem(f"{i + 1}: {control_name}")
            widget.blockSignals(False)
        for widget in [
            self.prediction_widget.excitation_row_selector,
            self.prediction_widget.excitation_column_selector,
        ]:
            widget.blockSignals(True)
            widget.clear()
            for i, drive_name in enumerate(self.initialized_output_names):
                widget.addItem(f"{i + 1}: {drive_name}")
            widget.blockSignals(False)
        # Set up the prediction plots
        self.prediction_widget.excitation_display_plot.getPlotItem().clear()
        self.prediction_widget.response_display_plot.getPlotItem().clear()
        self.prediction_widget.excitation_display_plot.getPlotItem().addLegend()
        self.prediction_widget.response_display_plot.getPlotItem().addLegend()
        self.plot_data_items["response_prediction"] = multiline_plotter(
            np.arange(self.environment_parameters.fft_lines)
            * self.environment_parameters.frequency_spacing,
            np.zeros((4, self.environment_parameters.fft_lines)),
            widget=self.prediction_widget.response_display_plot,
            other_pen_options={"width": 2},
            names=["Real Prediction", "Real Spec", "Imag Prediction", "Imag Spec"],
        )
        self.plot_data_items[
            "prediction_warning_upper"
        ] = self.prediction_widget.response_display_plot.getPlotItem().plot(
            np.array([0, self.data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={
                "color": PlotWindow.WARNING_COLOR,
                "width": PlotWindow.WARNING_LINEWIDTH,
                "style": PlotWindow.WARNING_LINESTYLE,
            },
            name="Warning",
        )
        self.plot_data_items[
            "prediction_warning_lower"
        ] = self.prediction_widget.response_display_plot.getPlotItem().plot(
            np.array([0, self.data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={
                "color": PlotWindow.WARNING_COLOR,
                "width": PlotWindow.WARNING_LINEWIDTH,
                "style": PlotWindow.WARNING_LINESTYLE,
            },
        )
        self.plot_data_items[
            "prediction_abort_upper"
        ] = self.prediction_widget.response_display_plot.getPlotItem().plot(
            np.array([0, self.data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={
                "color": PlotWindow.ABORT_COLOR,
                "width": PlotWindow.ABORT_LINEWIDTH,
                "style": PlotWindow.ABORT_LINESTYLE,
            },
            name="Abort",
        )
        self.plot_data_items[
            "prediction_abort_lower"
        ] = self.prediction_widget.response_display_plot.getPlotItem().plot(
            np.array([0, self.data_acquisition_parameters.sample_rate / 2]),
            np.zeros(2),
            pen={
                "color": PlotWindow.ABORT_COLOR,
                "width": PlotWindow.ABORT_LINEWIDTH,
                "style": PlotWindow.ABORT_LINESTYLE,
            },
        )
        self.plot_data_items["excitation_prediction"] = multiline_plotter(
            np.arange(self.environment_parameters.fft_lines)
            * self.environment_parameters.frequency_spacing,
            np.zeros((2, self.environment_parameters.fft_lines)),
            widget=self.prediction_widget.excitation_display_plot,
            other_pen_options={"width": 1},
            names=["Real Prediction", "Imag Prediction"],
        )
        # Create the interactive control law if necessary
        if self.definition_widget.control_function_generator_selector.currentIndex() == 3:
            control_class = getattr(
                self.python_control_module,
                self.definition_widget.control_function_input.itemText(
                    self.definition_widget.control_function_input.currentIndex()
                ),
            )
            self.log(f"Building Interactive UI for class {control_class.__name__}")
            ui_class = control_class.get_ui_class()
            if ui_class == self.interactive_control_law_widget.__class__:
                print("initializing data acquisition and environment parameters")
                self.interactive_control_law_widget.initialize_parameters(
                    self.data_acquisition_parameters, self.environment_parameters
                )
            else:
                if self.interactive_control_law_widget is not None:
                    self.interactive_control_law_widget.close()
                self.interactive_control_law_window = QtWidgets.QDialog(self.definition_widget)
                self.interactive_control_law_widget = ui_class(
                    self.log_name,
                    self.environment_command_queue,
                    self.interactive_control_law_window,
                    self,
                    self.data_acquisition_parameters,
                    self.environment_parameters,
                )
            self.interactive_control_law_window.show()
        return self.environment_parameters

    # %% Test Predictions

    def show_max_voltage_prediction(self):
        """Shows the prediction with the largest RMS voltage"""
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmax([float(widget.item(v).text()) for v in range(widget.count())])
        self.prediction_widget.excitation_row_selector.setCurrentIndex(index)
        self.prediction_widget.excitation_column_selector.setCurrentIndex(index)

    def show_min_voltage_prediction(self):
        """Shows the prediction with the smallest RMS voltage"""
        widget = self.prediction_widget.excitation_voltage_list
        index = np.argmin([float(widget.item(v).text()) for v in range(widget.count())])
        self.prediction_widget.excitation_row_selector.setCurrentIndex(index)
        self.prediction_widget.excitation_column_selector.setCurrentIndex(index)

    def show_max_error_prediction(self):
        """Shows the prediction with the largest error"""
        widget = self.prediction_widget.response_error_list
        index = np.argmax([float(widget.item(v).text()) for v in range(widget.count())])
        self.prediction_widget.response_row_selector.setCurrentIndex(index)
        self.prediction_widget.response_column_selector.setCurrentIndex(index)

    def show_min_error_prediction(self):
        """Shows the prediction with the smallest error"""
        widget = self.prediction_widget.response_error_list
        index = np.argmin([float(widget.item(v).text()) for v in range(widget.count())])
        self.prediction_widget.response_row_selector.setCurrentIndex(index)
        self.prediction_widget.response_column_selector.setCurrentIndex(index)

    def update_response_error_prediction_selector(self, item):
        """Updates the selection when an item is double-clicked"""
        index = self.prediction_widget.response_error_list.row(item)
        self.prediction_widget.response_row_selector.setCurrentIndex(index)
        self.prediction_widget.response_column_selector.setCurrentIndex(index)

    def update_excitation_prediction_selector(self, item):
        """Updates the selection when an item is double-clicked"""
        index = self.prediction_widget.excitation_voltage_list.row(item)
        self.prediction_widget.excitation_row_selector.setCurrentIndex(index)
        self.prediction_widget.excitation_column_selector.setCurrentIndex(index)

    def update_control_predictions(self):
        """Updates the control prediction with new data"""
        excite_row_index = self.prediction_widget.excitation_row_selector.currentIndex()
        excite_column_index = self.prediction_widget.excitation_column_selector.currentIndex()
        self.plot_data_items["excitation_prediction"][0].setData(
            self.frequencies,
            np.abs(np.real(self.excitation_prediction[:, excite_row_index, excite_column_index])),
        )
        row_index = self.prediction_widget.response_row_selector.currentIndex()
        column_index = self.prediction_widget.response_column_selector.currentIndex()
        self.plot_data_items["response_prediction"][0].setData(
            self.frequencies,
            np.abs(np.real(self.response_prediction[:, row_index, column_index])),
        )
        if row_index == column_index:
            warning_upper = abs(
                self.environment_parameters.specification_warning_matrix[1, :, row_index]
            )
            warning_lower = abs(
                self.environment_parameters.specification_warning_matrix[0, :, row_index]
            )
            abort_upper = abs(
                self.environment_parameters.specification_abort_matrix[1, :, row_index]
            )
            abort_lower = abs(
                self.environment_parameters.specification_abort_matrix[0, :, row_index]
            )
            self.plot_data_items["prediction_warning_upper"].setData(
                self.specification_frequency_lines, warning_upper
            )
            self.plot_data_items["prediction_warning_lower"].setData(
                self.specification_frequency_lines, warning_lower
            )
            self.plot_data_items["prediction_abort_upper"].setData(
                self.specification_frequency_lines, abort_upper
            )
            self.plot_data_items["prediction_abort_lower"].setData(
                self.specification_frequency_lines, abort_lower
            )
            self.plot_data_items["excitation_prediction"][1].setData(
                self.frequencies, np.zeros(self.frequencies.shape)
            )
            self.plot_data_items["response_prediction"][2].setData(
                self.frequencies, np.zeros(self.frequencies.shape)
            )
            self.plot_data_items["response_prediction"][3].setData(
                self.frequencies, np.zeros(self.frequencies.shape)
            )
        else:
            self.plot_data_items["prediction_warning_upper"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["prediction_warning_lower"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["prediction_abort_upper"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["prediction_abort_lower"].setData(
                np.array([0, self.definition_widget.sample_rate_display.value() / 2]),
                np.zeros(2),
            )
            self.plot_data_items["excitation_prediction"][1].setData(
                self.frequencies,
                np.abs(
                    np.imag(self.excitation_prediction[:, excite_row_index, excite_column_index])
                ),
            )
            self.plot_data_items["response_prediction"][2].setData(
                self.frequencies,
                np.abs(np.imag(self.response_prediction[:, row_index, column_index])),
            )
            self.plot_data_items["response_prediction"][3].setData(
                self.frequencies,
                np.abs(
                    np.imag(
                        self.environment_parameters.specification_cpsd_matrix[
                            :, row_index, column_index
                        ]
                    )
                ),
            )
        self.plot_data_items["response_prediction"][1].setData(
            self.frequencies,
            np.abs(
                np.real(
                    self.environment_parameters.specification_cpsd_matrix[
                        :, row_index, column_index
                    ]
                )
            ),
        )

    def recompute_prediction(self):
        """Sends a message to the environment process to recompute the prediction"""
        self.environment_command_queue.put(
            self.log_name, (RandomVibrationCommands.RECOMPUTE_PREDICTION, None)
        )

    # %% Run Control

    def start_control(self):
        """Runs the corresponding environment in the controller"""
        self.enable_control(False)
        self.controller_communication_queue.put(
            self.log_name, (GlobalCommands.START_ENVIRONMENT, self.environment_name)
        )
        self.environment_command_queue.put(
            self.log_name,
            (
                RandomVibrationCommands.START_CONTROL,
                db2scale(self.run_widget.current_test_level_selector.value()),
            ),
        )
        self.run_timer.start(250)
        self.run_start_time = time.time()
        self.run_level_start_time = self.run_start_time
        self.run_widget.test_progress_bar.setValue(0)
        if (
            self.run_widget.current_test_level_selector.value()
            >= self.run_widget.target_test_level_selector.value()
        ):
            self.controller_communication_queue.put(
                self.log_name, (GlobalCommands.AT_TARGET_LEVEL, self.environment_name)
            )

    def stop_control(self):
        """Stops the corresponding environment in the controller"""
        self.run_widget.stop_test_button.setEnabled(False)
        self.environment_command_queue.put(
            self.log_name, (RandomVibrationCommands.STOP_CONTROL, None)
        )
        self.run_timer.stop()

    def enable_control(self, enabled):
        """Enables or disables widgets to start or stop control if the control is running or not"""
        for widget in [
            self.run_widget.test_time_selector,
            self.run_widget.time_test_at_target_level_checkbox,
            self.run_widget.timed_test_radiobutton,
            self.run_widget.continuous_test_radiobutton,
            self.run_widget.target_test_level_selector,
            self.run_widget.start_test_button,
        ]:
            widget.setEnabled(enabled)
        for widget in [self.run_widget.stop_test_button]:
            widget.setEnabled(not enabled)
        if enabled:
            self.run_timer.stop()

    def update_run_time(self):
        """Updates the time that the control has been running on the GUI"""
        # Update the total run time
        current_time = time.time()
        time_elapsed = current_time - self.run_start_time
        time_at_level_elapsed = current_time - self.run_level_start_time
        self.run_widget.total_test_time_display.setText(
            str(datetime.timedelta(seconds=time_elapsed)).split(".", maxsplit=1)[0]
        )
        self.run_widget.time_at_level_display.setText(
            str(datetime.timedelta(seconds=time_at_level_elapsed)).split(".", maxsplit=1)[0]
        )
        # Check if we need to stop the test due to timeout
        if self.run_widget.timed_test_radiobutton.isChecked():
            check_time = self.run_widget.test_time_selector.time()
            check_time_seconds = (
                check_time.hour() * 3600 + check_time.minute() * 60 + check_time.second()
            )
            if self.run_widget.time_test_at_target_level_checkbox.isChecked():
                if (
                    self.run_widget.current_test_level_selector.value()
                    >= self.run_widget.target_test_level_selector.value()
                ):
                    self.run_widget.test_progress_bar.setValue(
                        int(time_at_level_elapsed / check_time_seconds * 100)
                    )
                    if time_at_level_elapsed > check_time_seconds:
                        self.run_widget.test_progress_bar.setValue(100)
                        self.stop_control()
                else:
                    self.run_widget.test_progress_bar.setValue(0)
            else:
                self.run_widget.test_progress_bar.setValue(
                    int(time_elapsed / check_time_seconds * 100)
                )
                if time_elapsed > check_time_seconds:
                    self.stop_control()

    def change_control_test_level(self):
        """Updates the test level of the control."""
        self.environment_command_queue.put(
            self.log_name,
            (
                RandomVibrationCommands.ADJUST_TEST_LEVEL,
                db2scale(self.run_widget.current_test_level_selector.value()),
            ),
        )
        self.run_level_start_time = time.time()
        # Check and see if we need to start streaming data
        if (
            self.run_widget.current_test_level_selector.value()
            >= self.run_widget.target_test_level_selector.value()
        ):
            self.controller_communication_queue.put(
                self.log_name, (GlobalCommands.AT_TARGET_LEVEL, self.environment_name)
            )

    def change_test_level_from_profile(self, test_level):
        """Sets the test level from a profile instruction

        Parameters
        ----------
        test_level :
            Value to set the test level to.
        """
        self.run_widget.current_test_level_selector.setValue(float(test_level))

    def change_specification_from_profile(self, new_specification_file):
        """
        Loads in a new specification and starts controlling to it

        Parameters
        ----------
        new_specification_file : str
            File path to a new specification file

        """
        self.select_spec_file(None, new_specification_file)
        environment_parameters = self.initialize_environment()
        self.environment_command_queue.put(
            self.log_name,
            (GlobalCommands.INITIALIZE_ENVIRONMENT_PARAMETERS, environment_parameters),
        )

    def show_magnitude_window(self, item):
        """Creates a window showing the magnitude of a signal when an item is double-clicked"""
        index = self.run_widget.test_response_error_list.row(item)
        self.create_window(None, index, index, 0)

    def create_window(
        self, event, row_index=None, column_index=None, datatype_index=None
    ):  # pylint: disable=unused-argument
        """Creates a subwindow to show a specific channel information

        Parameters
        ----------
        event :

        row_index :
            Row index in the CPSD matrix to display (Default value = None)
        column_index :
            Column index in the CPSD matrix to display (Default value = None)
        datatype_index :
            Data type to display (real,imag,mag,phase,etc) (Default value = None)

        """
        if row_index is None:
            row_index = self.run_widget.control_channel_1_selector.currentIndex()
        if column_index is None:
            column_index = self.run_widget.control_channel_2_selector.currentIndex()
        if datatype_index is None:
            datatype_index = self.run_widget.data_type_selector.currentIndex()
        self.plot_windows.append(
            PlotWindow(
                None,
                row_index,
                column_index,
                datatype_index,
                (self.specification_frequency_lines, self.specification_cpsd_matrix),
                self.run_widget.control_channel_1_selector.itemText(row_index),
                self.run_widget.control_channel_2_selector.itemText(column_index),
                self.run_widget.data_type_selector.itemText(datatype_index),
                (
                    self.specification_warning_matrix
                    if row_index == column_index and datatype_index == 0
                    else None
                ),
                (
                    self.specification_abort_matrix
                    if row_index == column_index and datatype_index == 0
                    else None
                ),
            )
        )

    def show_all_asds(self):
        """Creates a subwindow for each ASD in the CPSD matrix"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            self.create_window(None, i, i, 0)
        self.tile_windows()

    def show_all_csds_phscoh(self):
        """Creates a subwindow for each entry in the CPSD matrix showing phase and coherence"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            for j in range(self.specification_cpsd_matrix.shape[-1]):
                if i == j:
                    datatype_index = 0
                elif i < j:
                    datatype_index = 1
                elif i > j:
                    datatype_index = 2
                else:
                    raise ValueError("Invalid situation.  How did you get here?!")
                self.create_window(None, i, j, datatype_index)
        self.tile_windows()

    def show_all_csds_realimag(self):
        """Creates a subwindow for each entry in the CPSD matrix showing real and imaginary"""
        for i in range(self.specification_cpsd_matrix.shape[-1]):
            for j in range(self.specification_cpsd_matrix.shape[-1]):
                if i == j:
                    datatype_index = 0
                elif i < j:
                    datatype_index = 3
                elif i > j:
                    datatype_index = 4
                else:
                    raise ValueError("Invalid situation.  How did you get here?!")
                self.create_window(None, i, j, datatype_index)
        self.tile_windows()

    def tile_windows(self):
        """Tile subwindow equally across the screen"""
        screen_rect = QtWidgets.QApplication.desktop().screenGeometry()
        # Go through and remove any closed windows
        self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
        num_windows = len(self.plot_windows)
        ncols = int(np.ceil(np.sqrt(num_windows)))
        nrows = int(np.ceil(num_windows / ncols))
        window_width = int(screen_rect.width() / ncols)
        window_height = int(screen_rect.height() / nrows)
        for index, window in enumerate(self.plot_windows):
            window.resize(window_width, window_height)
            row_ind = index // ncols
            col_ind = index % ncols
            window.move(col_ind * window_width, row_ind * window_height)

    def close_windows(self):
        """Close all subwindows"""
        for window in self.plot_windows:
            window.close()

    def save_control_data_from_profile(self, filename):
        """Saves the control data to a file when requested by a profile argument"""
        self.save_spectral_data(None, filename)

    def save_spectral_data(self, clicked, filename=None):  # pylint: disable=unused-argument
        """Save Spectral Data from the Controller"""
        if filename is None:
            filename, _ = QtWidgets.QFileDialog.getSaveFileName(
                self.definition_widget,
                "Select File to Save Spectral Data",
                filter="NetCDF File (*.nc4)",
            )
        if filename == "":
            return
        labels = [
            ["node_number", str],
            ["node_direction", str],
            ["comment", str],
            ["serial_number", str],
            ["triax_dof", str],
            ["sensitivity", str],
            ["unit", str],
            ["make", str],
            ["model", str],
            ["expiration", str],
            ["physical_device", str],
            ["physical_channel", str],
            ["channel_type", str],
            ["minimum_value", str],
            ["maximum_value", str],
            ["coupling", str],
            ["excitation_source", str],
            ["excitation", str],
            ["feedback_device", str],
            ["feedback_channel", str],
            ["warning_level", str],
            ["abort_level", str],
        ]
        global_data_parameters: DataAcquisitionParameters
        global_data_parameters = self.data_acquisition_parameters
        netcdf_handle = nc4.Dataset(  # pylint: disable=no-member
            filename, "w", format="NETCDF4", clobber=True
        )
        # Create dimensions
        netcdf_handle.createDimension("response_channels", len(global_data_parameters.channel_list))
        netcdf_handle.createDimension(
            "output_channels",
            len(
                [
                    channel
                    for channel in global_data_parameters.channel_list
                    if channel.feedback_device is not None
                ]
            ),
        )
        netcdf_handle.createDimension("time_samples", None)
        netcdf_handle.createDimension(
            "num_environments", len(global_data_parameters.environment_names)
        )
        # Create attributes
        netcdf_handle.file_version = "3.0.0"
        netcdf_handle.sample_rate = global_data_parameters.sample_rate
        netcdf_handle.time_per_write = (
            global_data_parameters.samples_per_write / global_data_parameters.output_sample_rate
        )
        netcdf_handle.time_per_read = (
            global_data_parameters.samples_per_read / global_data_parameters.sample_rate
        )
        netcdf_handle.hardware = global_data_parameters.hardware
        netcdf_handle.hardware_file = (
            "None"
            if global_data_parameters.hardware_file is None
            else global_data_parameters.hardware_file
        )
        netcdf_handle.output_oversample = global_data_parameters.output_oversample
        for key, value in global_data_parameters.extra_parameters.items():
            setattr(netcdf_handle, key, value)
        # Create Variables
        var = netcdf_handle.createVariable("environment_names", str, ("num_environments",))
        this_environment_index = None
        for i, name in enumerate(global_data_parameters.environment_names):
            var[i] = name
            if name == self.environment_name:
                this_environment_index = i
        var = netcdf_handle.createVariable(
            "environment_active_channels",
            "i1",
            ("response_channels", "num_environments"),
        )
        var[...] = global_data_parameters.environment_active_channels.astype("int8")[
            global_data_parameters.environment_active_channels[:, this_environment_index],
            :,
        ]
        # Create channel table variables
        for label, netcdf_datatype in labels:
            var = netcdf_handle.createVariable(
                "/channels/" + label, netcdf_datatype, ("response_channels",)
            )
            channel_data = [
                getattr(channel, label) for channel in global_data_parameters.channel_list
            ]
            if netcdf_datatype == "i1":
                channel_data = np.array([1 if val else 0 for val in channel_data])
            else:
                channel_data = ["" if val is None else val for val in channel_data]
            for i, cd in enumerate(channel_data):
                var[i] = cd
        group_handle = netcdf_handle.createGroup(self.environment_name)
        self.environment_parameters.store_to_netcdf(group_handle)
        # Create Variables for Spectral Data
        group_handle.createDimension("drive_channels", self.last_transfer_function.shape[2])
        var = group_handle.createVariable(
            "frf_data_real",
            "f8",
            ("fft_lines", "specification_channels", "drive_channels"),
        )
        var[...] = self.last_transfer_function.real
        var = group_handle.createVariable(
            "frf_data_imag",
            "f8",
            ("fft_lines", "specification_channels", "drive_channels"),
        )
        var[...] = self.last_transfer_function.imag
        var = group_handle.createVariable(
            "frf_coherence", "f8", ("fft_lines", "specification_channels")
        )
        var[...] = self.last_coherence.real
        var = group_handle.createVariable(
            "response_cpsd_real",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.last_response_cpsd.real
        var = group_handle.createVariable(
            "response_cpsd_imag",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.last_response_cpsd.imag
        var = group_handle.createVariable(
            "drive_cpsd_real", "f8", ("fft_lines", "drive_channels", "drive_channels")
        )
        var[...] = self.last_reference_cpsd.real
        var = group_handle.createVariable(
            "drive_cpsd_imag", "f8", ("fft_lines", "drive_channels", "drive_channels")
        )
        var[...] = self.last_reference_cpsd.imag
        var = group_handle.createVariable(
            "response_noise_cpsd_real",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.last_response_noise.real
        var = group_handle.createVariable(
            "response_noise_cpsd_imag",
            "f8",
            ("fft_lines", "specification_channels", "specification_channels"),
        )
        var[...] = self.last_response_noise.imag
        var = group_handle.createVariable(
            "drive_noise_cpsd_real",
            "f8",
            ("fft_lines", "drive_channels", "drive_channels"),
        )
        var[...] = self.last_reference_noise.real
        var = group_handle.createVariable(
            "drive_noise_cpsd_imag",
            "f8",
            ("fft_lines", "drive_channels", "drive_channels"),
        )
        var[...] = self.last_reference_noise.imag
        netcdf_handle.close()

    # %% Miscellaneous

    def retrieve_metadata(
        self,
        netcdf_handle: nc4._netCDF4.Dataset = None,  # pylint: disable=c-extension-no-member
        environment_name: str = None,
    ):
        """Collects environment parameters from a netCDF dataset.

        This function retrieves parameters from a netCDF dataset that was written
        by the controller during streaming.  It must populate the widgets
        in the user interface with the proper information.

        This function is the "read" counterpart to the store_to_netcdf
        function in the AbstractMetadata class, which will write parameters to
        the netCDF file to document the metadata.

        Note that the entire dataset is passed to this function, so the function
        should collect parameters pertaining to the environment from a Group
        in the dataset sharing the environment's name, e.g.

        ``group = netcdf_handle.groups[self.environment_name]``
        ``self.definition_widget.parameter_selector.setValue(group.parameter)``

        Parameters
        ----------
        netcdf_handle : nc4._netCDF4.Dataset
            The netCDF dataset from which the data will be read.  It should have
            a group name with the enviroment's name.
        environment_name : str (optional)
            name of environment from which to retrieve metadata. Only needed if
            different from current environment.

        """
        group = super().retrieve_metadata(netcdf_handle, environment_name)

        # Control channels
        try:
            for i in group.variables["control_channel_indices"][...]:
                item = self.definition_widget.control_channels_selector.item(i)
                item.setCheckState(Qt.Checked)
        except KeyError:
            print("no variable control_channel_indices, please select control channels manually")
        # Other data
        try:
            self.response_transformation_matrix = group.variables["response_transformation_matrix"][
                ...
            ].data
        except KeyError:
            self.response_transformation_matrix = None
        try:
            self.output_transformation_matrix = group.variables["reference_transformation_matrix"][
                ...
            ].data
        except KeyError:
            self.output_transformation_matrix = None
        self.define_transformation_matrices(None, dialog=False)

        # environment_name is passed when the saved environment doesn't match the
        # current environment
        if environment_name is None:
            # Spinboxes
            self.definition_widget.samples_per_frame_selector.setValue(group.samples_per_frame)
            self.definition_widget.ramp_time_spinbox.setValue(group.test_level_ramp_time)
            self.definition_widget.cola_overlap_percentage_selector.setValue(
                group.cola_overlap * 100
            )
            self.definition_widget.cola_exponent_selector.setValue(group.cola_window_exponent)
            self.definition_widget.cpsd_overlap_selector.setValue(group.cpsd_overlap * 100)
            self.definition_widget.cpsd_frames_selector.setValue(group.frames_in_cpsd)
            # Checkboxes
            self.definition_widget.update_transfer_function_during_control_selector.setChecked(
                bool(group.update_tf_during_control)
            )
            self.definition_widget.auto_abort_checkbox.setChecked(
                bool(group.allow_automatic_aborts)
            )
            # Comboboxes
            self.definition_widget.cola_window_selector.setCurrentIndex(
                self.definition_widget.cola_window_selector.findText(group.cola_window)
            )
            self.definition_widget.cpsd_computation_window_selector.setCurrentIndex(
                self.definition_widget.cpsd_computation_window_selector.findText(group.cpsd_window)
            )
            # Specification
            self.specification_frequency_lines = group.variables["specification_frequency_lines"][
                ...
            ].data
            self.specification_cpsd_matrix = (
                group.variables["specification_cpsd_matrix_real"][...].data
                + 1j * group.variables["specification_cpsd_matrix_imag"][...].data
            )
            self.specification_warning_matrix = group.variables["specification_warning_matrix"][
                ...
            ].data
            self.specification_abort_matrix = group.variables["specification_abort_matrix"][
                ...
            ].data
            self.select_python_module(None, group.control_python_script)
            index = self.definition_widget.control_function_input.findText(
                group.control_python_function
            )
            if (
                index == -1
            ):  # error handling (older revisions of rattlesnake may be missing newer control laws)
                index = 0
                default = self.definition_widget.control_function_input.itemText(index)
                print(
                    f'Warning: control function "{group.control_python_function}" not found, '
                    f'defaulting to "{default}"'
                )
            self.definition_widget.control_function_input.setCurrentIndex(index)
            self.definition_widget.control_parameters_text_input.setText(
                group.control_python_function_parameters
            )
            self.show_specification()

    def update_gui(self, queue_data: tuple):
        """Update the environment's graphical user interface

        This function will receive data from the gui_update_queue that
        specifies how the user interface should be updated.  Data will usually
        be received as ``(instruction,data)`` pairs, where the ``instruction`` notes
        what operation should be taken or which widget should be modified, and
        the ``data`` notes what data should be used in the update.

        Parameters
        ----------
        queue_data : tuple
            A tuple containing ``(instruction,data)`` pairs where ``instruction``
            defines and operation or widget to be modified and ``data`` contains
            the data used to perform the operation.
        """
        if super().update_gui(queue_data):
            return
        message, data = queue_data
        if message == "control_predictions":
            (
                _,
                self.excitation_prediction,
                self.response_prediction,
                _,
                rms_voltage_prediction,
                rms_db_error_prediction,
            ) = data
            self.update_control_predictions()
            for widget, widget_data in zip(
                [
                    self.prediction_widget.excitation_voltage_list,
                    self.prediction_widget.response_error_list,
                ],
                [rms_voltage_prediction, rms_db_error_prediction],
            ):
                widget.clear()
                widget.addItems([f"{d:.3f}" for d in widget_data])
            # Now compute if any channels are erroring or not
            with np.errstate(invalid="ignore"):
                lines_out = (
                    self.environment_parameters.percent_lines_out / 100
                ) * self.environment_parameters.fft_lines
                for i in range(self.prediction_widget.response_error_list.count()):
                    item = self.prediction_widget.response_error_list.item(i)
                    if (
                        sum(
                            self.response_prediction[:, i, i]
                            > self.environment_parameters.specification_abort_matrix[1, :, i]
                        )
                        > lines_out
                    ):
                        item.setBackground(QColor(255, 125, 125))
                    elif (
                        sum(
                            self.response_prediction[:, i, i]
                            < self.environment_parameters.specification_abort_matrix[0, :, i]
                        )
                        > lines_out
                    ):
                        item.setBackground(QColor(255, 125, 125))
                    elif (
                        sum(
                            self.response_prediction[:, i, i]
                            > self.environment_parameters.specification_warning_matrix[1, :, i]
                        )
                        > lines_out
                    ):
                        item.setBackground(QColor(255, 255, 125))
                    elif (
                        sum(
                            self.response_prediction[:, i, i]
                            < self.environment_parameters.specification_warning_matrix[0, :, i]
                        )
                        > lines_out
                    ):
                        item.setBackground(QColor(255, 255, 125))
                    else:
                        item.setBackground(QColor(255, 255, 255))
        elif message == "control_update":
            (
                frames,
                total_frames,
                self.frequencies,
                self.last_transfer_function,
                self.last_coherence,
                self.last_response_cpsd,
                self.last_reference_cpsd,
                self.last_condition,
            ) = data
            self.update_sysid_plots(
                update_time=False, update_transfer_function=True, update_noise=True
            )
            self.system_id_widget.current_frames_spinbox.setValue(frames)
            self.system_id_widget.total_frames_spinbox.setValue(total_frames)
            self.system_id_widget.progressBar.setValue(int(frames / total_frames * 100))
            self.plot_data_items["sum_asds_control"].setData(
                self.frequencies, np.einsum("ijj", self.last_response_cpsd).real
            )
            # Go through and remove any closed windows
            self.plot_windows = [window for window in self.plot_windows if window.isVisible()]
            for window in self.plot_windows:
                window.update_plot(self.last_response_cpsd)
        elif message == "interactive_control_sysid_update":
            if self.interactive_control_law_widget is not None:
                self.interactive_control_law_widget.update_ui_sysid(*data)
        elif message == "interactive_control_update":
            if self.interactive_control_law_widget is not None:
                self.interactive_control_law_widget.update_ui_control(data)
        elif message == "update_test_response_error_list":
            rms_db_error, warning_channels, abort_channels = data
            self.run_widget.test_response_error_list.clear()
            self.run_widget.test_response_error_list.addItems([f"{d:.3f}" for d in rms_db_error])
            for index in warning_channels:
                item = self.run_widget.test_response_error_list.item(index)
                item.setBackground(QColor(255, 255, 125))
            for index in abort_channels:
                item = self.run_widget.test_response_error_list.item(index)
                item.setBackground(QColor(255, 125, 125))
        elif message == "enable_control":
            self.enable_control(True)
        elif message == "enable":
            widget = None
            for parent in [
                self.definition_widget,
                self.system_id_widget,
                self.prediction_widget,
                self.run_widget,
            ]:
                try:
                    widget = getattr(parent, data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError(f"Cannot Enable Widget {data}: not found in UI")
            widget.setEnabled(True)
        elif message == "disable":
            widget = None
            for parent in [
                self.definition_widget,
                self.system_id_widget,
                self.prediction_widget,
                self.run_widget,
            ]:
                try:
                    widget = getattr(parent, data)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError(f"Cannot Disable Widget {data}: not found in UI")
            widget.setEnabled(False)
        else:
            widget = None
            for parent in [
                self.definition_widget,
                self.system_id_widget,
                self.prediction_widget,
                self.run_widget,
            ]:
                try:
                    widget = getattr(parent, message)
                    break
                except AttributeError:
                    continue
            if widget is None:
                raise ValueError(f"Cannot Update Widget {message}: not found in UI")
            if isinstance(widget, QtWidgets.QDoubleSpinBox):
                widget.setValue(data)
            elif isinstance(widget, QtWidgets.QSpinBox):
                widget.setValue(data)
            elif isinstance(widget, QtWidgets.QLineEdit):
                widget.setText(data)
            elif isinstance(widget, QtWidgets.QListWidget):
                widget.clear()
                widget.addItems([f"{d:.3f}" for d in data])

    @staticmethod
    def create_environment_template(
        environment_name: str, workbook: openpyxl.workbook.workbook.Workbook
    ):
        """Creates a template worksheet in an Excel workbook defining the
        environment.

        This function creates a template worksheet in an Excel workbook that
        when filled out could be read by the controller to re-create the
        environment.

        This function is the "write" counterpart to the
        ``set_parameters_from_template`` function in the ``RandomVibrationUI`` class,
        which reads the values from the template file to populate the user
        interface.

        Parameters
        ----------
        environment_name : str :
            The name of the environment that will specify the worksheet's name
        workbook : openpyxl.workbook.workbook.Workbook :
            A reference to an ``openpyxl`` workbook.

        """
        worksheet = workbook.create_sheet(environment_name)
        worksheet.cell(1, 1, "Control Type")
        worksheet.cell(1, 2, "Random")
        worksheet.cell(2, 1, "Samples Per Frame:")
        worksheet.cell(2, 2, "# Number of Samples per Measurement Frame")
        worksheet.cell(3, 1, "Test Level Ramp Time:")
        worksheet.cell(3, 2, "# Time taken to Ramp between test levels")
        worksheet.cell(4, 1, "COLA Window:")
        worksheet.cell(4, 2, "# Window used for Constant Overlap and Add process")
        worksheet.cell(5, 1, "COLA Overlap %:")
        worksheet.cell(5, 2, "# Overlap used in Constant Overlap and Add process")
        worksheet.cell(6, 1, "COLA Window Exponent:")
        worksheet.cell(
            6,
            2,
            "# Exponent Applied to the COLA Window (use 0.5 unless you "
            "are sure you don't want to!)",
        )
        worksheet.cell(7, 1, "Update System ID During Control:")
        worksheet.cell(
            7,
            2,
            "# Continue updating transfer function while the controller is controlling (Y/N)",
        )
        worksheet.cell(8, 1, "Frames in CPSD:")
        worksheet.cell(8, 2, "# Frames used to compute the CPSD matrix")
        worksheet.cell(9, 1, "CPSD Window:")
        worksheet.cell(9, 2, "# Window used to compute the CPSD matrix")
        worksheet.cell(10, 1, "CPSD Overlap %:")
        worksheet.cell(10, 2, "# Overlap percentage for CPSD calculations")
        worksheet.cell(11, 1, "Allow Automatic Aborts")
        worksheet.cell(12, 1, "Control Python Script:")
        worksheet.cell(12, 2, "# Path to the Python script containing the control law")
        worksheet.cell(13, 1, "Control Python Function:")
        worksheet.cell(
            13,
            2,
            "# Function or class name within the Python Script that will serve as the control law",
        )
        worksheet.cell(14, 1, "Control Parameters:")
        worksheet.cell(14, 2, "# Extra parameters used in the control law")
        worksheet.cell(15, 1, "Control Channels (1-based):")
        worksheet.cell(16, 1, "System ID Averaging:")
        worksheet.cell(
            16,
            2,
            "# Averaging Type used for system ID.  Should be Linear or Exponential",
        )
        worksheet.cell(17, 1, "Noise Averages:")
        worksheet.cell(17, 2, "# Number of Averages used when characterizing noise")
        worksheet.cell(18, 1, "System ID Averages:")
        worksheet.cell(18, 2, "# Number of Averages used when computing the FRF")
        worksheet.cell(19, 1, "Exponential Averaging Coefficient:")
        worksheet.cell(19, 2, "# Averaging Coefficient for Exponential Averaging (if used)")
        worksheet.cell(20, 1, "System ID Estimator:")
        worksheet.cell(
            20,
            2,
            "# Technique used to compute system ID.  Should be one of H1, H2, H3, or Hv.",
        )
        worksheet.cell(21, 1, "System ID Level (V RMS):")
        worksheet.cell(
            21,
            2,
            "# RMS Value of Flat Voltage Spectrum used for System Identification.",
        )
        worksheet.cell(22, 1, "System ID Signal Type:")
        worksheet.cell(23, 1, "System ID Window:")
        worksheet.cell(
            23,
            2,
            "# Window used to compute FRFs during system ID.  Should be one of Hann or None",
        )
        worksheet.cell(24, 1, "System ID Overlap %:")
        worksheet.cell(24, 2, "# Overlap to use in the system identification")
        worksheet.cell(25, 1, "System ID Burst On %:")
        worksheet.cell(25, 2, "# Percentage of a frame that the burst random is on for")
        worksheet.cell(26, 1, "System ID Burst Pretrigger %:")
        worksheet.cell(
            26,
            2,
            "# Percentage of a frame that occurs before the burst starts in a burst random signal",
        )
        worksheet.cell(27, 1, "System ID Ramp Fraction %:")
        worksheet.cell(
            27,
            2,
            '# Percentage of the "System ID Burst On %" that will be used to ramp up to full level',
        )
        worksheet.cell(28, 1, "Specification File:")
        worksheet.cell(28, 2, "# Path to the file containing the Specification")
        worksheet.cell(29, 1, "Response Transformation Matrix:")
        worksheet.cell(
            29,
            2,
            "# Transformation matrix to apply to the response channels.  Type None if there "
            "is none.  Otherwise, make this a 2D array in the spreadsheet and move the Output "
            "Transformation Matrix line down so it will fit.  The number of columns should be the "
            "number of physical control channels.",
        )
        worksheet.cell(30, 1, "Output Transformation Matrix:")
        worksheet.cell(
            30,
            2,
            "# Transformation matrix to apply to the outputs.  Type None if there is none.  "
            "Otherwise, make this a 2D array in the spreadsheet.  The number of columns should be "
            "the number of physical output channels in the environment.",
        )

    def set_parameters_from_template(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        """
        Collects parameters for the user interface from the Excel template file

        This function reads a filled out template worksheet to create an
        environment.  Cells on this worksheet contain parameters needed to
        specify the environment, so this function should read those cells and
        update the UI widgets with those parameters.

        This function is the "read" counterpart to the
        ``create_environment_template`` function in the ``RandomVibrationUI`` class,
        which writes a template file that can be filled out by a user.


        Parameters
        ----------
        worksheet : openpyxl.worksheet.worksheet.Worksheet
            An openpyxl worksheet that contains the environment template.
            Cells on this worksheet should contain the parameters needed for the
            user interface.

        """
        self.definition_widget.samples_per_frame_selector.setValue(int(worksheet.cell(2, 2).value))
        self.definition_widget.ramp_time_spinbox.setValue(float(worksheet.cell(3, 2).value))
        self.definition_widget.cola_window_selector.setCurrentIndex(
            self.definition_widget.cola_window_selector.findText(worksheet.cell(4, 2).value)
        )
        self.definition_widget.cola_overlap_percentage_selector.setValue(
            float(worksheet.cell(5, 2).value)
        )
        self.definition_widget.cola_exponent_selector.setValue(float(worksheet.cell(6, 2).value))
        self.definition_widget.update_transfer_function_during_control_selector.setChecked(
            worksheet.cell(7, 2).value.upper() == "Y"
        )
        self.definition_widget.cpsd_frames_selector.setValue(int(worksheet.cell(8, 2).value))
        self.definition_widget.cpsd_computation_window_selector.setCurrentIndex(
            self.definition_widget.cpsd_computation_window_selector.findText(
                worksheet.cell(9, 2).value
            )
        )
        self.definition_widget.cpsd_overlap_selector.setValue(float(worksheet.cell(10, 2).value))
        self.definition_widget.auto_abort_checkbox.setChecked(
            worksheet.cell(11, 2).value.upper() == "Y"
        )
        self.select_python_module(None, worksheet.cell(12, 2).value)
        self.definition_widget.control_function_input.setCurrentIndex(
            self.definition_widget.control_function_input.findText(worksheet.cell(13, 2).value)
        )
        self.definition_widget.control_parameters_text_input.setText(
            "" if worksheet.cell(14, 2).value is None else str(worksheet.cell(14, 2).value)
        )
        column_index = 2
        while True:
            value = worksheet.cell(15, column_index).value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                break
            item = self.definition_widget.control_channels_selector.item(int(value) - 1)
            item.setCheckState(Qt.Checked)
            column_index += 1
        self.system_id_widget.averagingTypeComboBox.setCurrentIndex(
            self.system_id_widget.averagingTypeComboBox.findText(worksheet.cell(16, 2).value)
        )
        self.system_id_widget.noiseAveragesSpinBox.setValue(int(worksheet.cell(17, 2).value))
        self.system_id_widget.systemIDAveragesSpinBox.setValue(int(worksheet.cell(18, 2).value))
        self.system_id_widget.averagingCoefficientDoubleSpinBox.setValue(
            float(worksheet.cell(19, 2).value)
        )
        self.system_id_widget.estimatorComboBox.setCurrentIndex(
            self.system_id_widget.estimatorComboBox.findText(worksheet.cell(20, 2).value)
        )
        self.system_id_widget.levelDoubleSpinBox.setValue(float(worksheet.cell(21, 2).value))
        # this should be a temporary solution - template file rework needed
        low, high = worksheet.cell(21, 3).value, worksheet.cell(21, 4).value
        sigma = worksheet.cell(21, 5).value
        if low is not None:
            self.system_id_widget.lowFreqCutoffSpinBox.setValue(int(low))
        if high is not None:
            self.system_id_widget.highFreqCutoffSpinBox.setValue(int(high))
        if sigma is not None:
            self.definition_widget.sigma_clipping_selector.setValue(
                float(sigma)
            )  # TODO: sigma clipping and bandwidths should get
            # their own rows, but how to maintain backward compatibility?
        self.system_id_widget.signalTypeComboBox.setCurrentIndex(
            self.system_id_widget.signalTypeComboBox.findText(worksheet.cell(22, 2).value)
        )
        self.system_id_widget.windowComboBox.setCurrentIndex(
            self.system_id_widget.windowComboBox.findText(worksheet.cell(23, 2).value)
        )
        self.system_id_widget.overlapDoubleSpinBox.setValue(float(worksheet.cell(24, 2).value))
        self.system_id_widget.onFractionDoubleSpinBox.setValue(float(worksheet.cell(25, 2).value))
        self.system_id_widget.pretriggerDoubleSpinBox.setValue(float(worksheet.cell(26, 2).value))
        self.system_id_widget.rampFractionDoubleSpinBox.setValue(float(worksheet.cell(27, 2).value))

        # Now we need to find the transformation matrices' sizes
        response_channels = self.definition_widget.control_channels_display.value()
        output_channels = self.definition_widget.output_channels_display.value()
        output_transform_row = 30
        if (
            isinstance(worksheet.cell(29, 2).value, str)
            and worksheet.cell(29, 2).value.lower() == "none"
        ):
            self.response_transformation_matrix = None
        else:
            while True:
                if worksheet.cell(output_transform_row, 1).value == "Output Transformation Matrix:":
                    break
                output_transform_row += 1
            response_size = output_transform_row - 29
            response_transformation = []
            for i in range(response_size):
                response_transformation.append([])
                for j in range(response_channels):
                    response_transformation[-1].append(float(worksheet.cell(29 + i, 2 + j).value))
            self.response_transformation_matrix = np.array(response_transformation)
        if (
            isinstance(worksheet.cell(output_transform_row, 2).value, str)
            and worksheet.cell(output_transform_row, 2).value.lower() == "none"
        ):
            self.output_transformation_matrix = None
        else:
            output_transformation = []
            i = 0
            while True:
                if worksheet.cell(output_transform_row + i, 2).value is None or (
                    isinstance(worksheet.cell(output_transform_row + i, 2).value, str)
                    and worksheet.cell(output_transform_row + i, 2).value.strip() == ""
                ):
                    break
                output_transformation.append([])
                for j in range(output_channels):
                    output_transformation[-1].append(
                        float(worksheet.cell(output_transform_row + i, 2 + j).value)
                    )
                i += 1
            self.output_transformation_matrix = np.array(output_transformation)
        self.define_transformation_matrices(None, dialog=False)
        self.select_spec_file(None, worksheet.cell(28, 2).value)
