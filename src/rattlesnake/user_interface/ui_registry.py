from rattlesnake.environment.environment_utilities import ControlTypes
from rattlesnake.user_interface.transient_sys_id_ui import TransientUI
from rattlesnake.user_interface.random_vibration_sys_id_ui import RandomVibrationUI
from rattlesnake.user_interface.sine_sys_id_ui import SineUI
from rattlesnake.user_interface.time_ui import TimeUI
from rattlesnake.user_interface.modal_ui import ModalUI
import openpyxl
from qtpy import QtWidgets, uic, QtGui
from rattlesnake.user_interface.ui_utilities import environment_select_ui_path
from rattlesnake.environment.environment_utilities import combined_environments_capable

ENVIRONMENT_UIS = {}

ENVIRONMENT_UIS[ControlTypes.TIME] = TimeUI
ENVIRONMENT_UIS[ControlTypes.MODAL] = ModalUI
ENVIRONMENT_UIS[ControlTypes.SINE] = SineUI
ENVIRONMENT_UIS[ControlTypes.RANDOM] = RandomVibrationUI
ENVIRONMENT_UIS[ControlTypes.TRANSIENT] = TransientUI


def save_combined_environments_profile_template(filename, environment_data):
    """Creates a spreadsheet template that can be completed and loaded to define a test"""
    # Create the header
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Channel Table"
    hardware_worksheet = workbook.create_sheet("Hardware")
    # Create the header
    worksheet.cell(row=1, column=2, value="Test Article Definition")
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    worksheet.cell(row=1, column=5, value="Instrument Definition")
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=11)
    worksheet.cell(row=1, column=12, value="Channel Definition")
    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=19)
    worksheet.cell(row=1, column=20, value="Output Feedback")
    worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=21)
    worksheet.cell(row=1, column=22, value="Limits")
    worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=23)
    for col_idx, val in enumerate(
        [
            "Channel Index",
            "Node Number",
            "Node Direction",
            "Comment",
            "Serial Number",
            "Triax DoF",
            "Sensitivity  (mV/EU)",
            "Engineering Unit",
            "Make",
            "Model",
            "Calibration Exp Date",
            "Physical Device",
            "Physical Channel",
            "Type",
            "Minimum Value (V)",
            "Maximum Value (V)",
            "Coupling",
            "Current Excitation Source",
            "Current Excitation Value",
            "Physical Device",
            "Physical Channel",
            "Warning Level (EU)",
            "Abort Level (EU)",
        ]
    ):
        worksheet.cell(row=2, column=1 + col_idx, value=val)
    # Fill out the hardware worksheet
    hardware_worksheet.cell(1, 1, "Hardware Type")
    hardware_worksheet.cell(1, 2, "# Enter hardware index here")
    hardware_worksheet.cell(
        1,
        3,
        "Hardware Indices: 0 - NI DAQmx; 1 - LAN XI; 2 - Data Physics Quattro; "
        "3 - Data Physics 900 Series; 4 - Exodus Modal Solution; 5 - State Space Integration; "
        "6 - SDynPy System Integration",
    )
    hardware_worksheet.cell(2, 1, "Hardware File")
    hardware_worksheet.cell(
        2,
        2,
        "# Path to Hardware File (Depending on Hardware Device: 0 - Not Used; 1 - Not Used; "
        "2 - Path to DpQuattro.dll library file; 3 - Not Used; 4 - Path to Exodus Eigensolution; "
        "5 - Path to State Space File; 6 - Path to SDynPy system file)",
    )
    hardware_worksheet.cell(3, 1, "Sample Rate")
    hardware_worksheet.cell(3, 2, "# Sample Rate of Data Acquisition System")
    hardware_worksheet.cell(4, 1, "Time Per Read")
    hardware_worksheet.cell(4, 2, "# Number of seconds per Read from the Data Acquisition System")
    hardware_worksheet.cell(5, 1, "Time Per Write")
    hardware_worksheet.cell(5, 2, "# Number of seconds per Write to the Data Acquisition System")
    hardware_worksheet.cell(6, 1, "Maximum Acquisition Processes")
    hardware_worksheet.cell(
        6,
        2,
        "# Maximum Number of Acquisition Processes to start to pull data from hardware",
    )
    hardware_worksheet.cell(
        6,
        3,
        "Only Used by LAN-XI Hardware.  This row can be deleted if LAN-XI is not used",
    )
    hardware_worksheet.cell(7, 1, "Integration Oversampling")
    hardware_worksheet.cell(
        7, 2, "# For virtual control, an integration oversampling can be specified"
    )
    hardware_worksheet.cell(
        7,
        3,
        "Only used for virtual control (Exodus, State Space, or SDynPy).  "
        "This row can be deleted if these are not used.",
    )
    hardware_worksheet.cell(8, 1, "Task Trigger")
    hardware_worksheet.cell(8, 2, "# Start trigger type")
    hardware_worksheet.cell(
        8,
        3,
        "Task Triggers: 0 - Internal, 1 - PFI0 with external trigger, 2 - PFI0 with Analog Output "
        "trigger.  Only used for NI hardware.  This row can be deleted if NI is not used.",
    )
    hardware_worksheet.cell(9, 1, "Task Trigger Output Channel")
    hardware_worksheet.cell(9, 2, "# Physical device and channel that generates a trigger signal")
    hardware_worksheet.cell(
        9,
        3,
        "Only used if Task Triggers is 2.  Only used for NI hardware.  "
        "This row can be deleted if it is not used.",
    )

    # Now do the environment
    worksheet.cell(row=1, column=24, value="Environments")
    for row, (value, name) in enumerate(environment_data):
        ENVIRONMENT_UIS[value].create_environment_template(name, workbook)
        worksheet.cell(row=2, column=24 + row, value=name)
    # Now create a profile page
    profile_sheet = workbook.create_sheet("Test Profile")
    profile_sheet.cell(1, 1, "Time (s)")
    profile_sheet.cell(1, 2, "Environment")
    profile_sheet.cell(1, 3, "Operation")
    profile_sheet.cell(1, 4, "Data")

    workbook.save(filename)


class EnvironmentSelect(QtWidgets.QDialog):
    """QDialog for selecting the environments in a combined environments run"""

    def __init__(self, parent=None):
        """
        Constructor for the EnvironmentSelect dialog box.

        Parameters
        ----------
        parent : QWidget, optional
            Parent widget to the dialog. The default is None.

        """
        super(QtWidgets.QDialog, self).__init__(parent)
        uic.loadUi(environment_select_ui_path, self)
        self.setWindowIcon(QtGui.QIcon("logo/Rattlesnake_Icon.png"))

        self.add_environment_button.clicked.connect(self.add_environment)
        self.remove_environment_button.clicked.connect(self.remove_environment)
        self.load_profile_button.clicked.connect(self.load_profile)
        self.save_profile_template_button.clicked.connect(self.save_profile_template)
        self.loaded_profile = None

    def add_environment(self):
        """Adds a row to the environment table"""
        selected_row = self.environment_display_table.rowCount()
        self.environment_display_table.insertRow(selected_row)
        combobox = QtWidgets.QComboBox()
        for control_type in combined_environments_capable:
            combobox.addItem(control_type.name.title(), control_type.value)
        self.environment_display_table.setCellWidget(selected_row, 0, combobox)

    def remove_environment(self):
        """Removes a row from the environment table"""
        selected_row = self.environment_display_table.currentRow()
        if selected_row >= 0:
            self.environment_display_table.removeRow(selected_row)

    def save_profile_template(self):
        """Saves a template for the given environments table

        This template can be filled out by a user and then loaded."""
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save Combined Environments Template", filter="Excel File (*.xlsx)"
        )
        if filename == "":
            return
        # Now do the environments
        environment_data = []
        for row in range(self.environment_display_table.rowCount()):
            combobox = self.environment_display_table.cellWidget(row, 0)
            value = ControlTypes(combobox.currentData())
            name = self.environment_display_table.item(row, 1).text()
            environment_data.append((value, name))
        save_combined_environments_profile_template(filename, environment_data)

    def load_profile(self):
        """Loads a profile from an excel spreadsheet."""
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Load Combined Environments Profile", filter="Excel File (*.xlsx)"
        )
        if filename == "":
            return
        else:
            self.loaded_profile = filename
            self.accept()

    @staticmethod
    def select_environment(parent=None):
        """Creates the dialog box and then parses the output.

        Note that there are variable numbers of outputs for this function

        Parameters
        ----------
        parent : QWidget
            Parent to the dialog box (Default value = None)

        Returns
        -------
        result : int
            A flag specifying the outcome of the dialog box.  Will be 1 if the
            dialog was accepted, zero if cancelled, and -1 if a profile was
            loaded instead.
        environment_table : list of lists
            A list of environment type, environment name pairs that will be
            used to specify the environments in a test.
        loaded_profile : str
            File name to the profile file that needs to be loaded.  Only
            output if result == -1
        """
        dialog = EnvironmentSelect(parent)
        result = 1 if (dialog.exec_() == QtWidgets.QDialog.Accepted) else 0
        if dialog.loaded_profile is None:
            environment_table = []
            if result:
                for row in range(dialog.environment_display_table.rowCount()):
                    combobox = dialog.environment_display_table.cellWidget(row, 0)
                    value = ControlTypes(combobox.currentData())
                    name = dialog.environment_display_table.item(row, 1).text()
                    environment_table.append([value, name])
            # print(environment_table)
            return result, environment_table
        else:
            result = -1
            workbook = openpyxl.load_workbook(dialog.loaded_profile)
            environment_sheets = [
                sheet
                for sheet in workbook
                if (sheet.title not in ["Channel Table", "Hardware", "Test Profile"])
                and sheet.cell(1, 1).value == "Control Type"
            ]
            environment_table = [
                (ControlTypes[sheet.cell(1, 2).value.upper()], sheet.title)
                for sheet in environment_sheets
            ]
            workbook.close()
            return result, environment_table, dialog.loaded_profile
